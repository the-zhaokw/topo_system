# 缺陷管理 API - 已启用，支持完整的 CRUD 操作

import os
import re
import logging
import uuid
import pandas as pd
from io import BytesIO
from flask import Blueprint, request, jsonify, send_file, make_response
from datetime import datetime
from flask_jwt_extended import jwt_required, get_jwt_identity
from logging_decorators import log_api_call, log_business_operation
from functools import wraps
from enhanced_app import db as _app_db, User as _AppUser


def _check_perm(user, perm_code):
    if not user:
        return False
    if user.is_super_admin:
        return True
    pos = user.get_position_info()
    if pos and (pos.is_admin or pos.is_manager):
        return True
    return user.check_permission(perm_code)


def require_bug_permission(perm_code):
    """缺陷子路由权限校验装饰器"""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            current_user_id = get_jwt_identity()
            user = _app_db.session.query(_AppUser).get(current_user_id)
            if not user:
                return jsonify({'error': '用户不存在'}), 404
            if not _check_perm(user, perm_code):
                return jsonify({'error': '权限不足', 'code': 'PERMISSION_DENIED', 'required_permission': perm_code}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator


from werkzeug.utils import secure_filename
from sqlalchemy.orm import joinedload

# 延迟导入数据库和装饰器，模型延迟导入以避免循环导入
def get_db():
    from enhanced_app import get_db_instance
    return get_db_instance()

def get_create_audit_log():
    from enhanced_app import create_audit_log
    return create_audit_log

def get_create_notification():
    """获取通知创建函数"""
    from enhanced_app import create_notification
    return create_notification

def get_models():
    """延迟获取数据库模型"""
    from enhanced_app import User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications
    return User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications

def parse_mentions(content):
    """解析评论内容中的@提及用户，返回用户名列表"""
    # 匹配 @username 格式的用户名
    mention_pattern = r'@([a-zA-Z0-9_-]+)'
    mentions = re.findall(mention_pattern, content)
    return list(set(mentions))  # 去重

def parse_tags(tags_value):
    """解析tags字段，返回逗号分隔的字符串
    支持的输入格式：
    - 数组：['数据库', '严重']
    - JSON字符串：'["数据库", "严重"]'
    - 逗号分隔字符串：'数据库,严重'
    - 普通字符串：'数据库'
    返回：逗号分隔的字符串，如 '数据库,严重'
    """
    if not tags_value:
        return None
    
    if isinstance(tags_value, list):
        # 直接是数组
        tags_list = tags_value
    elif isinstance(tags_value, str):
        # 是字符串，尝试解析
        tags_value = tags_value.strip()
        if not tags_value:
            return None
        
        # 尝试解析JSON数组
        if tags_value.startswith('['):
            try:
                import json
                parsed = json.loads(tags_value)
                if isinstance(parsed, list):
                    tags_list = parsed
                else:
                    tags_list = [str(parsed)]
            except:
                # JSON解析失败，作为逗号分隔处理
                tags_list = [t.strip() for t in tags_value.split(',') if t.strip()]
        else:
            # 逗号分隔的字符串
            tags_list = [t.strip() for t in tags_value.split(',') if t.strip()]
    else:
        return None
    
    # 过滤空值并返回逗号分隔的字符串
    valid_tags = [t for t in tags_list if t and str(t).strip()]
    return ','.join(valid_tags) if valid_tags else None

# 创建bugs蓝图
bugs_bp = Blueprint('bugs', __name__, url_prefix='/bugs')

# 获取缺陷列表
@bugs_bp.route('/', methods=['GET'])
@jwt_required()
def get_bugs():
    """获取缺陷列表"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    user_role = current_user.role if current_user else None
    
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status')
    severity = request.args.get('severity')
    priority = request.args.get('priority')
    assignee = request.args.get('assignee', type=int)
    resolver = request.args.get('resolver', type=int)
    assigned_to_me = request.args.get('assigned_to_me', type=int)
    search = request.args.get('search')
    user_id = request.args.get('user_id', type=int)
    version = request.args.get('version')
    module = request.args.get('module')
    issue_type = request.args.get('issue_type')
    reproduce_frequency = request.args.get('reproduce_frequency')
    found_build = request.args.get('found_build')
    test_version = request.args.get('test_version')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    sort = request.args.get('sort', 'created_at')
    order = request.args.get('order', 'desc')
    filter_type = request.args.get('filter_type')
    
    query = Bug.query
    
    if filter_type == 'my_handling':
        if user_role in ['tester', 'test_engineer']:
            open_statuses = [BugStatus.NEW.value, BugStatus.ASSIGNED.value, BugStatus.IN_PROGRESS.value, BugStatus.FIXED.value, BugStatus.REOPENED.value, BugStatus.RESOLVED.value]
            query = query.filter(
                db.or_(
                    db.and_(
                        Bug.reported_by == current_user_id,
                        Bug.resolved_by.is_(None),
                        Bug.status.in_(open_statuses)
                    ),
                    db.and_(
                        Bug.verifier_id == current_user_id,
                        Bug.status.in_([BugStatus.RESOLVED.value, BugStatus.FIXED.value])
                    )
                )
            )
        else:
            open_statuses = [BugStatus.NEW.value, BugStatus.ASSIGNED.value, BugStatus.IN_PROGRESS.value, BugStatus.FIXED.value, BugStatus.REOPENED.value, BugStatus.RESOLVED.value]
            query = query.filter(
                Bug.resolved_by == current_user_id,
                Bug.status.in_(open_statuses)
            )
    elif filter_type == 'to_claim':
        open_statuses = [BugStatus.NEW.value, BugStatus.ASSIGNED.value, BugStatus.IN_PROGRESS.value, BugStatus.FIXED.value, BugStatus.REOPENED.value]
        query = query.filter(
            Bug.resolved_by.is_(None),
            Bug.verifier_id.is_(None),
            Bug.status.in_(open_statuses)
        )
    elif filter_type == 'all_open':
        open_statuses = [BugStatus.NEW.value, BugStatus.ASSIGNED.value, BugStatus.IN_PROGRESS.value, BugStatus.FIXED.value, BugStatus.REOPENED.value, BugStatus.RESOLVED.value, BugStatus.VERIFIED.value]
        query = query.filter(Bug.status.in_(open_statuses))
    
    if project_id:
        query = query.filter_by(project_id=project_id)
    
    # 如果指定了user_id，筛选该用户相关的bug（创建者、解决者、验证者、分配给）
    if user_id:
        query = query.filter(
            (Bug.reported_by == user_id) |
            (Bug.assigned_to == user_id) |
            (Bug.resolved_by == user_id) |
            (Bug.verifier_id == user_id) |
            (Bug.verified_by == user_id)
        )
    
    # 应用筛选
    if status:
        status_list = status.split(',')
        # 处理特殊的 'open' 状态，表示所有未关闭的bug
        if 'open' in status_list:
            # 排除 closed, resolved, verified 状态
            closed_statuses = [BugStatus.CLOSED.value, BugStatus.RESOLVED.value, BugStatus.VERIFIED.value]
            query = query.filter(~Bug.status.in_(closed_statuses))
            # 从列表中移除 'open'，只保留其他明确的状态
            status_list = [s for s in status_list if s != 'open']
            if status_list:
                query = query.filter(Bug.status.in_(status_list))
        else:
            query = query.filter(Bug.status.in_(status_list))
    if severity:
        query = query.filter_by(severity=severity)
    if priority:
        query = query.filter_by(priority=priority)
    if assignee:
        if isinstance(assignee, int):
            assignee_list = [assignee]
        else:
            assignee_list = [int(a) for a in assignee.split(',') if a.isdigit()]
        query = query.filter(Bug.assigned_to.in_(assignee_list))
    if resolver:
        query = query.filter(Bug.resolved_by == resolver)
    if assigned_to_me:
        open_statuses = [BugStatus.NEW.value, BugStatus.ASSIGNED.value, BugStatus.IN_PROGRESS.value, BugStatus.FIXED.value, BugStatus.REOPENED.value]
        query = query.filter(
            (Bug.assigned_to == assigned_to_me) | (Bug.resolved_by == assigned_to_me),
            Bug.status.in_(open_statuses)
        )
    if search:
        query = query.filter(Bug.title.contains(search) | Bug.description.contains(search))
    if version:
        query = query.filter_by(version=version)
    if module:
        query = query.filter_by(module=module)
    if issue_type:
        query = query.filter_by(issue_type=issue_type)
    if reproduce_frequency:
        query = query.filter_by(reproduce_frequency=reproduce_frequency)
    if found_build:
        query = query.filter_by(found_build=found_build)
    if test_version:
        query = query.filter_by(test_version=test_version)
    
    # 排序
    sort_column = getattr(Bug, sort, Bug.created_at)
    if order == 'asc':
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())
    
    # 分页
    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    bugs = pagination.items
    
    # 构建包含关联信息的bug列表
    bug_list = []
    for bug in bugs:
        bug_dict = bug.to_dict()
        # 添加项目名称
        if bug.project_id:
            project = Project.query.get(bug.project_id)
            bug_dict['project_name'] = project.name if project else '未知项目'
        else:
            bug_dict['project_name'] = '未知项目'
        
        # 添加创建人名称
        if bug.reported_by:
            reporter = User.query.get(bug.reported_by)
            bug_dict['reporter_name'] = f'{reporter.first_name or ""} {reporter.last_name or ""}'.strip() or reporter.username if reporter else '未知'
        else:
            bug_dict['reporter_name'] = '未知'
        
        # 添加指派人名称
        if bug.assigned_to:
            assignee_user = User.query.get(bug.assigned_to)
            bug_dict['assignee_name'] = f'{assignee_user.first_name or ""} {assignee_user.last_name or ""}'.strip() or assignee_user.username if assignee_user else '未知'
        else:
            bug_dict['assignee_name'] = '未分配'
        
        bug_list.append(bug_dict)
    
    return jsonify({
        'success': True,
        'message': '获取缺陷列表成功',
        'data': {
            'bugs': bug_list,
            'total': pagination.total,
            'page': page,
            'per_page': per_page
        }
    })

# 创建缺陷
@bugs_bp.route('/', methods=['POST'])
@jwt_required()
@require_bug_permission('bug:create')
def create_bug():
    """创建缺陷"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    create_audit_log = get_create_audit_log()
    
    current_user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data:
        return jsonify({'success': False, 'message': '请求数据不能为空', 'data': None}), 400
    
    # 验证必填字段
    required_fields = ['title', 'project_id']
    for field in required_fields:
        if field not in data or not data[field]:
            return jsonify({'success': False, 'message': f'{field} 字段为必填项', 'data': None}), 400
    
    # 验证项目是否存在
    project = Project.query.get(data['project_id'])
    if not project:
        return jsonify({'success': False, 'message': '指定的项目不存在', 'data': None}), 404
    
    # 验证当前用户权限（管理员或项目成员）
    current_user = User.query.get(current_user_id)
    if current_user.role not in ['admin', 'manager']:
        member = ProjectMember.query.filter_by(
            project_id=data['project_id'],
            user_id=current_user_id
        ).first()
        if not member:
            return jsonify({'success': False, 'message': '无权在此项目中创建缺陷', 'data': None}), 403
    
    # 处理状态
    status = BugStatus.NEW.value
    if 'status' in data and data['status']:
        status_value = data['status'].lower()
        valid_statuses = [s.value for s in BugStatus]
        if status_value not in valid_statuses:
            return jsonify({'success': False, 'message': f'无效的状态值，可选值: {valid_statuses}', 'data': None}), 400
        status = status_value
    
    # 处理优先级
    priority = Priority.MEDIUM.value
    if 'priority' in data and data['priority']:
        priority_value = data['priority'].lower()
        valid_priorities = [p.value for p in Priority]
        if priority_value not in valid_priorities:
            return jsonify({'success': False, 'message': f'无效的优先级值，可选值: {valid_priorities}', 'data': None}), 400
        priority = priority_value
    
    # 处理严重程度
    severity = Severity.HIGH.value
    if 'severity' in data and data['severity']:
        severity_value = data['severity'].lower()
        valid_severities = [sv.value for sv in Severity]
        if severity_value not in valid_severities:
            return jsonify({'success': False, 'message': f'无效的严重程度值，可选值: {valid_severities}', 'data': None}), 400
        severity = severity_value
    
    # 验证分配人是否存在
    assigned_to = None
    if 'assigned_to' in data and data['assigned_to']:
        assignee = User.query.get(data['assigned_to'])
        if not assignee:
            return jsonify({'success': False, 'message': '指定的分配用户不存在', 'data': None}), 404
        assigned_to = data['assigned_to']

    # 验证验证者是否存在，如果未提供则默认使用创建者
    verifier_id = int(current_user_id)
    if 'verifier_id' in data and data['verifier_id']:
        verifier = User.query.get(data['verifier_id'])
        if not verifier:
            return jsonify({'success': False, 'message': '指定的验证者用户不存在', 'data': None}), 404
        verifier_id = data['verifier_id']

    # 处理截止日期
    deadline = None
    if 'deadline' in data and data['deadline']:
        try:
            deadline = datetime.strptime(data['deadline'], '%Y-%m-%d')
        except ValueError:
            try:
                deadline = datetime.fromisoformat(data['deadline'].replace('Z', '+00:00'))
            except ValueError:
                return jsonify({'success': False, 'message': '截止日期格式不正确', 'data': None}), 400
    
    try:
        bug = Bug(
            title=data['title'],
            description=data.get('description', ''),
            status=status,
            priority=priority,
            severity=severity,
            project_id=data['project_id'],
            reported_by=int(current_user_id),
            assigned_to=assigned_to,
            verifier_id=verifier_id,
            version=data.get('version'),
            tags=parse_tags(data.get('tags')),
            issue_type=data.get('issue_type'),
            reproduce_frequency=data.get('reproduce_frequency'),
            found_build=data.get('found_build'),
            test_version=data.get('test_version'),
            module=data.get('module'),
            reproduce_steps=data.get('reproduce_steps'),
            expected_result=data.get('expected_result'),
            actual_result=data.get('actual_result'),
            steps_to_reproduce=data.get('steps_to_reproduce'),
            customer_mr_number=data.get('customer_mr_number'),
            plan_resolve_version=data.get('plan_resolve_version'),
            estimated_hours=data.get('estimated_hours'),
            deadline=deadline
        )
        
        db.session.add(bug)
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='create_bug',
            resource_type='bug',
            resource_id=bug.id,
            details=f'创建缺陷: {bug.title}',
            request=request
        )
        
        # 如果分配了用户，发送通知（异步发送，不阻塞请求）
        if assigned_to:
            try:
                # 使用线程异步发送通知，避免阻塞请求
                import threading
                def send_notification_async():
                    try:
                        create_notification = get_create_notification()
                        notification = create_notification(
                            user_id=assigned_to,
                            notification_type='bug_assigned',
                            title=f'新缺陷分配给您: {bug.title}',
                            content=f'缺陷 #{bug.id} "{bug.title}" 已分配给您，请及时处理。\n\n项目: {project.name if project else "未知"}\n优先级: {bug.priority.value if hasattr(bug.priority, "value") else str(bug.priority)}\n严重程度: {bug.severity.value if hasattr(bug.severity, "value") else str(bug.severity)}',
                            related_bug_id=bug.id
                        )
                        logger = logging.getLogger(__name__)
                        notification_id = notification.id if notification else None
                        logger.info(f"创建bug时发送通知给分配用户 {assigned_to}, notification_id: {notification_id}")
                    except Exception as e:
                        logger = logging.getLogger(__name__)
                        logger.error(f"创建bug时发送通知失败: {str(e)}")
                
                # 启动异步线程发送通知
                notification_thread = threading.Thread(target=send_notification_async)
                notification_thread.daemon = True
                notification_thread.start()
                
            except Exception as e:
                logger = logging.getLogger(__name__)
                logger.error(f"启动通知线程失败: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': '缺陷创建成功',
            'data': {'bug': bug.to_dict()}
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'创建缺陷失败: {str(e)}', 'data': None}), 500

# 获取缺陷的评论列表
@bugs_bp.route('/<int:bug_id>/comments', methods=['GET'])
@jwt_required()
def get_bug_comments(bug_id):
    """获取缺陷的评论列表"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    comments = Comment.query.filter_by(
        commentable_type='bug',
        commentable_id=bug_id
    ).order_by(Comment.created_at.asc()).all()
    
    comments_data = []
    for comment in comments:
        comment_dict = comment.to_dict()
        comments_data.append(comment_dict)
    
    return jsonify({
        'success': True,
        'message': '获取评论列表成功',
        'data': {'comments': comments_data}
    })

# 创建缺陷评论
@bugs_bp.route('/<int:bug_id>/comments', methods=['POST'])
@jwt_required()
def create_bug_comment(bug_id):
    """创建缺陷评论"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    data = request.get_json()
    if not data or not data.get('content'):
        return jsonify({'success': False, 'message': '评论内容不能为空', 'data': None}), 400
    
    current_user_id = get_jwt_identity()
    
    comment = Comment(
        content=data['content'],
        created_by=int(current_user_id),
        commentable_type='bug',
        commentable_id=bug_id
    )
    
    db = get_db()
    db.session.add(comment)
    db.session.commit()
    
    comment_dict = comment.to_dict()
    
    return jsonify({
        'success': True,
        'message': '评论创建成功',
        'data': {'comment': comment_dict}
    }), 201

# 获取单个缺陷详情
@bugs_bp.route('/<int:bug_id>', methods=['GET'])
@jwt_required()
def get_bug_by_id(bug_id):
    """获取单个缺陷详情"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    bug = Bug.query.get(bug_id)
    
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    bug_dict = bug.to_dict()
    
    # 添加关联信息
    if bug.project:
        bug_dict['project_name'] = bug.project.name
    if bug.assignee:
        bug_dict['assignee_name'] = bug.assignee.username
    if bug.reporter:
        bug_dict['creator_name'] = bug.reporter.username
    if bug.resolver:
        bug_dict['resolver_name'] = bug.resolver.username
    if bug.verifier:
        bug_dict['verifier_name'] = bug.verifier.username
    
    # 将标签字符串转换为数组
    if bug.tags:
        tags_str = bug.tags.strip()
        # 尝试解析JSON数组格式
        if tags_str.startswith('['):
            try:
                import json
                parsed = json.loads(tags_str)
                if isinstance(parsed, list):
                    bug_dict['tags'] = [t.strip() for t in parsed if t.strip()]
                else:
                    bug_dict['tags'] = []
            except:
                # JSON解析失败，作为逗号分隔处理
                bug_dict['tags'] = [t.strip() for t in tags_str.split(',') if t.strip()]
        else:
            # 逗号分隔的字符串
            bug_dict['tags'] = [t.strip() for t in tags_str.split(',') if t.strip()]
    else:
        bug_dict['tags'] = []
    
    # 获取附件列表，使用 joinedload 预加载 uploader 关系
    attachments = Attachment.query.options(
        joinedload(Attachment.uploader)
    ).filter_by(
        attachable_type='bug',
        attachable_id=bug_id
    ).all()
    bug_dict['attachments'] = [attachment.to_dict() for attachment in attachments]
    
    return jsonify({
        'success': True,
        'message': '获取缺陷详情成功',
        'data': {'bug': bug_dict}
    })

# 更新缺陷
@bugs_bp.route('/<int:bug_id>', methods=['PUT'])
@jwt_required()
@require_bug_permission('bug:edit')
def update_bug(bug_id):
    """更新缺陷"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    from enhanced_app import Activity
    
    db = get_db()
    create_audit_log = get_create_audit_log()
    
    current_user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data:
        return jsonify({'success': False, 'message': '请求数据不能为空', 'data': None}), 400
    
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    current_user = User.query.get(current_user_id)
    if current_user.role not in ['admin', 'manager']:
        member = ProjectMember.query.filter_by(
            project_id=bug.project_id,
            user_id=current_user_id
        ).first()
        if not member:
            return jsonify({'success': False, 'message': '无权更新此缺陷', 'data': None}), 403
    
    updateable_fields = [
        'title', 'description', 'priority', 'severity', 'version', 
        'module', 'tags', 'issue_type', 'reproduce_frequency', 
        'found_build', 'test_version', 'reproduce_steps',
        'expected_result', 'actual_result', 'resolution',
        'resolution_version', 'estimated_hours', 'actual_hours',
        'test_case_id', 'customer_mr_number', 'plan_resolve_version',
        'resolve_build', 'deadline', 'related_bug_id', 'parent_bug_id',
        'assigned_to', 'resolved_by', 'verifier_id'
    ]
    
    enum_fields = {
        'status': BugStatus,
        'priority': Priority,
        'severity': Severity
    }
    
    field_labels = {
        'title': '标题',
        'description': '描述',
        'priority': '优先级',
        'severity': '严重程度',
        'version': '版本',
        'module': '模块',
        'tags': '标签',
        'issue_type': '问题类型',
        'reproduce_frequency': '重现频率',
        'found_build': '发现构建',
        'test_version': '测试版本',
        'reproduce_steps': '重现步骤',
        'expected_result': '预期结果',
        'actual_result': '实际结果',
        'resolution': '解决方案',
        'resolution_version': '解决版本',
        'estimated_hours': '预计工时',
        'actual_hours': '实际工时',
        'test_case_id': '测试用例ID',
        'customer_mr_number': '客户MR编号',
        'plan_resolve_version': '计划解决版本',
        'resolve_build': '解决构建',
        'deadline': '期限',
        'related_bug_id': '关联Bug',
        'parent_bug_id': '父Bug'
    }
    
    field_changes = []
    
    # 先获取所有旧值，避免在循环中被覆盖
    old_values = {}
    for field in updateable_fields:
        if field in data:
            old_values[field] = getattr(bug, field, None)
    
    for field in updateable_fields:
        if field in data:
            old_value = old_values[field]
            new_value = data[field]
            
            old_value_display = None
            new_value_display = None
            
            if field in enum_fields:
                old_value_display = str(old_value) if old_value else ''
                if new_value:
                    new_value_str = str(new_value).lower()
                    valid_values = [e.value for e in enum_fields[field]]
                    if new_value_str in valid_values:
                        new_value_display = new_value_str
                    else:
                        new_value_display = str(new_value)
                else:
                    new_value_display = ''
            elif field in ['assigned_to', 'resolved_by', 'verifier_id']:
                if old_value:
                    old_user = User.query.get(int(old_value))
                    old_value_display = old_user.username if old_user else str(old_value)
                else:
                    old_value_display = '未分配'
                
                new_value_int = int(new_value) if new_value else None
                if new_value_int:
                    new_user = User.query.get(new_value_int)
                    new_value_display = new_user.username if new_user else str(new_value_int)
                else:
                    new_value_display = '未分配'
            elif field == 'deadline':
                if old_value:
                    old_value_display = old_value.strftime('%Y-%m-%d') if hasattr(old_value, 'strftime') else str(old_value)
                else:
                    old_value_display = '无'
                if data[field]:
                    try:
                        new_value = datetime.strptime(data[field], '%Y-%m-%d')
                        new_value_display = new_value.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            new_value = datetime.fromisoformat(data[field].replace('Z', '+00:00'))
                            new_value_display = new_value.strftime('%Y-%m-%d')
                        except ValueError:
                            new_value = None
                            new_value_display = '无'
                else:
                    new_value = None
                    new_value_display = '无'
            else:
                old_value_display = str(old_value) if old_value is not None else '无'
                new_value_display = str(new_value) if new_value is not None else '无'
            
            # 只有当值确实发生变化时才记录
            if old_value_display != new_value_display:
                field_label = field_labels.get(field, field)
                change_info = {
                    'field': field,
                    'field_label': field_label,
                    'old_value': old_value_display,
                    'new_value': new_value_display
                }
                field_changes.append(change_info)
            
            if field in data:
                if field in enum_fields:
                    value = data[field]
                    if value:
                        enum_cls = enum_fields[field]
                        value_str = str(value).lower()
                        valid_values = [e.value for e in enum_cls]
                        if value_str in valid_values:
                            setattr(bug, field, value_str)
                        else:
                            setattr(bug, field, str(value))
                    else:
                        setattr(bug, field, None)
                elif field == 'assigned_to':
                    bug.assigned_to = int(data[field]) if data[field] else None
                elif field == 'resolved_by':
                    bug.resolved_by = int(data[field]) if data[field] else None
                elif field == 'verifier_id':
                    bug.verifier_id = int(data[field]) if data[field] else None
                elif field == 'deadline' and data[field]:
                    try:
                        bug.deadline = datetime.strptime(data[field], '%Y-%m-%d')
                    except ValueError:
                        try:
                            bug.deadline = datetime.fromisoformat(data[field].replace('Z', '+00:00'))
                        except ValueError:
                            bug.deadline = None
                else:
                    # 特殊处理tags字段
                    if field == 'tags':
                        setattr(bug, field, parse_tags(data[field]))
                    else:
                        setattr(bug, field, data[field])
    
    bug.updated_at = datetime.utcnow()
    
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Updating bug {bug_id} with data: {data}")
    
    try:
        db.session.commit()
        
        logger.info(f"Bug {bug_id} updated successfully")

        if field_changes:
            resolved_by_change = next((fc for fc in field_changes if fc.get('field') == 'resolved_by'), None)
            if resolved_by_change and resolved_by_change.get('new_value') and resolved_by_change.get('new_value') != '未分配':
                new_resolver_name = resolved_by_change['new_value']
                new_resolver = User.query.filter_by(username=new_resolver_name).first()
                if new_resolver and new_resolver.email:
                    try:
                        from enhanced_app import send_email_notification as send_email
                        email_subject = f"[TOPO系统] 您被指定为Bug解决者: {bug.title}"
                        email_body = f"""亲爱的 {new_resolver.username}，

您已被指定为Bug "#{bug.id} {bug.title}" 的解决者。

指派人: {current_user.username}
指派时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}

请登录TOPO系统查看详情并及时处理。

此邮件由TOPO系统自动发送。
"""
                        html_body = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #007bff;">Bug解决者指派通知</h2>
        <p>亲爱的 <strong>{new_resolver.username}</strong>，</p>
        <p>您已被指定为Bug <strong>#{bug.id} {bug.title}</strong> 的解决者。</p>
        <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;"><strong>指派人</strong></td>
                <td style="padding: 8px; border: 1px solid #ddd;">{current_user.username}</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;"><strong>指派时间</strong></td>
                <td style="padding: 8px; border: 1px solid #ddd;">{datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
        </table>
        <p>请登录TOPO系统查看详情并及时处理。</p>
        <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
        <p style="font-size: 12px; color: #666;">
            此邮件由TOPO系统自动发送。<br>
            请勿回复此邮件。
        </p>
    </div>
</body>
</html>
"""
                        send_email(new_resolver.email, email_subject, email_body, html_body)
                        logger.info(f"Bug #{bug.id} 指定解决者 {new_resolver.username}，邮件通知已发送")
                    except Exception as e:
                        logger.error(f"发送Bug解决者指派邮件通知失败: {str(e)}")
                
                try:
                    import threading
                    def send_notification_async():
                        try:
                            create_notification = get_create_notification()
                            project = Project.query.get(bug.project_id)
                            notification = create_notification(
                                user_id=new_resolver.id,
                                notification_type='bug_assigned',
                                title=f'您被指定为Bug解决者: {bug.title}',
                                content=f'Bug #{bug.id} "{bug.title}" 已将您指定为解决者，请及时处理。\n\n项目: {project.name if project else "未知"}\n优先级: {bug.priority.value if hasattr(bug.priority, "value") else str(bug.priority)}\n严重程度: {bug.severity.value if hasattr(bug.severity, "value") else str(bug.severity)}',
                                related_bug_id=bug.id
                            )
                            logger.info(f"Bug #{bug.id} 指定解决者 {new_resolver.username}，系统通知已发送, notification_id: {notification.id if notification else None}")
                        except Exception as e:
                            logger.error(f"发送Bug解决者系统通知失败: {str(e)}")
                    
                    notification_thread = threading.Thread(target=send_notification_async)
                    notification_thread.daemon = True
                    notification_thread.start()
                except Exception as e:
                    logger.error(f"启动系统通知线程失败: {str(e)}")
        
        import json
        
        if field_changes:
            field_changes_json = json.dumps(field_changes, ensure_ascii=False)
            
            activity = Activity(
                action='update_bug',
                description=f'更新缺陷: {bug.title}',
                performed_by=int(current_user_id),
                target_type='bug',
                target_id=bug_id,
                field_changes=field_changes_json
            )
            db.session.add(activity)
            db.session.commit()
        
        create_audit_log(
            user_id=current_user_id,
            action='update_bug',
            resource_type='bug',
            resource_id=bug_id,
            details=f'更新缺陷: {bug.title}',
            request=request
        )
        
        return jsonify({
            'success': True,
            'message': '缺陷更新成功',
            'data': {'bug': bug.to_dict()}
        }), 200
    except Exception as e:
        import traceback
        db.session.rollback()
        logger.error(f"更新缺陷失败: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'更新缺陷失败: {str(e)}', 'data': None}), 500

# 更新缺陷状态
@bugs_bp.route('/<int:bug_id>/status', methods=['PUT'])
@jwt_required()
@require_bug_permission('bug:update_status')
def update_bug_status(bug_id):
    """更新缺陷状态"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    create_audit_log = get_create_audit_log()
    
    current_user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data or 'status' not in data:
        return jsonify({'success': False, 'message': '状态字段为必填项', 'data': None}), 400
    
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    current_user = User.query.get(current_user_id)
    if current_user.role != 'admin':
        member = ProjectMember.query.filter_by(
            project_id=bug.project_id,
            user_id=current_user_id
        ).first()
        if not member:
            is_resolver = bug.resolved_by == current_user_id
            is_reporter = bug.reported_by == current_user_id
            if not is_resolver and not is_reporter:
                return jsonify({'success': False, 'message': '无权更新此缺陷状态', 'data': None}), 403

    old_status = str(bug.status) if bug.status else ''
    new_status_value = data['status']
    
    valid_statuses = [status.value for status in BugStatus]
    if new_status_value not in valid_statuses:
        return jsonify({'success': False, 'message': f'无效的状态值，可选值: {valid_statuses}', 'data': None}), 400
    
    bug.status = new_status_value.lower()
    bug.updated_at = datetime.utcnow()
    
    if new_status_value == 'resolved' and not bug.resolved_at:
        bug.resolved_at = datetime.utcnow()
        bug.resolved_by = int(current_user_id)
    elif new_status_value == 'closed' and not bug.closed_at:
        bug.closed_at = datetime.utcnow()
    
    try:
        db.session.commit()
        
        # 如果状态变为resolved，发送邮件通知报告人
        if new_status_value == 'resolved' and old_status != 'resolved':
            try:
                reporter = User.query.get(bug.reported_by)
                if reporter and reporter.email:
                    from enhanced_app import send_email_notification as send_email
                    email_subject = f"[TOPO系统] Bug已解决: {bug.title}"
                    email_body = f"""亲爱的 {reporter.username}，

您报告的Bug "#{bug.id} {bug.title}" 已经被解决。

解决者: {current_user.username}
解决时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}

请登录TOPO系统查看详情。

此邮件由TOPO系统自动发送。
"""
                    html_body = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #28a745;">Bug已解决通知</h2>
        <p>亲爱的 <strong>{reporter.username}</strong>，</p>
        <p>您报告的Bug <strong>#{bug.id} {bug.title}</strong> 已经被解决。</p>
        <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;"><strong>解决者</strong></td>
                <td style="padding: 8px; border: 1px solid #ddd;">{current_user.username}</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;"><strong>解决时间</strong></td>
                <td style="padding: 8px; border: 1px solid #ddd;">{datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
        </table>
        <p>请登录TOPO系统查看详情。</p>
        <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
        <p style="font-size: 12px; color: #666;">
            此邮件由TOPO系统自动发送。<br>
            请勿回复此邮件。
        </p>
    </div>
</body>
</html>
"""
                    send_email(reporter.email, email_subject, email_body, html_body)
                    logger.info(f"Bug #{bug.id} 已解决，邮件通知已发送给报告人: {reporter.email}")
            except Exception as e:
                logger.error(f"发送Bug解决邮件通知失败: {str(e)}")
        
        create_audit_log(
            user_id=current_user_id,
            action='bug_status_update',
            resource_type='bug',
            resource_id=bug_id,
            details=f'更新缺陷状态: {old_status} -> {new_status_value}',
            request=request
        )

        import json
        field_changes = [{
            'field': 'status',
            'field_label': '状态',
            'old_value': old_status,
            'new_value': new_status_value
        }]
        field_changes_json = json.dumps(field_changes, ensure_ascii=False)

        activity = Activity(
            action='update_bug',
            description=f'更新缺陷状态: {old_status} -> {new_status_value}',
            performed_by=int(current_user_id),
            target_type='bug',
            target_id=bug_id,
            field_changes=field_changes_json
        )
        db.session.add(activity)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '状态更新成功',
            'data': {'bug': bug.to_dict()}
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新状态失败: {str(e)}', 'data': None}), 500

# 缺陷状态转换
@bugs_bp.route('/<int:bug_id>/transition', methods=['POST'])
@jwt_required()
@log_business_operation()
def transition_bug_status(bug_id):
    """缺陷状态转换"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    create_audit_log = get_create_audit_log()
    
    current_user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data:
        return jsonify({'success': False, 'message': '请求数据不能为空', 'data': None}), 400
    
    new_status = data.get('status')
    if not new_status:
        return jsonify({'success': False, 'message': '目标状态不能为空', 'data': None}), 400
    
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    current_user = User.query.get(current_user_id)
    if current_user.role != 'admin':
        member = ProjectMember.query.filter_by(
            project_id=bug.project_id,
            user_id=current_user_id
        ).first()
        if not member:
            is_resolver = int(bug.resolved_by) == int(current_user_id) if bug.resolved_by else False
            is_reporter = int(bug.reported_by) == int(current_user_id) if bug.reported_by else False
            if not is_resolver and not is_reporter:
                return jsonify({'success': False, 'message': '无权操作此缺陷', 'data': None}), 403
    
    old_status = str(bug.status) if bug.status else ''
    new_status_lower = new_status.lower()
    
    valid_statuses = [status.value for status in BugStatus]
    if new_status_lower not in valid_statuses:
        return jsonify({'success': False, 'message': f'无效的状态值，可选值: {valid_statuses}', 'data': None}), 400
    
    bug.status = new_status_lower
    bug.updated_at = datetime.utcnow()
    
    if new_status_lower == 'resolved':
        bug.resolved_at = datetime.utcnow()
        bug.resolved_by = int(current_user_id)
        if data.get('resolution'):
            bug.resolution = data['resolution']
        if not bug.verifier_id:
            bug.verifier_id = bug.reported_by
    elif new_status_lower == 'closed':
        bug.closed_at = datetime.utcnow()
    elif new_status_lower == 'reopened':
        bug.reopened_count = (bug.reopened_count or 0) + 1
    
    try:
        db.session.commit()
        
        # 如果状态变为resolved，发送邮件通知报告人
        if new_status_lower == 'resolved' and old_status != 'resolved':
            try:
                reporter = User.query.get(bug.reported_by)
                if reporter and reporter.email:
                    from enhanced_app import send_email_notification as send_email
                    email_subject = f"[TOPO系统] Bug已解决: {bug.title}"
                    email_body = f"""亲爱的 {reporter.username}，

您报告的Bug "#{bug.id} {bug.title}" 已经被解决。

解决者: {current_user.username}
解决时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}

请登录TOPO系统查看详情。

此邮件由TOPO系统自动发送。
"""
                    html_body = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #28a745;">Bug已解决通知</h2>
        <p>亲爱的 <strong>{reporter.username}</strong>，</p>
        <p>您报告的Bug <strong>#{bug.id} {bug.title}</strong> 已经被解决。</p>
        <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;"><strong>解决者</strong></td>
                <td style="padding: 8px; border: 1px solid #ddd;">{current_user.username}</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;"><strong>解决时间</strong></td>
                <td style="padding: 8px; border: 1px solid #ddd;">{datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
        </table>
        <p>请登录TOPO系统查看详情。</p>
        <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
        <p style="font-size: 12px; color: #666;">
            此邮件由TOPO系统自动发送。<br>
            请勿回复此邮件。
        </p>
    </div>
</body>
</html>
"""
                    send_email(reporter.email, email_subject, email_body, html_body)
                    logger.info(f"Bug #{bug.id} 已解决，邮件通知已发送给报告人: {reporter.email}")
            except Exception as e:
                logger.error(f"发送Bug解决邮件通知失败: {str(e)}")
        
        create_audit_log(
            user_id=current_user_id,
            action='bug_status_transition',
            resource_type='bug',
            resource_id=bug_id,
            details=f'Bug状态变更: {old_status} -> {new_status_lower}',
            request=request
        )

        import json
        field_changes = [{
            'field': 'status',
            'field_label': '状态',
            'old_value': old_status,
            'new_value': new_status_lower
        }]
        field_changes_json = json.dumps(field_changes, ensure_ascii=False)

        activity = Activity(
            action='update_bug',
            description=f'Bug状态变更: {old_status} -> {new_status_lower}',
            performed_by=int(current_user_id),
            target_type='bug',
            target_id=bug_id,
            field_changes=field_changes_json
        )
        db.session.add(activity)
        
        try:
            db.session.commit()
        except Exception as commit_error:
            db.session.rollback()
            logger.warning(f"Activity commit failed, skipping activity log: {commit_error}")
        
        try:
            bug_data = bug.to_dict()
        except Exception as dict_error:
            logger.warning(f"bug.to_dict() failed: {dict_error}")
            bug_data = {'id': bug.id, 'title': bug.title, 'status': new_status_lower}
        
        return jsonify({
            'success': True,
            'message': '状态转换成功',
            'data': {'bug': bug_data}
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'状态转换失败: {str(e)}', 'data': None}), 500

# 分配缺陷
@bugs_bp.route('/<int:bug_id>/assign', methods=['PUT'])
@jwt_required()
@require_bug_permission('bug:assign')
def assign_bug(bug_id):
    """分配缺陷"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    create_audit_log = get_create_audit_log()
    
    current_user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data or 'assignee_id' not in data:
        return jsonify({'success': False, 'message': '分配目标不能为空', 'data': None}), 400
    
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    current_user = User.query.get(current_user_id)
    if current_user.role not in ['admin', 'manager']:
        member = ProjectMember.query.filter_by(
            project_id=bug.project_id,
            user_id=current_user_id
        ).first()
        if not member:
            return jsonify({'success': False, 'message': '无权分配此缺陷', 'data': None}), 403
    
    assignee_id = data['assignee_id']
    if assignee_id:
        assignee = User.query.get(assignee_id)
        if not assignee:
            return jsonify({'success': False, 'message': '指定的分配用户不存在', 'data': None}), 404
    
    old_assignee = bug.assigned_to
    old_assignee_name = ''
    if old_assignee:
        old_assignee_user = User.query.get(old_assignee)
        old_assignee_name = old_assignee_user.username if old_assignee_user else str(old_assignee)
    
    new_assignee_name = ''
    if assignee_id:
        new_assignee_name = assignee.username if assignee else str(assignee_id)
    
    bug.assigned_to = assignee_id
    bug.updated_at = datetime.utcnow()
    
    try:
        db.session.commit()
        
        create_audit_log(
            user_id=current_user_id,
            action='assign_bug',
            resource_type='bug',
            resource_id=bug_id,
            details=f'分配缺陷: {bug.title}, 从用户 {old_assignee_name} 分配给用户 {new_assignee_name}',
            request=request
        )

        import json
        field_changes = [{
            'field': 'assigned_to',
            'field_label': '分配给',
            'old_value': old_assignee_name or '空',
            'new_value': new_assignee_name or '空'
        }]
        field_changes_json = json.dumps(field_changes, ensure_ascii=False)

        activity = Activity(
            action='update_bug',
            description=f'分配缺陷: {bug.title}, 从用户 {old_assignee_name} 分配给用户 {new_assignee_name}',
            performed_by=int(current_user_id),
            target_type='bug',
            target_id=bug_id,
            field_changes=field_changes_json
        )
        db.session.add(activity)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '分配成功',
            'data': {'bug': bug.to_dict()}
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'分配失败: {str(e)}', 'data': None}), 500

# 导出缺陷
@bugs_bp.route('/export', methods=['GET'])
@jwt_required()
def export_bugs():
    """导出缺陷列表为 Excel 文件"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    
    # 获取查询参数（与 get_bugs 相同）
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status')
    severity = request.args.get('severity')
    priority = request.args.get('priority')
    issue_type = request.args.get('issue_type')
    version = request.args.get('version')
    module = request.args.get('module')
    assignee = request.args.get('assignee', type=int)
    # 支持 keyword 或 search 参数
    keyword = request.args.get('keyword', '')
    search_param = request.args.get('search', '')
    search = keyword or search_param
    search = search.strip()
    
    # 构建查询
    if current_user.role in ['admin', 'manager', 'project_manager']:
        query = Bug.query
    else:
        user_projects = [m.project_id for m in ProjectMember.query.filter_by(user_id=current_user_id).all()]
        if not user_projects:
            return jsonify({'success': True, 'message': '无数据', 'data': {'bugs': []}}), 200
        query = Bug.query.filter(Bug.project_id.in_(user_projects))
    
    # 应用筛选
    if project_id:
        query = query.filter_by(project_id=project_id)
    if status:
        query = query.filter_by(status=status)
    if severity:
        query = query.filter_by(severity=severity)
    if priority:
        query = query.filter_by(priority=priority)
    if issue_type:
        query = query.filter_by(issue_type=issue_type)
    if version:
        query = query.filter_by(version=version)
    if module:
        query = query.filter_by(module=module)
    if assignee:
        query = query.filter_by(assigned_to=assignee)
    if search:
        query = query.filter(Bug.title.contains(search) | Bug.description.contains(search))
    
    # 获取所有匹配的缺陷
    bugs = query.all()
    
    # 缓存用户和项目信息以减少查询
    user_cache = {u.id: u for u in User.query.all()}
    project_cache = {p.id: p for p in Project.query.all()}
    
    # 转换为导出格式
    export_data = []
    for bug in bugs:
        project = project_cache.get(bug.project_id)
        reporter = user_cache.get(bug.reported_by)
        assignee = user_cache.get(bug.assigned_to) if bug.assigned_to else None
        resolver = user_cache.get(bug.resolved_by) if bug.resolved_by else None
        verifier = user_cache.get(bug.verifier_id) if bug.verifier_id else None
        
        export_data.append({
            'ID': bug.id,
            '标题': bug.title,
            '描述': bug.description,
            '状态': str(bug.status) if bug.status else '',
            '严重程度': str(bug.severity) if bug.severity else '',
            '优先级': str(bug.priority) if bug.priority else '',
            '项目': project.name if project else '',
            '负责人': assignee.username if assignee else '',
            '创建人': reporter.username if reporter else '',
            '创建时间': bug.created_at.strftime('%Y/%m/%d %H:%M:%S') if bug.created_at else '',
            '更新时间': bug.updated_at.strftime('%Y/%m/%d %H:%M:%S') if bug.updated_at else '',
            '归属版本': bug.version or '',
            '问题类型': bug.issue_type or '',
            '重现频率': bug.reproduce_frequency or '',
            '发现构建': bug.found_build or '',
            '测试版本': bug.test_version or '',
            '模块': bug.module or '',
            '客户 MR 编号': bug.customer_mr_number or '',
            '计划解决版本': bug.plan_resolve_version or '',
            '解决构建': bug.resolve_build or '',
            '解决者': resolver.username if resolver else '',
            '解决时间': bug.resolved_at.strftime('%Y/%m/%d %H:%M:%S') if bug.resolved_at else '',
            '验证者': verifier.username if verifier else '',
            '验证时间': bug.verified_at.strftime('%Y/%m/%d %H:%M:%S') if bug.verified_at else '',
            '截止日期': bug.deadline.strftime('%Y/%m/%d') if bug.deadline else '',
            '重新打开次数': bug.reopened_count or 0,
            '预计工时': bug.estimated_hours or '',
            '实际工时': bug.actual_hours or '',
            '关联测试用例': bug.test_case_id or '',
            '相关 Bug': bug.related_bug_id or '',
            '标签': bug.tags or '',
            '重现步骤': bug.steps_to_reproduce or '',
            '期望结果': bug.expected_result or '',
            '实际结果': bug.actual_result or '',
            '解决方案': bug.resolution or '',
            '解决版本': bug.resolution_version or '',
            '附件路径': bug.attachment_path or ''
        })
    
    # 创建 DataFrame 并导出为 Excel
    if not export_data:
        return jsonify({'success': False, 'message': '没有可导出的缺陷数据', 'data': None}), 404
    
    # 定义列顺序
    columns = [
        'ID', '标题', '描述', '状态', '严重程度', '优先级', '项目', '负责人',
        '创建人', '创建时间', '更新时间', '归属版本', '问题类型', '重现频率',
        '发现构建', '测试版本', '模块', '客户 MR 编号', '计划解决版本', '解决构建',
        '解决者', '解决时间', '验证者', '验证时间', '截止日期', '重新打开次数',
        '预计工时', '实际工时', '关联测试用例', '相关 Bug', '标签', '重现步骤',
        '期望结果', '实际结果', '解决方案', '解决版本', '附件路径'
    ]
    
    df = pd.DataFrame(export_data, columns=columns)
    
    output = BytesIO()
    try:
        df.to_excel(output, index=False, sheet_name='缺陷报告', engine='openpyxl')
        output.seek(0)
        
        filename = f'缺陷报告_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        
        response = make_response(send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        ))
        # 添加 CORS 头
        origin = request.headers.get('Origin')
        if origin:
            response.headers['Access-Control-Allow-Origin'] = origin
        else:
            response.headers['Access-Control-Allow-Origin'] = '*'
        return response
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'导出 Excel 失败：{str(e)}')
        if output:
            output.close()
        return jsonify({'success': False, 'message': f'导出 Excel 失败：{str(e)}', 'data': None}), 500

# 删除缺陷
@bugs_bp.route('/<int:bug_id>', methods=['DELETE'])
@jwt_required()
@require_bug_permission('bug:delete')
def delete_bug(bug_id):
    """删除缺陷"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    try:
        # 删除关联的评论
        Comment.query.filter_by(
            commentable_type='bug',
            commentable_id=bug_id
        ).delete()
        
        # 删除关联的附件记录
        Attachment.query.filter_by(
            attachable_type='bug',
            attachable_id=bug_id
        ).delete()
        
        # 删除缺陷
        db.session.delete(bug)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '缺陷删除成功', 'data': None}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'删除缺陷失败：{str(e)}', 'data': None}), 500

# 批量删除缺陷
@bugs_bp.route('/batch-delete', methods=['POST'])
@jwt_required()
@log_business_operation()
def batch_delete_bugs():
    """批量删除缺陷"""
    from flask import current_app
    
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    
    data = request.get_json()
    current_app.logger.info(f"批量删除请求数据: {data}")
    
    if not data or not data.get('bug_ids'):
        return jsonify({'success': False, 'message': '请提供要删除的缺陷ID列表', 'data': None}), 400
    
    bug_ids = data.get('bug_ids')
    current_app.logger.info(f"批量删除bug_ids: {bug_ids}")
    
    try:
        deleted_count = 0
        for bug_id in bug_ids:
            bug = Bug.query.get(bug_id)
            if bug:
                # 删除关联的评论
                Comment.query.filter_by(
                    commentable_type='bug',
                    commentable_id=bug_id
                ).delete()
                
                # 删除关联的附件记录
                Attachment.query.filter_by(
                    attachable_type='bug',
                    attachable_id=bug_id
                ).delete()
                
                # 删除缺陷
                db.session.delete(bug)
                deleted_count += 1
            else:
                current_app.logger.warning(f"未找到Bug ID: {bug_id}")
        
        db.session.commit()
        current_app.logger.info(f"成功删除 {deleted_count} 个缺陷")
        
        return jsonify({
            'success': True,
            'message': f'成功删除 {deleted_count} 个缺陷',
            'data': {'deleted_count': deleted_count}
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量删除缺陷失败: {str(e)}")
        return jsonify({'success': False, 'message': f'批量删除缺陷失败：{str(e)}', 'data': None}), 500

# 批量更新缺陷
@bugs_bp.route('/batch-update', methods=['PUT'])
@jwt_required()
@log_business_operation()
def batch_update_bugs():
    """批量更新缺陷"""
    from flask import current_app
    
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    
    data = request.get_json()
    current_app.logger.info(f"批量更新请求数据: {data}")
    
    if not data or not data.get('bug_ids'):
        return jsonify({'success': False, 'message': '请提供要更新的缺陷ID列表', 'data': None}), 400
    
    bug_ids = data.get('bug_ids')
    updates = data.get('updates', {})
    current_app.logger.info(f"批量更新bug_ids: {bug_ids}, updates: {updates}")
    
    if not updates:
        return jsonify({'success': False, 'message': '请提供要更新的字段', 'data': None}), 400
    
    try:
        bugs = Bug.query.filter(Bug.id.in_(bug_ids)).all()
        current_app.logger.info(f"查询到的Bug数量: {len(bugs)}")
        
        if not bugs:
            return jsonify({'success': False, 'message': '未找到要更新的缺陷', 'data': None}), 404
        
        allowed_fields = ['status', 'severity', 'priority', 'assigned_to', 'tags', 'version', 'module', 'deadline']
        
        enum_fields = {
            'status': BugStatus,
            'severity': Severity,
            'priority': Priority
        }
        
        update_count = 0
        for bug in bugs:
            for field, value in updates.items():
                if field in allowed_fields:
                    if field in enum_fields and value:
                        enum_cls = enum_fields[field]
                        value_str = str(value).lower()
                        valid_values = [e.value for e in enum_cls]
                        if value_str in valid_values:
                            setattr(bug, field, value_str)
                        else:
                            setattr(bug, field, str(value))
                    elif field == 'assigned_to' and value:
                        bug.assigned_to = int(value)
                    elif field == 'deadline' and value:
                        from datetime import datetime
                        try:
                            bug.deadline = datetime.strptime(value, '%Y-%m-%d')
                        except ValueError:
                            try:
                                bug.deadline = datetime.fromisoformat(value.replace('Z', '+00:00'))
                            except ValueError:
                                bug.deadline = None
                    else:
                        setattr(bug, field, value)
            update_count += 1
        
        db.session.commit()
        current_app.logger.info(f"成功更新 {update_count} 个缺陷")
        
        return jsonify({
            'success': True,
            'message': f'成功更新 {update_count} 个缺陷',
            'data': {'updated_count': update_count}
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量更新缺陷失败: {str(e)}")
        return jsonify({'success': False, 'message': f'批量更新缺陷失败：{str(e)}', 'data': None}), 500

# 导入缺陷
@bugs_bp.route('/import', methods=['POST'])
@jwt_required()
def import_bugs():
    """从 Excel 文件导入缺陷"""
    import csv
    from io import StringIO, BytesIO
    from datetime import datetime
    
    db = get_db()
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user:
        return jsonify({'success': False, 'message': '用户不存在', 'data': None}), 404
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请上传文件', 'data': None}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '文件名为空', 'data': None}), 400
    
    try:
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            if content.startswith('\ufeff'):
                content = content[1:]
            input_file = StringIO(content)
            reader = csv.DictReader(input_file)
            rows = list(reader)
            headers = reader.fieldnames if reader.fieldnames else []
        elif filename.endswith('.xlsx'):
            import openpyxl
            input_file = BytesIO(file.read())
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        if header and col_idx < len(row):
                            row_dict[header] = row[col_idx] if row[col_idx] is not None else ''
                        elif header:
                            row_dict[header] = ''
                    rows.append(row_dict)
        else:
            return jsonify({'success': False, 'message': '不支持的文件格式，请上传 CSV 或 Excel 文件', 'data': None}), 400
        
        required_columns = ['标题', '项目']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return jsonify({'success': False, 'message': f'文件缺少必要的列：{", ".join(missing_columns)}', 'data': None}), 400
        
        success_count = 0
        error_count = 0
        errors = []
        
        project_cache = {p.name.lower(): p for p in Project.query.all()}
        user_cache = {u.username.lower(): u for u in User.query.all()}
        
        status_map = {s.value.lower(): s for s in BugStatus}
        severity_map = {s.value.lower(): s for s in Severity}
        priority_map = {p.value.lower(): p for p in Priority}
        
        for row_idx, row in enumerate(rows, start=2):
            try:
                title = str(row.get('标题', '')).strip()
                if not title:
                    errors.append({'row': row_idx, 'error': '标题不能为空'})
                    error_count += 1
                    continue
                
                project_name = str(row.get('项目', '')).strip().lower()
                project = project_cache.get(project_name)
                if not project:
                    errors.append({'row': row_idx, 'error': f'项目不存在: {row.get("项目")}'})
                    error_count += 1
                    continue
                
                reporter_name = str(row.get('创建人', '')).strip().lower()
                if reporter_name:
                    reporter = user_cache.get(reporter_name)
                    if not reporter:
                        reporter = current_user
                else:
                    reporter = current_user
                
                status_value = str(row.get('状态', 'new')).strip().lower()
                status = status_map.get(status_value, BugStatus.NEW).value
                
                severity_value = str(row.get('严重程度', 'medium')).strip().lower()
                severity = severity_map.get(severity_value, Severity.MEDIUM).value
                
                priority_value = str(row.get('优先级', 'medium')).strip().lower()
                priority = priority_map.get(priority_value, Priority.MEDIUM).value
                
                assignee_name = str(row.get('负责人', '')).strip().lower()
                assigned_to = user_cache.get(assignee_name).id if assignee_name and user_cache.get(assignee_name) else None
                
                bug = Bug(
                    title=title,
                    description=str(row.get('描述', '')),
                    status=status,
                    severity=severity,
                    priority=priority,
                    project_id=project.id,
                    reported_by=reporter.id,
                    assigned_to=assigned_to,
                    version=str(row.get('归属版本', '')) or None,
                    tags=str(row.get('标签', '')) if row.get('标签') else None,
                    issue_type=str(row.get('问题类型', '')) or None,
                    reproduce_frequency=str(row.get('重现频率', '')) or None,
                    found_build=str(row.get('发现构建', '')) or None,
                    test_version=str(row.get('测试版本', '')) or None,
                    module=str(row.get('模块', '')) or None,
                    customer_mr_number=str(row.get('客户 MR 编号', '')) or None,
                    plan_resolve_version=str(row.get('计划解决版本', '')) or None,
                    resolve_build=str(row.get('解决构建', '')) or None,
                    resolution=str(row.get('解决方案', '')) or None,
                    resolution_version=str(row.get('解决版本', '')) or None,
                    reproduce_steps=str(row.get('重现步骤', '')) or None,
                    expected_result=str(row.get('期望结果', '')) or None,
                    actual_result=str(row.get('实际结果', '')) or None,
                    test_case_id=str(row.get('关联测试用例', '')) or None,
                    attachment_path=str(row.get('附件路径', '')) or None,
                )
                
                if row.get('创建时间'):
                    try:
                        bug.created_at = datetime.strptime(str(row.get('创建时间')), '%Y/%m/%d %H:%M:%S')
                    except:
                        pass
                
                if row.get('截止日期'):
                    try:
                        bug.deadline = datetime.strptime(str(row.get('截止日期')), '%Y/%m/%d')
                    except:
                        try:
                            bug.deadline = datetime.strptime(str(row.get('截止日期')), '%Y-%m-%d')
                        except:
                            pass
                
                if row.get('预计工时'):
                    try:
                        bug.estimated_hours = float(row.get('预计工时'))
                    except:
                        pass
                
                if row.get('实际工时'):
                    try:
                        bug.actual_hours = float(row.get('实际工时'))
                    except:
                        pass
                
                if row.get('重新打开次数'):
                    try:
                        bug.reopened_count = int(row.get('重新打开次数'))
                    except:
                        pass
                
                db.session.add(bug)
                success_count += 1
                
            except Exception as row_error:
                errors.append({'row': row_idx, 'error': str(row_error)})
                error_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {success_count} 个缺陷' + (f'，{error_count} 个失败' if error_count > 0 else ''),
            'data': {
                'imported_count': success_count,
                'error_count': error_count,
                'errors': errors[:50]
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"导入缺陷失败：{str(e)}")
        return jsonify({'success': False, 'message': f'导入缺陷失败：{str(e)}', 'data': None}), 500


# 上传 Bug 附件
@bugs_bp.route('/<int:bug_id>/attachments/upload', methods=['POST'])
@jwt_required()
@log_business_operation()
def upload_bug_attachment(bug_id):
    """上传 Bug 附件"""
    import logging
    logger = logging.getLogger(__name__)
    
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    current_user_id = get_jwt_identity()
    
    logger.info(f"收到上传请求 - Bug ID: {bug_id}, 用户 ID: {current_user_id}")
    
    try:
        # 验证 Bug 是否存在
        bug = Bug.query.get(bug_id)
        if not bug:
            logger.warning(f"Bug {bug_id} 不存在")
            return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
        
        # 检查权限
        current_user = User.query.get(current_user_id)
        if not current_user:
            logger.warning(f"用户 {current_user_id} 不存在")
            return jsonify({'success': False, 'message': '用户不存在', 'data': None}), 401
        
        if current_user.role != 'admin':
            member = ProjectMember.query.filter_by(
                project_id=bug.project_id,
                user_id=current_user_id
            ).first()
            if not member:
                logger.warning(f"用户 {current_user_id} 无权上传附件到 Bug {bug_id}")
                return jsonify({'success': False, 'message': '无权上传附件', 'data': None}), 403
        
        # 检查文件是否存在
        if 'file' not in request.files:
            logger.warning("请求中没有文件部分")
            return jsonify({'success': False, 'message': '未找到上传文件', 'data': None}), 400
        
        file = request.files['file']
        if file.filename == '':
            logger.warning("没有选择文件")
            return jsonify({'success': False, 'message': '未选择文件', 'data': None}), 400
        
        # 验证文件大小 (50MB)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        max_size = 50 * 1024 * 1024  # 50MB
        if file_size > max_size:
            logger.warning(f"文件大小 {file_size} 超过限制 {max_size}")
            return jsonify({'success': False, 'message': f'文件大小不能超过 50MB，当前文件大小: {file_size / (1024*1024):.2f}MB', 'data': None}), 400
        
        # 验证文件扩展名 - 允许所有文件类型，保留原始文件名
        original_filename = file.filename

        if not original_filename:
            logger.warning("文件名为空")
            return jsonify({'success': False, 'message': '文件名为空', 'data': None}), 400

        # 仅移除路径遍历等危险字符，保留原始文件名中的安全字符
        filename = secure_filename(original_filename)
        
        # 检查是否有路径遍历风险
        if '..' in original_filename or original_filename.startswith('/') or '\\' in original_filename:
            logger.warning(f"文件名包含路径遍历风险: {original_filename}")
            return jsonify({'success': False, 'message': '文件名包含非法字符', 'data': None}), 400
        
        # 如果secure_filename清空了文件名，说明包含完全非法的字符
        if not filename:
            logger.warning(f"文件名为空或包含非法字符: {original_filename}")
            return jsonify({'success': False, 'message': '文件名包含非法字符', 'data': None}), 400

        # 使用原始文件名保存（已通过安全检查）
        unique_filename = original_filename
        
        # 创建上传目录 - 使用绝对路径
        base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')
        upload_dir = os.path.join(base_dir, 'bugs', str(bug_id))
        
        # 确保目录存在
        try:
            os.makedirs(upload_dir, exist_ok=True)
        except Exception as e:
            logger.error(f"创建上传目录失败: {str(e)}")
            return jsonify({'success': False, 'message': f'创建上传目录失败: {str(e)}', 'data': None}), 500
        
        # 保存文件
        file_path = os.path.abspath(os.path.join(upload_dir, unique_filename))
        
        # 验证路径安全性
        if not file_path.startswith(os.path.abspath(base_dir)):
            logger.warning(f"文件路径不安全: {file_path}")
            return jsonify({'success': False, 'message': '无效的文件路径', 'data': None}), 400
        
        try:
            file.save(file_path)
        except Exception as e:
            logger.error(f"保存文件失败: {str(e)}")
            return jsonify({'success': False, 'message': f'保存文件失败: {str(e)}', 'data': None}), 500
        
        # 获取实际保存的文件大小
        saved_file_size = os.path.getsize(file_path)
        
        # 创建附件记录，使用原始文件名存储
        attachment = Attachment(
            filename=original_filename,
            file_path=f'bugs/{bug_id}/{unique_filename}',
            file_size=saved_file_size,
            mime_type=file.content_type or 'application/octet-stream',
            uploaded_by=current_user_id,
            created_by=current_user_id,
            attachable_type='bug',
            attachable_id=bug_id
        )
        
        db.session.add(attachment)
        
        # 创建活动记录
        activity = Activity(
            performed_by=current_user_id,
            action='upload_attachment',
            description=f"上传了附件：{original_filename}",
            target_type='bug',
            target_id=bug_id
        )
        db.session.add(activity)
        
        db.session.commit()
        
        logger.info(f"附件上传成功 - Bug ID: {bug_id}, 文件名: {original_filename}, 大小: {saved_file_size}")
        
        return jsonify({
            'success': True,
            'message': '附件上传成功',
            'data': {
                'attachment': attachment.to_dict()
            }
        }), 201
        
    except Exception as e:
        logger.error(f"上传附件时发生错误: {str(e)}", exc_info=True)
        db.session.rollback()
        return jsonify({'success': False, 'message': f'上传附件失败: {str(e)}', 'data': None}), 500


# 获取 Bug 附件列表
@bugs_bp.route('/<int:bug_id>/attachments', methods=['GET'])
@jwt_required()
def get_bug_attachments(bug_id):
    """获取 Bug 附件列表"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    # 验证 Bug 是否存在
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    # 获取所有关联到此 Bug 的附件，使用 joinedload 预加载 uploader 关系
    attachments = Attachment.query.options(
        joinedload(Attachment.uploader)
    ).filter_by(
        attachable_type='bug',
        attachable_id=bug_id
    ).all()
    
    return jsonify({
        'success': True,
        'message': '获取附件列表成功',
        'data': {
            'attachments': [attachment.to_dict() for attachment in attachments]
        }
    })


# 下载 Bug 附件
@bugs_bp.route('/<int:bug_id>/attachments/<int:attachment_id>', methods=['GET'])
@jwt_required()
def download_bug_attachment(bug_id, attachment_id):
    """下载 Bug 附件"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    # 获取附件
    attachment = Attachment.query.get(attachment_id)
    if not attachment:
        return jsonify({'success': False, 'message': '附件不存在', 'data': None}), 404
    
    # 验证附件是否属于此 Bug
    if attachment.attachable_type != 'bug' or attachment.attachable_id != bug_id:
        return jsonify({'success': False, 'message': '附件不属于此缺陷', 'data': None}), 404
    
    # 验证 Bug 是否存在
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    # 检查权限
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if current_user.role != 'admin':
        member = ProjectMember.query.filter_by(
            project_id=bug.project_id,
            user_id=current_user_id
        ).first()
        if not member:
            return jsonify({'success': False, 'message': '无权下载附件', 'data': None}), 403
    
    # 构建文件路径
    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')
    file_path = os.path.join(upload_dir, attachment.file_path)
    
    if not os.path.exists(file_path):
        return jsonify({'success': False, 'message': '文件不存在', 'data': None}), 404
    
    return send_file(
        file_path,
        as_attachment=True,
        download_name=attachment.filename,
        mimetype=attachment.mime_type
    )


# 删除 Bug 附件
@bugs_bp.route('/<int:bug_id>/attachments/<int:attachment_id>', methods=['DELETE'])
@jwt_required()
@log_business_operation()
def delete_bug_attachment(bug_id, attachment_id):
    """删除 Bug 附件"""
    User, Project, ProjectMember, BugStatus, Severity, Priority, Bug, Comment, Attachment, Activity, send_mention_notifications = get_models()
    
    db = get_db()
    current_user_id = get_jwt_identity()
    
    # 获取附件
    attachment = Attachment.query.get(attachment_id)
    if not attachment:
        return jsonify({'success': False, 'message': '附件不存在', 'data': None}), 404
    
    # 验证附件是否属于此 Bug
    if attachment.attachable_type != 'bug' or attachment.attachable_id != bug_id:
        return jsonify({'success': False, 'message': '附件不属于此缺陷', 'data': None}), 404
    
    # 验证 Bug 是否存在
    bug = Bug.query.get(bug_id)
    if not bug:
        return jsonify({'success': False, 'message': '缺陷不存在', 'data': None}), 404
    
    # 检查权限（只有管理员和上传者可删除）
    current_user = User.query.get(current_user_id)
    if current_user.role not in ['admin', 'manager'] and attachment.uploaded_by != current_user_id:
        return jsonify({'success': False, 'message': '无权删除此附件', 'data': None}), 403
    
    # 删除文件
    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')
    file_path = os.path.join(upload_dir, attachment.file_path)
    
    if os.path.exists(file_path):
        os.remove(file_path)
    
    # 删除附件记录
    db.session.delete(attachment)
    
    # 创建活动记录
    activity = Activity(
        performed_by=current_user_id,
        action='delete_attachment',
        description=f"删除了附件：{attachment.filename}",
        target_type='bug',
        target_id=bug_id
    )
    db.session.add(activity)
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '附件删除成功',
        'data': None
    })

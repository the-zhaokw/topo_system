"""
物料管理系统API接口
包含物料分类、物料主数据、仓库库位、库存管理、序列号管理等功能
"""

from flask import Blueprint, request, jsonify, current_app
from flask_restful import Api, Resource
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import or_, func
from datetime import datetime, timezone, timedelta

# 统一权限系统
from utils.permission_unified import (
    require_perm as _require_perm,
    require_any as _require_any,
    require_admin as _require_admin,
    check_perm as _check_perm,
    check_module as _check_module,
    filter_query_by_perm as _filter_query_by_perm,
    is_system_admin as _is_system_admin,
)

# 延迟导入数据库和模型以避免循环导入
def get_db():
    """延迟获取数据库实例，避免循环导入"""
    from enhanced_app import get_db_instance
    return get_db_instance()

def get_app():
    from enhanced_app import app
    return app

def get_models():
    from enhanced_app import MaterialCategory, Material, Warehouse, Location, Inventory, SerialNumber, InventoryTransaction, MaterialRelationship
    return MaterialCategory, Material, Warehouse, Location, Inventory, SerialNumber, InventoryTransaction, MaterialRelationship

# 创建物料管理蓝图
materials_bp = Blueprint('materials', __name__, url_prefix='/materials')
materials_api = Api(materials_bp)


def _check_perm_or_admin(user, perm_code):
    """内部权限校验：管理员/超管直接放行，否则校验细分权限。"""
    if not user:
        return False
    return _check_perm(user, perm_code)


def require_material_perm(perm_code):
    """物料子路由权限校验装饰器。"""
    from functools import wraps
    def decorator(f):
        @wraps(f)
        @jwt_required()
        def wrapper(*args, **kwargs):
            current_user_id = get_jwt_identity()
            try:
                current_user_id = int(current_user_id)
            except (TypeError, ValueError):
                pass
            from enhanced_app import User
            user = User.query.get(current_user_id)
            if not user:
                return {'error': '用户不存在'}, 404
            if not _check_perm_or_admin(user, perm_code):
                return {'error': '权限不足', 'code': 'PERMISSION_DENIED', 'required_permission': perm_code}, 403
            return f(*args, **kwargs)
        return wrapper
    return decorator

class MaterialCategoryResource(Resource):
    """物料分类管理API"""

    method_decorators = {
        'get': [require_material_perm('material:view')],
        'post': [require_material_perm('material:category_manage')],
        'put': [require_material_perm('material:category_manage')],
        'delete': [require_material_perm('material:category_manage')],
    }

    def get(self, category_id=None):
        """获取物料分类列表或单个分类详情"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, _, _, _, _, _, _ = get_models()
        
        with app.app_context():
            if category_id:
                category = db.session.get(MaterialCategory, category_id)
                if not category:
                    return {'error': '物料分类不存在'}, 404
                
                # 获取子分类
                children = db.session.query(MaterialCategory).filter_by(parent_id=category_id).all()
                
                return {
                    'id': category.id,
                    'name': category.name,
                    'code': category.code,
                    'description': category.description,
                    'parent_id': category.parent_id,
                    'level': category.level,
                    'children': [{
                        'id': child.id,
                        'name': child.name,
                        'code': child.code,
                        'description': child.description,
                        'parent_id': child.parent_id,
                        'level': child.level,
                        'created_at': child.created_at.isoformat() if child.created_at else None
                    } for child in children],
                    'created_at': category.created_at.isoformat() if category.created_at else None
                }
            
            # 获取查询参数
            parent_id = request.args.get('parent_id')
            
            # 构建查询
            query = db.session.query(MaterialCategory)
            
            if parent_id:
                query = query.filter_by(parent_id=parent_id)
            else:
                # 默认获取顶级分类
                query = query.filter_by(parent_id=None)
            
            categories = query.all()
            
            # 为每个分类获取子分类数量
            result = []
            for cat in categories:
                child_count = db.session.query(MaterialCategory).filter_by(parent_id=cat.id).count()
                material_count = db.session.query(Material).filter_by(category_id=cat.id).count()
                
                result.append({
                    'id': cat.id,
                    'name': cat.name,
                    'code': cat.code,
                    'description': cat.description,
                    'parent_id': cat.parent_id,
                    'level': cat.level,
                    'child_count': child_count,
                    'material_count': material_count,
                    'has_children': child_count > 0,
                    'created_at': cat.created_at.isoformat() if cat.created_at else None
                })
            
            return result
    
    def post(self):
        """创建物料分类"""
        db = get_db()
        app = get_app()
        MaterialCategory, _, _, _, _, _, _, _ = get_models()
        
        with app.app_context():
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['name', 'code']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查编码是否重复
            if db.session.query(MaterialCategory).filter_by(code=data['code']).first():
                return {'error': '分类编码已存在'}, 400
            
            # 创建分类
            category = MaterialCategory(
                name=data['name'],
                code=data['code'],
                description=data.get('description', ''),
                parent_id=data.get('parent_id'),
                level=data.get('level', 1)
            )
            
            db.session.add(category)
            db.session.commit()
            
            return {
                'id': category.id,
                'name': category.name,
                'code': category.code,
                'description': category.description,
                'parent_id': category.parent_id,
                'level': category.level,
                'message': '物料分类创建成功'
            }, 201
    
    def put(self, category_id):
        """更新物料分类"""
        db = get_db()
        app = get_app()
        MaterialCategory, _, _, _, _, _, _, _ = get_models()
        
        with app.app_context():
            data = request.get_json()
            
            category = db.session.get(MaterialCategory, category_id)
            if not category:
                return {'error': '物料分类不存在'}, 404
            
            # 检查编码是否重复（排除当前分类）
            if 'code' in data and data['code'] != category.code:
                if db.session.query(MaterialCategory).filter(MaterialCategory.code == data['code'], MaterialCategory.id != category_id).first():
                    return {'error': '分类编码已存在'}, 400
            
            # 更新字段
            if 'name' in data:
                category.name = data['name']
            if 'code' in data:
                category.code = data['code']
            if 'description' in data:
                category.description = data['description']
            if 'parent_id' in data:
                category.parent_id = data['parent_id']
            if 'level' in data:
                category.level = data['level']
            
            db.session.commit()
            
            return {
                'id': category.id,
                'name': category.name,
                'code': category.code,
                'description': category.description,
                'parent_id': category.parent_id,
                'level': category.level,
                'message': '物料分类更新成功'
            }
    
    def delete(self, category_id):
        """删除物料分类"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, _, _, _, _, _, _ = get_models()
        
        with app.app_context():
            category = db.session.get(MaterialCategory, category_id)
            if not category:
                return {'error': '物料分类不存在'}, 404
            
            # 检查是否有子分类
            sub_categories = db.session.query(MaterialCategory).filter_by(parent_id=category_id).all()
            if sub_categories:
                return {'error': '该分类下存在子分类，无法删除'}, 400
            
            # 检查是否有物料使用该分类
            materials = db.session.query(Material).filter_by(category_id=category_id).all()
            if materials:
                return {'error': '该分类下存在物料，无法删除'}, 400
            
            db.session.delete(category)
            db.session.commit()
            
            return {'message': '物料分类删除成功'}

class MaterialResource(Resource):
    """物料主数据管理API"""

    method_decorators = {
        'get': [require_material_perm('material:view')],
        'post': [require_material_perm('material:create')],
        'put': [require_material_perm('material:edit')],
        'delete': [require_material_perm('material:delete')],
    }

    def get(self, material_id=None):
        """获取物料列表或单个物料详情"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, _, _, _, _, _, _ = get_models()
        from sqlalchemy import or_
        
        with app.app_context():
            if material_id:
                material = db.session.get(Material, material_id)
                if not material:
                    return {'error': '物料不存在'}, 404
                
                return self._serialize_material(material, db)
            
            # 获取查询参数
            category_id = request.args.get('category_id')
            keyword = request.args.get('keyword')
            
            # 构建查询
            query = db.session.query(Material)
            
            if category_id:
                query = query.filter_by(category_id=category_id)
            
            if keyword:
                query = query.filter(or_(
                    Material.name.ilike(f'%{keyword}%'),
                    Material.code.ilike(f'%{keyword}%'),
                    Material.spec.ilike(f'%{keyword}%')
                ))
            
            materials = query.all()
            return [self._serialize_material(material, db) for material in materials]
    
    def post(self):
        """创建物料"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, _, _, _, _, _, _ = get_models()
        
        with app.app_context():
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['material_code', 'name', 'category_id', 'unit']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查物料编码是否重复
            if db.session.query(Material).filter_by(code=data['material_code']).first():
                return {'error': '物料编码已存在'}, 400
            
            # 检查分类是否存在
            category = db.session.get(MaterialCategory, data['category_id'])
            if not category:
                return {'error': '物料分类不存在'}, 400
            
            # 创建物料
            material = Material(
                code=data['material_code'],
                name=data['name'],
                category_id=data['category_id'],
                spec=data.get('specification', ''),
                unit=data['unit'],
                description=data.get('description', ''),
                safety_stock=data.get('safety_stock', 0.0),
                supplier=data.get('supplier', ''),
                supplier_code=data.get('supplier_code', ''),
                unit_cost=data.get('unit_cost', 0.0),
                created_by=data.get('created_by', 1)
            )
        
        db.session.add(material)
        db.session.commit()
        
        return self._serialize_material(material, db), 201
    
    def put(self, material_id):
        """更新物料"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, _, _, _, _, _, _ = get_models()
        
        with app.app_context():
            material = db.session.get(Material, material_id)
            if not material:
                return {'error': '物料不存在'}, 404
            
            data = request.get_json()
            
            # 更新字段
            if 'material_code' in data:
                # 检查物料编码是否重复（排除当前物料）
                if db.session.query(Material).filter(
                    Material.code == data['material_code'],
                    Material.id != material_id
                ).first():
                    return {'error': '物料编码已存在'}, 400
                material.code = data['material_code']
            
            if 'name' in data:
                material.name = data['name']
            if 'category_id' in data:
                # 检查分类是否存在
                category = db.session.get(MaterialCategory, data['category_id'])
                if not category:
                    return {'error': '物料分类不存在'}, 400
                material.category_id = data['category_id']
            if 'specification' in data:
                material.spec = data['specification']
            if 'unit' in data:
                material.unit = data['unit']
            if 'description' in data:
                material.description = data['description']
            if 'safety_stock' in data:
                material.safety_stock = data['safety_stock']
            if 'supplier' in data:
                material.supplier = data['supplier']
            if 'supplier_code' in data:
                material.supplier_code = data['supplier_code']
            if 'unit_cost' in data:
                material.unit_cost = data['unit_cost']
            if 'status' in data:
                material.status = data['status']
            
            db.session.commit()
            
            return {
                'id': material.id,
                'material_code': material.code,
                'name': material.name,
                'category_id': material.category_id,
                'message': '物料更新成功'
            }
    
    def delete(self, material_id):
        """删除物料"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, Warehouse, Location, Inventory, SerialNumber, _, _ = get_models()
        
        with app.app_context():
            material = db.session.get(Material, material_id)
            if not material:
                return {'error': '物料不存在'}, 404
            
            # 检查是否有库存记录
            inventory = db.session.query(Inventory).filter(Inventory.material_id == material_id).first()
            if inventory:
                return {'error': '该物料存在库存记录，无法删除'}, 400
            
            # 检查是否有序列号记录
            serial_numbers = db.session.query(SerialNumber).filter(SerialNumber.material_id == material_id).all()
            if serial_numbers:
                return {'error': '该物料存在序列号记录，无法删除'}, 400
            
            db.session.delete(material)
            db.session.commit()
            
            return {'message': '物料删除成功'}
    
    def _serialize_material(self, material, db):
        """序列化物料对象"""
        MaterialCategory, Material, Warehouse, Location, Inventory, SerialNumber, _, _ = get_models()
        
        # 处理分类信息
        category = None
        # 先尝试直接访问关联对象
        if hasattr(material, 'category') and material.category:
            category = material.category
        # 如果没有直接关联对象，但有category_id，则查询获取
        elif hasattr(material, 'category_id') and material.category_id:
            category = db.session.get(MaterialCategory, material.category_id)
        
        # 获取库存信息
        inventory = db.session.query(Inventory).filter(Inventory.material_id == material.id).all()
        total_quantity = sum([inv.quantity for inv in inventory])
        
        # 计算总价值
        total_value = total_quantity * (material.unit_cost or 0)
        
        # 检查是否低于安全库存（库存预警）
        safety_stock = material.safety_stock or 0
        is_low_stock = total_quantity < safety_stock
        
        # 获取序列号数量
        serial_count = db.session.query(SerialNumber).filter(SerialNumber.material_id == material.id).count()
        
        return {
            'id': material.id,
            'material_code': material.code,
            'name': material.name,
            'category_id': material.category_id,
            'category_name': category.name if category else '',
            'specification': material.spec,
            'unit': material.unit,
            'status': material.status,
            'description': material.description,
            'safety_stock': safety_stock,
            'supplier': material.supplier,
            'supplier_code': material.supplier_code,
            'unit_cost': material.unit_cost,
            'total_quantity': total_quantity,
            'total_value': total_value,
            'is_low_stock': is_low_stock,
            'serial_count': serial_count,
            'inventory_locations': [{
                'warehouse_name': db.session.get(Warehouse, inv.warehouse_id).name if db.session.get(Warehouse, inv.warehouse_id) else '',
                'location_name': db.session.get(Location, inv.location_id).name if db.session.get(Location, inv.location_id) else '',
                'quantity': inv.quantity
            } for inv in inventory],
            'created_at': material.created_at.isoformat() if material.created_at else None
        }

class WarehouseResource(Resource):
    """仓库管理API"""

    method_decorators = {
        'get': [require_material_perm('material:warehouse_manage')],
        'post': [require_material_perm('material:warehouse_manage')],
        'put': [require_material_perm('material:warehouse_manage')],
        'delete': [require_material_perm('material:warehouse_manage')],
    }

    def get(self, warehouse_id=None):
        """获取仓库列表或单个仓库详情"""
        db = get_db()
        app = get_app()
        _, _, Warehouse, _, _, _, _, _ = get_models()
        
        with app.app_context():
            if warehouse_id:
                warehouse = db.session.get(Warehouse, warehouse_id)
                if not warehouse:
                    return {'error': '仓库不存在'}, 404
                
                return self._serialize_warehouse(warehouse)
            
            warehouses = db.session.query(Warehouse).all()
            return [self._serialize_warehouse(warehouse) for warehouse in warehouses]
    
    def post(self):
        """创建仓库"""
        db = get_db()
        app = get_app()
        _, _, Warehouse, _, _, _, _, _ = get_models()
        
        with app.app_context():
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['code', 'name']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查仓库编码是否重复
            if db.session.query(Warehouse).filter_by(code=data['code']).first():
                return {'error': '仓库编码已存在'}, 400
            
            # 创建仓库
            warehouse = Warehouse(
                code=data['code'],
                name=data['name'],
                location=data.get('address', ''),
                contact=data.get('contact_person', ''),
                phone=data.get('contact_phone', ''),
                description=data.get('description', ''),
                status='active' if data.get('is_active', True) else 'inactive'
            )
            
            db.session.add(warehouse)
            db.session.commit()
            
            return {
                'id': warehouse.id,
                'code': warehouse.code,
                'name': warehouse.name,
                'message': '仓库创建成功'
            }, 201
    
    def put(self, warehouse_id):
        """更新仓库"""
        db = get_db()
        app = get_app()
        _, _, Warehouse, _, _, _, _, _ = get_models()
        
        with app.app_context():
            warehouse = db.session.get(Warehouse, warehouse_id)
            if not warehouse:
                return {'error': '仓库不存在'}, 404
            
            data = request.get_json()
            
            # 更新字段
            if 'code' in data:
                # 检查编码是否重复（排除当前仓库）
                if db.session.query(Warehouse).filter(
                    Warehouse.code == data['code'], 
                    Warehouse.id != warehouse_id
                ).first():
                    return {'error': '仓库编码已存在'}, 400
                warehouse.code = data['code']
            
            if 'name' in data:
                warehouse.name = data['name']
            if 'address' in data:
                warehouse.location = data['address']  # 将address映射到location字段
            if 'contact_person' in data:
                warehouse.contact = data['contact_person']  # 将contact_person映射到contact字段
            if 'contact_phone' in data:
                warehouse.phone = data['contact_phone']  # 将contact_phone映射到phone字段
            if 'description' in data:
                warehouse.description = data['description']
            if 'is_active' in data:
                warehouse.status = 'active' if data['is_active'] else 'inactive'

            db.session.commit()
            
            return {
                'id': warehouse.id,
                'code': warehouse.code,
                'name': warehouse.name,
                'message': '仓库更新成功'
            }
    
    def delete(self, warehouse_id):
        """删除仓库"""
        db = get_db()
        app = get_app()
        _, _, Warehouse, Location, Inventory, _, _, _ = get_models()
        
        with app.app_context():
            warehouse = db.session.get(Warehouse, warehouse_id)
            if not warehouse:
                return {'error': '仓库不存在'}, 404
            
            # 检查是否有库位
            locations = db.session.query(Location).filter_by(warehouse_id=warehouse_id).all()
            if locations:
                return {'error': '该仓库下存在库位，无法删除'}, 400
            
            # 检查是否有库存记录
            inventory = db.session.query(Inventory).filter_by(warehouse_id=warehouse_id).first()
            if inventory:
                return {'error': '该仓库存在库存记录，无法删除'}, 400
            
            db.session.delete(warehouse)
            db.session.commit()
            
            return {'message': '仓库删除成功'}
    
    def _serialize_warehouse(self, warehouse):
        """序列化仓库对象"""
        return {
            'id': warehouse.id,
            'code': warehouse.code,
            'name': warehouse.name,
            'type': getattr(warehouse, 'type', 'normal') or 'normal',
            'address': warehouse.location,
            'contact_person': warehouse.contact,
            'contact_phone': warehouse.phone,
            'description': warehouse.description,
            'is_active': warehouse.status == 'active' if warehouse.status else True,
            'created_at': warehouse.created_at.isoformat() if warehouse.created_at else None
        }

class LocationResource(Resource):
    """库位管理API"""

    method_decorators = {
        'get': [require_material_perm('material:warehouse_manage')],
        'post': [require_material_perm('material:warehouse_manage')],
        'put': [require_material_perm('material:warehouse_manage')],
        'delete': [require_material_perm('material:warehouse_manage')],
    }

    def get(self, location_id=None):
        """获取库位列表或单个库位详情"""
        db = get_db()
        app = get_app()
        
        with app.app_context():
            # 在app上下文内获取模型
            from enhanced_app import Location, Warehouse
            
            if location_id:
                location = db.session.get(Location, location_id)
                if not location:
                    return {'error': '库位不存在'}, 404
                
                return self._serialize_location(location, db)
            
            # 获取查询参数
            warehouse_id = request.args.get('warehouse_id')
            
            # 构建查询
            query = db.session.query(Location)
            if warehouse_id:
                query = query.filter_by(warehouse_id=warehouse_id)
            
            locations = query.all()
            return [self._serialize_location(location, db) for location in locations]
    
    def _serialize_location(self, location, db):
        """序列化库位对象"""
        # 获取关联的仓库信息
        warehouse = db.session.get(Warehouse, location.warehouse_id)
        return {
            'id': location.id,
            'warehouse_id': location.warehouse_id,
            'warehouse_name': warehouse.name if warehouse else None,
            'code': location.code,
            'name': location.name,
            'area': location.area,
            'zone': location.zone,
            'rack': location.rack,
            'level': location.level,
            'position': location.position,
            'capacity': location.capacity,
            'description': location.description,
            'created_at': location.created_at.isoformat() if location.created_at else None,
            'updated_at': location.updated_at.isoformat() if location.updated_at else None
        }
    
    def post(self):
        """创建库位"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, Warehouse, Location, Inventory, SerialNumber, InventoryTransaction, MaterialRelationship = get_models()
        
        with app.app_context():
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['warehouse_id', 'code', 'name']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查仓库是否存在
            warehouse = db.session.get(Warehouse, data['warehouse_id'])
            if not warehouse:
                return {'error': '仓库不存在'}, 400
            
            # 检查库位编码是否重复（在同一仓库内）
            if db.session.query(Location).filter_by(
                warehouse_id=data['warehouse_id'], 
                code=data['code']
            ).first():
                return {'error': '该仓库下库位编码已存在'}, 400
            
            # 创建库位
            location = Location(
                warehouse_id=data['warehouse_id'],
                code=data['code'],
                name=data['name'],
                area=data.get('area', ''),
                zone=data.get('zone', ''),
                rack=data.get('rack', ''),
                level=data.get('level', ''),
                capacity=data.get('capacity', 0),
                type=data.get('type', ''),
                status=data.get('status', 'active'),
                description=data.get('description', '')
            )
            
            db.session.add(location)
            db.session.commit()
            
            return {
                'id': location.id,
                'warehouse_id': location.warehouse_id,
                'warehouse_name': warehouse.name,
                'code': location.code,
                'name': location.name,
                'message': '库位创建成功'
            }, 201
    
    def put(self, location_id):
        """更新库位"""
        db = get_db()
        app = get_app()
        MaterialCategory, Material, Warehouse, Location, Inventory, SerialNumber, InventoryTransaction, MaterialRelationship = get_models()
        
        with app.app_context():
            location = db.session.get(Location, location_id)
            if not location:
                return {'error': '库位不存在'}, 404
            
            data = request.get_json()
            
            # 更新字段
            if 'warehouse_id' in data:
                warehouse = db.session.get(Warehouse, data['warehouse_id'])
                if not warehouse:
                    return {'error': '仓库不存在'}, 400
                location.warehouse_id = data['warehouse_id']
            
            if 'code' in data:
                # 检查编码是否重复（在同一仓库内，排除当前库位）
                if db.session.query(Location).filter(
                    Location.warehouse_id == location.warehouse_id,
                    Location.code == data['code'],
                    Location.id != location_id
                ).first():
                    return {'error': '该仓库下库位编码已存在'}, 400
                location.code = data['code']
            
            if 'name' in data:
                location.name = data['name']
            if 'area' in data:
                location.area = data['area']
            if 'shelf' in data:
                location.rack = data['shelf']  # 修正：使用正确的字段名
            if 'level' in data:
                location.level = data['level']
            if 'position' in data:
                location.position = data['position']
            if 'capacity' in data:
                location.capacity = data['capacity']
            if 'description' in data:
                location.description = data['description']
            
            db.session.commit()
            
            return {
                'id': location.id,
                'warehouse_id': location.warehouse_id,
                'code': location.code,
                'name': location.name,
                'message': '库位更新成功'
            }
    
    def delete(self, location_id):
        """删除库位"""
        db = get_db()
        app = get_app()
        Location, _, Inventory, _, _, _, _, _ = get_models()
        
        with app.app_context():
            location = db.session.get(Location, location_id)
            if not location:
                return {'error': '库位不存在'}, 404
            
            # 检查是否有库存记录
            inventory = db.session.query(Inventory).filter_by(location_id=location_id).first()
            if inventory:
                return {'error': '该库位存在库存记录，无法删除'}, 400
            
            db.session.delete(location)
            db.session.commit()
            
            return {'message': '库位删除成功'}
    
    def _serialize_location(self, location, db):
        """序列化库位对象"""
        Warehouse, _, _, _, _, _, _, _ = get_models()
        
        warehouse = db.session.get(Warehouse, location.warehouse_id)
        
        return {
            'id': location.id,
            'warehouse_id': location.warehouse_id,
            'warehouse_name': warehouse.name if warehouse else '',
            'warehouse_code': warehouse.code if warehouse else '',
            'code': location.code,  # 使用正确的字段名
            'name': location.name,
            'area': location.area,
            'zone': location.zone,
            'rack': location.rack,
            'level': location.level,
            'capacity': location.capacity,
            'created_at': location.created_at.isoformat() if location.created_at else None
        }

class InventoryResource(Resource):
    """库存管理API"""

    method_decorators = {
        'get': [require_material_perm('material:inventory_view')],
        'post': [require_material_perm('material:inventory_edit')],
        'put': [require_material_perm('material:inventory_edit')],
        'delete': [require_material_perm('material:inventory_edit')],
    }

    def get(self):
        """获取库存列表"""
        from enhanced_app import app, Inventory
        with app.app_context():
            db = get_db()
            # 获取查询参数
            material_id = request.args.get('material_id')
            warehouse_id = request.args.get('warehouse_id')
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 10, type=int)
            
            # 构建查询
            query = db.session.query(Inventory)
            
            if material_id:
                query = query.filter(Inventory.material_id == material_id)
            if warehouse_id:
                query = query.filter_by(warehouse_id=warehouse_id)
            
            # 获取总数
            total = query.count()
            
            # 分页
            inventory_items = query.offset((page - 1) * per_page).limit(per_page).all()
            
            return {
                'inventory': [self._serialize_inventory(item, db) for item in inventory_items],
                'total': total,
                'page': page,
                'per_page': per_page
            }
    
    def post(self):
        """创建库存记录"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['material_id', 'warehouse_id', 'quantity']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查物料是否存在
            material = db.session.get(Material, data['material_id'])
            if not material:
                return {'error': '物料不存在'}, 400
            
            # 检查仓库是否存在
            warehouse = db.session.get(Warehouse, data['warehouse_id'])
            if not warehouse:
                return {'error': '仓库不存在'}, 400
            
            # 检查库位是否存在（如果提供了库位ID）
            if data.get('location_id'):
                location = db.session.get(Location, data['location_id'])
                if not location:
                    return {'error': '库位不存在'}, 400
            
            # 创建库存记录
            inventory = Inventory(
                material_id=data['material_id'],
                warehouse_id=data['warehouse_id'],
                location_id=data.get('location_id'),
                quantity=data['quantity'],
                locked_quantity=data.get('locked_quantity', 0)
                # status字段已移除，不再支持
            )
            
            try:
                db.session.add(inventory)
                db.session.commit()
                
                # 创建入库交易记录
                transaction = InventoryTransaction(
                    transaction_type='in',
                    material_id=data['material_id'],
                    to_warehouse_id=data['warehouse_id'],
                    to_location_id=data.get('location_id'),
                    quantity=data['quantity'],
                    reference_id=f"INV_{inventory.id}",
                    remarks=f"初始入库: {data.get('remarks', '')}",
                    created_by=data.get('created_by', 'system'),
                    created_at=datetime.now(timezone.utc)
                )
                db.session.add(transaction)
                db.session.commit()
                
                return self._serialize_inventory(inventory, db), 201
            except Exception as e:
                db.session.rollback()
                return {'error': f'创建库存记录失败: {str(e)}'}, 500
    
    def put(self, inventory_id):
        """更新库存记录"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            inventory = db.session.get(Inventory, inventory_id)
            if not inventory:
                return {'error': '库存记录不存在'}, 404
            
            data = request.get_json()
            
            # 更新字段
            if 'quantity' in data:
                inventory.quantity = data['quantity']
            if 'locked_quantity' in data:
                inventory.locked_quantity = data['locked_quantity']
            # status字段已移除，不再支持
            if 'location_id' in data:
                # 检查库位是否存在
                if data['location_id']:
                    location = db.session.get(Location, data['location_id'])
                    if not location:
                        return {'error': '库位不存在'}, 400
                inventory.location_id = data['location_id']
            
            try:
                db.session.commit()
                return self._serialize_inventory(inventory, db)
            except Exception as e:
                db.session.rollback()
                return {'error': f'更新库存记录失败: {str(e)}'}, 500
    
    def delete(self, inventory_id):
        """删除库存记录"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            inventory = db.session.get(Inventory, inventory_id)
            if not inventory:
                return {'error': '库存记录不存在'}, 404
            
            # 检查是否有交易记录
            transactions = db.session.query(InventoryTransaction).filter(
                or_(InventoryTransaction.from_warehouse_id == inventory.warehouse_id,
                    InventoryTransaction.to_warehouse_id == inventory.warehouse_id),
                InventoryTransaction.material_id == inventory.material_id
            ).count()
            
            if transactions > 0:
                return {'error': '存在交易记录的库存不能删除'}, 400
            
            try:
                db.session.delete(inventory)
                db.session.commit()
                return {'message': '库存记录删除成功'}
            except Exception as e:
                db.session.rollback()
                return {'error': f'删除库存记录失败: {str(e)}'}, 500
    
    def _serialize_inventory(self, inventory, db):
        """序列化库存对象"""
        from enhanced_app import Material, Warehouse, Location
        material = db.session.get(Material, inventory.material_id)
        warehouse = db.session.get(Warehouse, inventory.warehouse_id)
        location = db.session.get(Location, inventory.location_id) if inventory.location_id else None
        
        return {
            'id': inventory.id,
            'material_id': inventory.material_id,
            'material_code': material.code if material else '',
            'material_name': material.name if material else '',
            'warehouse_id': inventory.warehouse_id,
            'warehouse_name': warehouse.name if warehouse else '',
            'location_id': inventory.location_id,
            'location_name': location.name if location else '',
            'quantity': float(inventory.quantity),
            'locked_quantity': float(inventory.locked_quantity),
            'available_quantity': float(inventory.quantity - inventory.locked_quantity),
            'last_updated': inventory.last_updated.isoformat() if inventory.last_updated else None
        }

class InventoryStatsResource(Resource):
    """库存统计API"""

    method_decorators = {'get': [require_material_perm('material:view_reports')]}

    def get(self):
        """获取库存统计数据"""
        try:
            from enhanced_app import app, Inventory, Material, Warehouse
            with app.app_context():
                db = get_db()

                try:
                    total_materials = db.session.query(Inventory).distinct(Inventory.material_id).count()
                except:
                    total_materials = 0

                try:
                    total_quantity = db.session.query(db.func.sum(Inventory.quantity)).scalar() or 0
                except:
                    total_quantity = 0

                try:
                    total_warehouses = db.session.query(Warehouse).count()
                except:
                    total_warehouses = 0

                return {
                    'stats': {
                        'total_materials': total_materials,
                        'total_quantity': float(total_quantity),
                        'low_stock_count': 0,
                        'total_warehouses': total_warehouses
                    }
                }
        except Exception as e:
            import logging
            logging.error(f"Error getting inventory stats: {str(e)}")
            return {'stats': {'total_materials': 0, 'total_quantity': 0, 'low_stock_count': 0, 'total_warehouses': 0}}


class InventoryTransactionResource(Resource):
    """库存交易管理API"""

    method_decorators = {
        'get': [require_material_perm('material:inventory_view')],
        'post': [require_material_perm('material:inventory_edit')],
        'put': [require_material_perm('material:inventory_edit')],
        'delete': [require_material_perm('material:inventory_edit')],
    }

    def get(self, transaction_id=None):
        """获取库存交易记录"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            if transaction_id:
                transaction = db.session.get(InventoryTransaction, transaction_id)
                if not transaction:
                    return {'error': '交易记录不存在'}, 404
                
                return self._serialize_transaction(transaction, db)
            
            # 获取查询参数
            material_id = request.args.get('material_id')
            transaction_type = request.args.get('transaction_type')
            
            # 构建查询
            query = db.session.query(InventoryTransaction)
            
            if material_id:
                query = query.filter(InventoryTransaction.material_id == material_id)
            if transaction_type:
                query = query.filter_by(transaction_type=transaction_type)
            
            # 按时间倒序排列
            transactions = query.order_by(InventoryTransaction.created_at.desc()).limit(100).all()
            return [self._serialize_transaction(transaction, db) for transaction in transactions]
    
    def post(self):
        """创建库存交易（入库、出库、调拨）"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['transaction_type', 'material_id', 'quantity']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查物料是否存在
            material = db.session.get(Material, data['material_id'])
            if not material:
                return {'error': '物料不存在'}, 400
            
            # 验证交易类型
            valid_transaction_types = ['in', 'out', 'transfer']
            if data['transaction_type'] not in valid_transaction_types:
                return {'error': f'交易类型必须是以下之一: {valid_transaction_types}'}, 400
            
            # 验证仓库和库位
            if data['transaction_type'] == 'in':
                if not data.get('to_warehouse_id'):
                    return {'error': '入库交易必须指定目标仓库'}, 400
                warehouse = db.session.get(Warehouse, data['to_warehouse_id'])
                if not warehouse:
                    return {'error': '目标仓库不存在'}, 400
                if data.get('to_location_id'):
                    location = db.session.get(Location, data['to_location_id'])
                    if not location:
                        return {'error': '目标库位不存在'}, 400
            elif data['transaction_type'] == 'out':
                if not data.get('from_warehouse_id'):
                    return {'error': '出库交易必须指定源仓库'}, 400
                warehouse = db.session.get(Warehouse, data['from_warehouse_id'])
                if not warehouse:
                    return {'error': '源仓库不存在'}, 400
                if data.get('from_location_id'):
                    location = db.session.get(Location, data['from_location_id'])
                    if not location:
                        return {'error': '源库位不存在'}, 400
            elif data['transaction_type'] == 'transfer':
                if not data.get('from_warehouse_id') or not data.get('to_warehouse_id'):
                    return {'error': '调拨交易必须指定源仓库和目标仓库'}, 400
                from_warehouse = db.session.get(Warehouse, data['from_warehouse_id'])
                to_warehouse = db.session.get(Warehouse, data['to_warehouse_id'])
                if not from_warehouse or not to_warehouse:
                    return {'error': '源仓库或目标仓库不存在'}, 400
            
            # 执行库存操作
            try:
                if data['transaction_type'] == 'in':
                    result = self._process_inbound_transaction(data, db)
                elif data['transaction_type'] == 'out':
                    result = self._process_outbound_transaction(data, db)
                elif data['transaction_type'] == 'transfer':
                    result = self._process_transfer_transaction(data, db)
                
                return result
            except Exception as e:
                db.session.rollback()
                return {'error': f'库存操作失败: {str(e)}'}, 500
    
    def _process_inbound_transaction(self, data, db):
        """处理入库交易"""
        # 查找或创建库存记录
        inventory = db.session.query(Inventory).filter_by(
            material_id=data['material_id'],
            warehouse_id=data['to_warehouse_id'],
            location_id=data.get('to_location_id')
        ).first()
        
        if inventory:
            # 更新现有库存
            inventory.quantity += data['quantity']
        else:
            # 创建新库存记录
            inventory = Inventory(
                material_id=data['material_id'],
                warehouse_id=data['to_warehouse_id'],
                location_id=data.get('to_location_id'),
                quantity=data['quantity'],
                locked_quantity=0
                # status字段已移除，不再支持
            )
            db.session.add(inventory)
        
        # 创建交易记录
        transaction = InventoryTransaction(
            transaction_type='in',
            material_id=data['material_id'],
            serial_number=data.get('serial_number'),
            to_warehouse_id=data['to_warehouse_id'],
            to_location_id=data.get('to_location_id'),
            quantity=data['quantity'],
            reference_id=data.get('reference_id', ''),
            remarks=data.get('remarks', ''),
            created_by=data.get('created_by', 'system'),
            created_at=datetime.now(timezone.utc)
        )
        db.session.add(transaction)
        
        # 如果有序列号，更新序列号状态
        if data.get('serial_number'):
            serial_number = db.session.query(SerialNumber).filter_by(
                serial_number=data['serial_number']
            ).first()
            if serial_number:
                serial_number.current_status = 'in_stock'
                serial_number.current_location_id = data.get('to_location_id')
        
        db.session.commit()
        
        return {
            'id': transaction.id,
            'transaction_type': transaction.transaction_type,
            'material_id': transaction.material_id,
            'material_name': db.session.get(Material, data['material_id']).name,
            'message': '入库交易创建成功'
        }, 201
    
    def _process_outbound_transaction(self, data, db):
        """处理出库交易"""
        # 查找库存记录
        inventory = db.session.query(Inventory).filter_by(
            material_id=data['material_id'],
            warehouse_id=data['from_warehouse_id'],
            location_id=data.get('from_location_id')
        ).first()
        
        if not inventory:
            return {'error': '库存记录不存在'}, 400
        
        # 检查库存数量是否足够
        available_quantity = inventory.quantity - inventory.locked_quantity
        if available_quantity < data['quantity']:
            return {'error': f'库存不足，可用数量: {available_quantity}'}, 400
        
        # 更新库存
        inventory.quantity -= data['quantity']
        
        # 创建交易记录
        transaction = InventoryTransaction(
            transaction_type='out',
            material_id=data['material_id'],
            serial_number=data.get('serial_number'),
            from_warehouse_id=data['from_warehouse_id'],
            from_location_id=data.get('from_location_id'),
            quantity=data['quantity'],
            reference_id=data.get('reference_id', ''),
            remarks=data.get('remarks', ''),
            created_by=data.get('created_by', 'system'),
            created_at=datetime.now(timezone.utc)
        )
        db.session.add(transaction)
        
        # 如果有序列号，更新序列号状态
        if data.get('serial_number'):
            serial_number = db.session.query(SerialNumber).filter_by(
                serial_number=data['serial_number']
            ).first()
            if serial_number:
                serial_number.current_status = 'in_use'
                serial_number.current_location_id = None
        
        db.session.commit()
        
        return {
            'id': transaction.id,
            'transaction_type': transaction.transaction_type,
            'material_id': transaction.material_id,
            'material_name': db.session.get(Material, data['material_id']).name,
            'message': '出库交易创建成功'
        }, 201
    
    def _process_transfer_transaction(self, data, db):
        """处理调拨交易"""
        # 查找源库存记录
        from_inventory = db.session.query(Inventory).filter_by(
            material_id=data['material_id'],
            warehouse_id=data['from_warehouse_id'],
            location_id=data.get('from_location_id')
        ).first()
        
        if not from_inventory:
            return {'error': '源库存记录不存在'}, 400
        
        # 检查源库存数量是否足够
        available_quantity = from_inventory.quantity - from_inventory.locked_quantity
        if available_quantity < data['quantity']:
            return {'error': f'源库存不足，可用数量: {available_quantity}'}, 400
        
        # 更新源库存
        from_inventory.quantity -= data['quantity']
        
        # 查找或创建目标库存记录
        to_inventory = db.session.query(Inventory).filter_by(
            material_id=data['material_id'],
            warehouse_id=data['to_warehouse_id'],
            location_id=data.get('to_location_id')
        ).first()
        
        if to_inventory:
            # 更新目标库存
            to_inventory.quantity += data['quantity']
        else:
            # 创建目标库存记录
            to_inventory = Inventory(
                material_id=data['material_id'],
                warehouse_id=data['to_warehouse_id'],
                location_id=data.get('to_location_id'),
                quantity=data['quantity'],
                locked_quantity=0
                # status字段已移除，不再支持
            )
            db.session.add(to_inventory)
        
        # 创建交易记录
        transaction = InventoryTransaction(
            transaction_type='transfer',
            material_id=data['material_id'],
            serial_number=data.get('serial_number'),
            from_warehouse_id=data['from_warehouse_id'],
            from_location_id=data.get('from_location_id'),
            to_warehouse_id=data['to_warehouse_id'],
            to_location_id=data.get('to_location_id'),
            quantity=data['quantity'],
            reference_id=data.get('reference_id', ''),
            remarks=data.get('remarks', ''),
            created_by=data.get('created_by', 'system'),
            created_at=datetime.now(timezone.utc)
        )
        db.session.add(transaction)
        
        # 如果有序列号，更新序列号位置
        if data.get('serial_number'):
            serial_number = db.session.query(SerialNumber).filter_by(
                serial_number=data['serial_number']
            ).first()
            if serial_number:
                serial_number.current_location_id = data.get('to_location_id')
        
        db.session.commit()
        
        return {
            'id': transaction.id,
            'transaction_type': transaction.transaction_type,
            'material_id': transaction.material_id,
            'material_name': db.session.get(Material, data['material_id']).name,
            'message': '调拨交易创建成功'
        }, 201
    
    def put(self, transaction_id):
        """更新库存交易记录"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            transaction = db.session.get(InventoryTransaction, transaction_id)
            if not transaction:
                return {'error': '交易记录不存在'}, 404
            
            data = request.get_json()
            
            # 更新字段
            if 'remarks' in data:
                transaction.remarks = data['remarks']
            if 'reference_id' in data:
                transaction.reference_id = data['reference_id']
            
            try:
                db.session.commit()
                return self._serialize_transaction(transaction, db)
            except Exception as e:
                db.session.rollback()
                return {'error': f'更新交易记录失败: {str(e)}'}, 500
    
    def delete(self, transaction_id):
        """删除库存交易记录"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            transaction = db.session.get(InventoryTransaction, transaction_id)
            if not transaction:
                return {'error': '交易记录不存在'}, 404
            
            try:
                db.session.delete(transaction)
                db.session.commit()
                return {'message': '交易记录删除成功'}
            except Exception as e:
                db.session.rollback()
                return {'error': f'删除交易记录失败: {str(e)}'}, 500
    
    def _serialize_transaction(self, transaction, db):
        """序列化交易对象"""
        material = db.session.get(Material, transaction.material_id)
        from_warehouse = db.session.get(Warehouse, transaction.from_warehouse_id) if transaction.from_warehouse_id else None
        to_warehouse = db.session.get(Warehouse, transaction.to_warehouse_id) if transaction.to_warehouse_id else None
        from_location = db.session.get(Location, transaction.from_location_id) if transaction.from_location_id else None
        to_location = db.session.get(Location, transaction.to_location_id) if transaction.to_location_id else None
        serial_number = db.session.get(SerialNumber, transaction.serial_number_id) if transaction.serial_number_id else None
        
        return {
            'id': transaction.id,
            'transaction_type': transaction.transaction_type,
            'material_id': transaction.material_id,
            'material_code': material.code if material else '',
            'material_name': material.name if material else '',
            'serial_number_id': transaction.serial_number_id,
            'serial_number': serial_number.serial_number if serial_number else '',
            'from_warehouse_id': transaction.from_warehouse_id,
            'from_warehouse_name': from_warehouse.name if from_warehouse else '',
            'from_location_id': transaction.from_location_id,
            'from_location_name': from_location.name if from_location else '',
            'to_warehouse_id': transaction.to_warehouse_id,
            'to_warehouse_name': to_warehouse.name if to_warehouse else '',
            'to_location_id': transaction.to_location_id,
            'to_location_name': to_location.name if to_location else '',
            'quantity': float(transaction.quantity),
            'reference_id': transaction.reference_id,
            'remarks': transaction.remarks,
            'created_by': transaction.created_by,
            'created_at': transaction.created_at.isoformat() if transaction.created_at else None
        }

class SerialNumberResource(Resource):
    """序列号管理API"""

    method_decorators = {
        'get': [require_material_perm('material:serial_manage')],
        'post': [require_material_perm('material:serial_manage')],
        'put': [require_material_perm('material:serial_manage')],
        'delete': [require_material_perm('material:serial_manage')],
    }

    def get(self, serial_number_id=None):
        """获取序列号信息"""
        from enhanced_app import app, SerialNumber
        with app.app_context():
            db = get_db()
            if serial_number_id:
                serial_number = db.session.get(SerialNumber, serial_number_id)
                if not serial_number:
                    return {'error': '序列号不存在'}, 404
                
                return self._serialize_serial_number(serial_number, db)
            
            # 获取查询参数
            material_id = request.args.get('material_id')
            status = request.args.get('status')
            
            # 构建查询
            query = db.session.query(SerialNumber)
            
            if material_id:
                query = query.filter(SerialNumber.material_id == material_id)
            if status:
                query = query.filter_by(current_status=status)
            
            serial_numbers = query.all()
            return [self._serialize_serial_number(sn, db) for sn in serial_numbers]
    
    def post(self):
        """创建序列号"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['serial_number', 'material_id']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查序列号是否重复
            if db.session.query(SerialNumber).filter_by(serial_number=data['serial_number']).first():
                return {'error': '序列号已存在'}, 400
            
            # 检查物料是否存在
            material = db.session.get(Material, data['material_id'])
            if not material:
                return {'error': '物料不存在'}, 400
            
            # 创建序列号
            serial_number = SerialNumber(
                serial_number=data['serial_number'],
                material_id=data['material_id'],
                current_status=data.get('current_status', 'in_stock'),
                current_location_id=data.get('current_location_id'),
                supplier_batch=data.get('supplier_batch', ''),
                purchase_date=data.get('purchase_date')
            )
            
            db.session.add(serial_number)
            db.session.commit()
            
            return {
                'id': serial_number.id,
                'serial_number': serial_number.serial_number,
                'material_id': serial_number.material_id,
                'material_name': material.name,
                'message': '序列号创建成功'
            }, 201
    
    def put(self, serial_number_id):
        """更新序列号"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            serial_number = db.session.get(SerialNumber, serial_number_id)
            if not serial_number:
                return {'error': '序列号不存在'}, 404
            
            data = request.get_json()
            
            # 更新字段
            if 'serial_number' in data:
                # 检查序列号是否重复（排除当前序列号）
                if db.session.query(SerialNumber).filter(
                    SerialNumber.serial_number == data['serial_number'],
                    SerialNumber.id != serial_number_id
                ).first():
                    return {'error': '序列号已存在'}, 400
                serial_number.serial_number = data['serial_number']
            
            if 'material_id' in data:
                material = db.session.get(Material, data['material_id'])
                if not material:
                    return {'error': '物料不存在'}, 400
                serial_number.material_id = data['material_id']
            
            if 'current_status' in data:
                serial_number.current_status = data['current_status']
            if 'current_location_id' in data:
                serial_number.current_location_id = data['current_location_id']
            if 'supplier_batch' in data:
                serial_number.supplier_batch = data['supplier_batch']
            if 'purchase_date' in data:
                serial_number.purchase_date = data['purchase_date']
            
            db.session.commit()
            
            return {
                'id': serial_number.id,
                'serial_number': serial_number.serial_number,
                'material_id': serial_number.material_id,
                'message': '序列号更新成功'
            }
    
    def delete(self, serial_number_id):
        """删除序列号"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            serial_number = db.session.get(SerialNumber, serial_number_id)
            if not serial_number:
                return {'error': '序列号不存在'}, 404
            
            # 检查序列号是否被使用（有交易记录）
            transaction_count = db.session.query(InventoryTransaction).filter_by(
                serial_number=serial_number.serial_number
            ).count()
            
            relationship_count = 0
            
            if transaction_count > 0 or relationship_count > 0:
                return {'error': '序列号已被使用，无法删除'}, 400
            
            try:
                db.session.delete(serial_number)
                db.session.commit()
                return {'message': '序列号删除成功'}
            except Exception as e:
                db.session.rollback()
                return {'error': f'删除序列号失败: {str(e)}'}, 500
    
    def _serialize_serial_number(self, serial_number, db):
        """序列化序列号对象"""
        material = db.session.get(Material, serial_number.material_id)
        location = db.session.get(Location, serial_number.current_location_id) if serial_number.current_location_id else None
        
        return {
            'id': serial_number.id,
            'serial_number': serial_number.serial_number,
            'material_id': serial_number.material_id,
            'material_code': material.code if material else '',
            'material_name': material.name if material else '',
            'current_status': serial_number.current_status,
            'current_location_id': serial_number.current_location_id,
            'current_location_name': location.name if location else '',
            'supplier_batch': serial_number.supplier_batch,
            'purchase_date': serial_number.purchase_date.isoformat() if serial_number.purchase_date else None,
            'created_at': serial_number.created_at.isoformat() if serial_number.created_at else None
        }

class MaterialRelationshipResource(Resource):
    """物料关系管理API"""

    method_decorators = {
        'get': [require_material_perm('material:view')],
        'post': [require_material_perm('material:edit')],
        'put': [require_material_perm('material:edit')],
        'delete': [require_material_perm('material:delete')],
    }

    def get(self, relationship_id=None):
        """获取物料关系"""
        from enhanced_app import app, MaterialRelationship
        with app.app_context():
            db = get_db()
            if relationship_id:
                relationship = db.session.get(MaterialRelationship, relationship_id)
                if not relationship:
                    return {'error': '关系记录不存在'}, 404
                
                return self._serialize_relationship(relationship, db)
            
            # 获取查询参数
            source_material_id = request.args.get('source_material_id', type=int)
            target_material_id = request.args.get('target_material_id', type=int)
            
            # 构建查询
            query = db.session.query(MaterialRelationship)
            
            if source_material_id:
                query = query.filter_by(source_material_id=source_material_id)
            if target_material_id:
                query = query.filter_by(target_material_id=target_material_id)
            
            relationships = query.all()
            return [self._serialize_relationship(rel, db) for rel in relationships]
    
    def post(self):
        """创建物料关系"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['source_material_id', 'target_material_id', 'relationship_type']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 检查关系是否已存在
            existing_relationship = db.session.query(MaterialRelationship).filter_by(
                source_material_id=data['source_material_id'],
                target_material_id=data['target_material_id']
            ).first()
            
            if existing_relationship:
                return {'error': '该物料关系已存在'}, 400
            
            # 创建关系
            relationship = MaterialRelationship(
                source_material_id=data['source_material_id'],
                target_material_id=data['target_material_id'],
                relationship_type=data['relationship_type'],
                quantity=data.get('quantity', 1.0),
                description=data.get('description', '')
            )
            
            try:
                db.session.add(relationship)
                db.session.commit()
                
                return {
                    'id': relationship.id,
                    'source_material_id': relationship.source_material_id,
                    'target_material_id': relationship.target_material_id,
                    'message': '物料关系创建成功'
                }, 201
            except Exception as e:
                db.session.rollback()
                return {'error': f'创建物料关系失败: {str(e)}'}, 500
    
    def put(self, relationship_id):
        """更新物料关系"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            relationship = db.session.get(MaterialRelationship, relationship_id)
            if not relationship:
                return {'error': '关系记录不存在'}, 404
            
            data = request.get_json()
            
            # 更新字段
            if 'relationship_type' in data:
                relationship.relationship_type = data['relationship_type']
            if 'quantity' in data:
                relationship.quantity = data['quantity']
            if 'description' in data:
                relationship.description = data['description']
            
            try:
                db.session.commit()
                return {
                    'id': relationship.id,
                    'source_material_id': relationship.source_material_id,
                    'target_material_id': relationship.target_material_id,
                    'relationship_type': relationship.relationship_type,
                    'quantity': relationship.quantity,
                    'description': relationship.description,
                    'message': '物料关系更新成功'
                }
            except Exception as e:
                db.session.rollback()
                return {'error': f'更新物料关系失败: {str(e)}'}, 500
    
    def delete(self, relationship_id):
        """删除物料关系"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            relationship = db.session.get(MaterialRelationship, relationship_id)
            if not relationship:
                return {'error': '关系记录不存在'}, 404
            
            try:
                db.session.delete(relationship)
                db.session.commit()
                return {'message': '物料关系删除成功'}
            except Exception as e:
                db.session.rollback()
                return {'error': f'删除物料关系失败: {str(e)}'}, 500
    
    def _serialize_relationship(self, relationship, db):
        """序列化关系对象"""
        return {
            'id': relationship.id,
            'source_material_id': relationship.source_material_id,
            'target_material_id': relationship.target_material_id,
            'relationship_type': relationship.relationship_type,
            'quantity': float(relationship.quantity),
            'description': relationship.description,
            'created_at': relationship.created_at.isoformat() if relationship.created_at else None
        }

class SerialNumberTraceResource(Resource):
    """序列号溯源API"""

    method_decorators = {'get': [require_material_perm('material:serial_manage')]}

    def get(self, serial_number):
        """获取序列号全链路轨迹"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            
            # 查找序列号
            serial_number_obj = db.session.query(SerialNumber).filter_by(serial_number=serial_number).first()
            if not serial_number_obj:
                return {'error': '序列号不存在'}, 404
            
            # 获取序列号基本信息
            material = db.session.get(Material, serial_number_obj.material_id)
            current_location = db.session.get(Location, serial_number_obj.current_location_id) if serial_number_obj.current_location_id else None
            
            # 获取交易记录
            transactions = db.session.query(InventoryTransaction).filter_by(
                serial_number=serial_number
            ).order_by(InventoryTransaction.created_at.asc()).all()
            
            # 获取关系记录
            relationships = []
            
            return {
                'serial_number': serial_number_obj.serial_number,
                'material_info': {
                    'material_code': material.code if material else '',
                    'material_name': material.name if material else '',
                    'specification': material.spec if material else '',
                    'brand': getattr(material, 'brand', '') if material else ''
                },
                'current_status': serial_number_obj.current_status,
                'current_location': current_location.name if current_location else '',
                'supplier_batch': serial_number_obj.supplier_batch,
                'purchase_date': serial_number_obj.purchase_date.isoformat() if serial_number_obj.purchase_date else None,
                'transaction_history': [self._serialize_transaction_for_trace(t, db) for t in transactions],
                'relationship_history': []
            }
    
    def _serialize_transaction_for_trace(self, transaction, db):
        """序列化交易记录用于溯源"""
        material = db.session.get(Material, transaction.material_id)
        from_warehouse = db.session.get(Warehouse, transaction.from_warehouse_id) if transaction.from_warehouse_id else None
        to_warehouse = db.session.get(Warehouse, transaction.to_warehouse_id) if transaction.to_warehouse_id else None
        
        return {
            'transaction_type': transaction.transaction_type,
            'transaction_type_cn': self._get_transaction_type_cn(transaction.transaction_type),
            'material_name': material.name if material else '',
            'from_warehouse': from_warehouse.name if from_warehouse else '',
            'to_warehouse': to_warehouse.name if to_warehouse else '',
            'quantity': float(transaction.quantity),
            'reference_id': transaction.reference_id,
            'remarks': transaction.remarks,
            'created_by': transaction.created_by,
            'created_at': transaction.created_at.isoformat() if transaction.created_at else None
        }
    
    def _serialize_relationship_for_trace(self, relationship, db):
        """序列化关系记录用于溯源"""
        return {
            'source_material_id': relationship.source_material_id,
            'target_material_id': relationship.target_material_id,
            'relationship_type': relationship.relationship_type,
            'quantity': float(relationship.quantity),
            'description': relationship.description,
            'created_at': relationship.created_at.isoformat() if relationship.created_at else None
        }
    
    def _get_transaction_type_cn(self, transaction_type):
        """获取交易类型中文描述"""
        type_map = {
            'in': '入库',
            'out': '出库',
            'transfer': '调拨'
        }
        return type_map.get(transaction_type, transaction_type)

class MaterialReportResource(Resource):
    """物料报表API"""

    method_decorators = {'get': [require_material_perm('material:view_reports')]}

    def get(self):
        """获取库存汇总报表"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            Material, MaterialCategory, Inventory, Warehouse, _, _, _, _ = get_models()
            
            # 获取查询参数
            category_id = request.args.get('category_id')
            warehouse_id = request.args.get('warehouse_id')
            
            # 构建查询
            query = db.session.query(
                Material.code.label('material_code'),
                Material.name,
                MaterialCategory.name.label('category_name'),
                Warehouse.name.label('warehouse_name'),
                func.sum(Inventory.quantity).label('total_quantity'),
                func.sum(Inventory.locked_quantity).label('total_locked'),
                func.sum(Inventory.quantity - Inventory.locked_quantity).label('available_quantity')
            ).join(
                MaterialCategory, Material.category_id == MaterialCategory.id
            ).join(
                Inventory, Material.id == Inventory.material_id
            ).join(
                Warehouse, Inventory.warehouse_id == Warehouse.id
            ).group_by(
                Material.code, Material.name, MaterialCategory.name, Warehouse.name
            )
            
            if category_id:
                query = query.filter(Material.category_id == category_id)
            if warehouse_id:
                query = query.filter(Inventory.warehouse_id == warehouse_id)
            
            results = query.all()
            
            return [{
                'material_code': row.material_code,
                'material_name': row.name,
                'category_name': row.category_name,
                'warehouse_name': row.warehouse_name,
                'total_quantity': float(row.total_quantity or 0),
                'locked_quantity': float(row.total_locked or 0),
                'available_quantity': float(row.available_quantity or 0)
            } for row in results]

class MaterialFlowReportResource(Resource):
    """物料流水报表API"""

    method_decorators = {'get': [require_material_perm('material:view_reports')]}

    def get(self):
        """获取物料流水报表"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            
            # 获取查询参数
            material_id = request.args.get('material_id')
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            transaction_type = request.args.get('transaction_type')
            
            # 构建查询
            query = db.session.query(InventoryTransaction)
            
            if material_id:
                query = query.filter(InventoryTransaction.material_id == material_id)
            if transaction_type:
                query = query.filter_by(transaction_type=transaction_type)
            if start_date:
                query = query.filter(InventoryTransaction.created_at >= start_date)
            if end_date:
                query = query.filter(InventoryTransaction.created_at <= end_date)
            
            # 按时间倒序排列
            transactions = query.order_by(InventoryTransaction.created_at.desc()).limit(500).all()
            
            return [self._serialize_transaction_for_report(t, db) for t in transactions]
    
    def _serialize_transaction_for_report(self, transaction, db):
        """序列化交易记录用于报表"""
        Material, _, _, Warehouse, _, _, _, _ = get_models()
        material = db.session.get(Material, transaction.material_id)
        from_warehouse = db.session.get(Warehouse, transaction.from_warehouse_id) if transaction.from_warehouse_id else None
        to_warehouse = db.session.get(Warehouse, transaction.to_warehouse_id) if transaction.to_warehouse_id else None
        
        # 处理serial_number字段，确保返回字符串而不是对象
        serial_number_str = ''
        if transaction.serial_number:
            # 如果serial_number是SerialNumber对象，获取其serial_number属性
            if hasattr(transaction.serial_number, 'serial_number'):
                serial_number_str = transaction.serial_number.serial_number
            else:
                serial_number_str = str(transaction.serial_number)
        
        return {
            'transaction_id': transaction.id,
            'transaction_type': transaction.transaction_type,
            'transaction_type_cn': self._get_transaction_type_cn(transaction.transaction_type),
            'material_code': material.code if material else '',
            'material_name': material.name if material else '',
            'serial_number': serial_number_str,
            'from_warehouse': from_warehouse.name if from_warehouse else '',
            'to_warehouse': to_warehouse.name if to_warehouse else '',
            'quantity': float(transaction.quantity),
            'reference_id': transaction.reference_id,
            'remarks': transaction.remarks,
            'created_by': transaction.created_by,
            'created_at': transaction.created_at.isoformat() if transaction.created_at else None
        }
    
    def _get_transaction_type_cn(self, transaction_type):
        """获取交易类型中文描述"""
        type_map = {
            'in': '入库',
            'out': '出库',
            'transfer': '调拨'
        }
        return type_map.get(transaction_type, transaction_type)

class MaterialAlertReportResource(Resource):
    """物料预警报表API"""

    method_decorators = {'get': [require_material_perm('material:view_reports')]}

    def get(self):
        """获取物料预警报表"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            
            # 获取查询参数
            alert_type = request.args.get('alert_type')  # low_stock, over_stock, stagnant
            warehouse_id = request.args.get('warehouse_id')
            category_id = request.args.get('category_id')
            
            # 低库存预警（库存数量低于安全库存）
            low_stock_alerts = db.session.query(
                Material.code.label('material_code'),
                Material.name,
                Material.safety_stock,
                Material.min_stock,
                Material.max_stock,
                func.sum(Inventory.quantity).label('current_stock'),
                Warehouse.name.label('warehouse_name'),
                MaterialCategory.name.label('category_name')
            ).join(
                Inventory, Material.id == Inventory.material_id
            ).join(
                Warehouse, Inventory.warehouse_id == Warehouse.id
            ).join(
                MaterialCategory, Material.category_id == MaterialCategory.id
            ).group_by(
                Material.id, Material.code, Material.name, Material.safety_stock, 
                Material.min_stock, Material.max_stock, Warehouse.name, MaterialCategory.name
            ).having(
                func.sum(Inventory.quantity) < Material.safety_stock
            )
            
            # 超储预警（库存数量超过最大库存）
            over_stock_alerts = db.session.query(
                Material.code.label('material_code'),
                Material.name,
                Material.safety_stock,
                Material.min_stock,
                Material.max_stock,
                func.sum(Inventory.quantity).label('current_stock'),
                Warehouse.name.label('warehouse_name'),
                MaterialCategory.name.label('category_name')
            ).join(
                Inventory, Material.id == Inventory.material_id
            ).join(
                Warehouse, Inventory.warehouse_id == Warehouse.id
            ).join(
                MaterialCategory, Material.category_id == MaterialCategory.id
            ).group_by(
                Material.id, Material.code, Material.name, Material.safety_stock, 
                Material.min_stock, Material.max_stock, Warehouse.name, MaterialCategory.name
            ).having(
                func.sum(Inventory.quantity) > Material.max_stock
            )
            
            # 呆滞料预警（长期未发生移动的物料）
            stagnant_alerts = db.session.query(
                Material.code.label('material_code'),
                Material.name,
                func.max(InventoryTransaction.created_at).label('last_transaction_date'),
                func.sum(Inventory.quantity).label('current_stock'),
                Warehouse.name.label('warehouse_name'),
                MaterialCategory.name.label('category_name')
            ).join(
                Inventory, Material.id == Inventory.material_id
            ).join(
                Warehouse, Inventory.warehouse_id == Warehouse.id
            ).join(
                MaterialCategory, Material.category_id == MaterialCategory.id
            ).outerjoin(
                InventoryTransaction, Material.id == InventoryTransaction.material_id
            ).group_by(
                Material.id, Material.code, Material.name, Warehouse.name, MaterialCategory.name
            ).having(
                (func.max(InventoryTransaction.created_at) < (datetime.now(timezone.utc) - timedelta(days=90))) |
                (func.max(InventoryTransaction.created_at).is_(None))
            )
            
            # 应用过滤条件
            if warehouse_id:
                low_stock_alerts = low_stock_alerts.filter(Inventory.warehouse_id == warehouse_id)
                over_stock_alerts = over_stock_alerts.filter(Inventory.warehouse_id == warehouse_id)
                stagnant_alerts = stagnant_alerts.filter(Inventory.warehouse_id == warehouse_id)
            
            if category_id:
                low_stock_alerts = low_stock_alerts.filter(Material.category_id == category_id)
                over_stock_alerts = over_stock_alerts.filter(Material.category_id == category_id)
                stagnant_alerts = stagnant_alerts.filter(Material.category_id == category_id)
            
            # 根据预警类型返回对应的数据
            if alert_type == 'low_stock':
                results = low_stock_alerts.all()
                alert_data = [self._serialize_low_stock_alert(row) for row in results]
            elif alert_type == 'over_stock':
                results = over_stock_alerts.all()
                alert_data = [self._serialize_over_stock_alert(row) for row in results]
            elif alert_type == 'stagnant':
                results = stagnant_alerts.all()
                alert_data = [self._serialize_stagnant_alert(row) for row in results]
            else:
                # 返回所有预警类型
                low_results = low_stock_alerts.all()
                over_results = over_stock_alerts.all()
                stagnant_results = stagnant_alerts.all()
                
                alert_data = {
                    'low_stock_alerts': [self._serialize_low_stock_alert(row) for row in low_results],
                    'over_stock_alerts': [self._serialize_over_stock_alert(row) for row in over_results],
                    'stagnant_alerts': [self._serialize_stagnant_alert(row) for row in stagnant_results]
                }
            
            return alert_data
    
    def _serialize_low_stock_alert(self, row):
        """序列化低库存预警数据"""
        return {
            'material_code': row.material_code,
            'material_name': row.name,
            'category_name': row.category_name,
            'warehouse_name': row.warehouse_name,
            'safety_stock': float(getattr(row, 'safety_stock', 0)),
            'min_stock': float(getattr(row, 'min_stock', 0)),
            'current_stock': float(row.current_stock or 0),
            'alert_type': '低库存预警',
            'severity': 'high' if float(row.current_stock or 0) <= 0 else 'medium'
        }
    
    def _serialize_over_stock_alert(self, row):
        """序列化超储预警数据"""
        return {
            'material_code': row.material_code,
            'material_name': row.name,
            'category_name': row.category_name,
            'warehouse_name': row.warehouse_name,
            'max_stock': float(getattr(row, 'max_stock', 0)),
            'current_stock': float(row.current_stock or 0),
            'alert_type': '超储预警',
            'severity': 'medium'
        }
    
    def _serialize_stagnant_alert(self, row):
        """序列化呆滞料预警数据"""
        return {
            'material_code': row.material_code,
            'material_name': row.name,
            'category_name': row.category_name,
            'warehouse_name': row.warehouse_name,
            'last_transaction_date': row.last_transaction_date.isoformat() if getattr(row, 'last_transaction_date', None) else '无交易记录',
            'current_stock': float(row.current_stock or 0),
            'alert_type': '呆滞料预警',
            'severity': 'low'
        }
    
    def _serialize_transaction_for_report(self, transaction, db):
        """序列化交易记录用于报表"""
        material = db.session.get(Material, transaction.material_id)
        from_warehouse = db.session.get(Warehouse, transaction.from_warehouse_id) if transaction.from_warehouse_id else None
        to_warehouse = db.session.get(Warehouse, transaction.to_warehouse_id) if transaction.to_warehouse_id else None
        
        return {
            'transaction_id': transaction.id,
            'transaction_type': transaction.transaction_type,
            'transaction_type_cn': self._get_transaction_type_cn(transaction.transaction_type),
            'material_code': material.code if material else '',
            'material_name': material.name if material else '',
            'serial_number': transaction.serial_number,
            'from_warehouse': from_warehouse.name if from_warehouse else '',
            'to_warehouse': to_warehouse.name if to_warehouse else '',
            'quantity': float(transaction.quantity),
            'reference_id': transaction.reference_id,
            'remarks': transaction.remarks,
            'created_by': transaction.created_by,
            'created_at': transaction.created_at.isoformat() if transaction.created_at else None
        }
    
    def _get_transaction_type_cn(self, transaction_type):
        """获取交易类型中文描述"""
        type_map = {
            'in': '入库',
            'out': '出库',
            'transfer': '调拨'
        }
        return type_map.get(transaction_type, transaction_type)

class InventoryCheckResource(Resource):
    """库存盘点API"""

    method_decorators = {
        'get': [require_material_perm('material:inventory_view')],
        'post': [require_material_perm('material:inventory_edit')],
        'put': [require_material_perm('material:inventory_edit')],
        'delete': [require_material_perm('material:inventory_edit')],
    }

    def get(self, check_id=None):
        """获取盘点记录"""
        from enhanced_app import app, InventoryCheck, InventoryCheckDetail
        with app.app_context():
            db = get_db()
            
            if check_id:
                # 获取单个盘点记录详情
                check = db.session.get(InventoryCheck, check_id)
                if not check:
                    return {'error': '盘点记录不存在'}, 404
                
                # 获取盘点明细
                details = db.session.query(InventoryCheckDetail).filter_by(check_id=check_id).all()
                
                return {
                    'check': check.to_dict(),
                    'details': [detail.to_dict() for detail in details]
                }
            else:
                # 获取盘点记录列表
                page = int(request.args.get('page', 1))
                per_page = int(request.args.get('per_page', 20))
                warehouse_id = request.args.get('warehouse_id')
                check_type = request.args.get('check_type')
                status = request.args.get('status')
                
                query = db.session.query(InventoryCheck)
                
                if warehouse_id:
                    query = query.filter_by(warehouse_id=warehouse_id)
                if check_type:
                    query = query.filter_by(check_type=check_type)
                if status:
                    query = query.filter_by(status=status)
                
                checks = query.order_by(InventoryCheck.created_at.desc()).paginate(
                    page=page, per_page=per_page, error_out=False
                )
                
                return {
                    'checks': [check.to_dict() for check in checks.items],
                    'total': checks.total,
                    'pages': checks.pages,
                    'current_page': page
                }
    
    def post(self):
        """创建盘点单"""
        from enhanced_app import app, InventoryCheck, Inventory, Material, Location
        with app.app_context():
            db = get_db()
            data = request.get_json()
            
            # 验证必填字段
            required_fields = ['warehouse_id', 'check_type', 'check_date']
            for field in required_fields:
                if field not in data or not data[field]:
                    return {'error': f'字段{field}是必填的'}, 400
            
            # 生成盘点单号
            check_number = f"CHK{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # 创建盘点单
            check = InventoryCheck(
                check_number=check_number,
                warehouse_id=data['warehouse_id'],
                check_type=data['check_type'],
                check_date=datetime.strptime(data['check_date'], '%Y-%m-%d').date(),
                created_by=data.get('created_by', 'system'),
                status='draft'
            )
            
            db.session.add(check)
            db.session.flush()  # 获取check的ID
            
            # 根据盘点类型生成盘点明细
            if data['check_type'] == 'full':
                # 全盘：盘点仓库中所有物料
                inventories = db.session.query(Inventory).filter_by(
                    warehouse_id=data['warehouse_id']
                ).all()
            else:
                # 循环盘点：只盘点部分物料（这里简化处理，实际可根据策略选择）
                inventories = db.session.query(Inventory).filter_by(
                    warehouse_id=data['warehouse_id']
                ).limit(50).all()
            
            # 创建盘点明细
            for inventory in inventories:
                detail = InventoryCheckDetail(
                    check_id=check.id,
                    material_id=inventory.material_id,
                    location_id=inventory.location_id,
                    system_quantity=inventory.quantity
                )
                db.session.add(detail)
            
            # 更新盘点单的总项数
            check.total_items = len(inventories)
            
            db.session.commit()
            
            return {
                'message': '盘点单创建成功',
                'check_id': check.id,
                'check_number': check.check_number,
                'total_items': check.total_items
            }, 201
    
    def put(self, check_id):
        """更新盘点单状态或明细"""
        from enhanced_app import app, InventoryCheck, InventoryCheckDetail
        with app.app_context():
            db = get_db()
            data = request.get_json()
            
            check = db.session.get(InventoryCheck, check_id)
            if not check:
                return {'error': '盘点记录不存在'}, 404
            
            # 更新盘点单状态
            if 'status' in data:
                check.status = data['status']
                if data['status'] == 'completed':
                    check.completed_at = datetime.utcnow()
            
            # 更新盘点明细
            if 'details' in data:
                for detail_data in data['details']:
                    detail = db.session.get(InventoryCheckDetail, detail_data['id'])
                    if detail and detail.check_id == check_id:
                        if 'actual_quantity' in detail_data:
                            detail.actual_quantity = detail_data['actual_quantity']
                            detail.difference = detail.actual_quantity - detail.system_quantity
                            detail.checked_by = detail_data.get('checked_by', 'system')
                            detail.checked_at = datetime.utcnow()
                            
                            # 更新已盘点项数
                            if detail.actual_quantity is not None and check.checked_items < check.total_items:
                                check.checked_items += 1
            
            db.session.commit()
            
            return {'message': '盘点单更新成功'}
    
    def delete(self, check_id):
        """删除盘点单"""
        from enhanced_app import app, InventoryCheck, InventoryCheckDetail
        with app.app_context():
            db = get_db()
            
            check = db.session.get(InventoryCheck, check_id)
            if not check:
                return {'error': '盘点记录不存在'}, 404
            
            # 只能删除草稿状态的盘点单
            if check.status != 'draft':
                return {'error': '只能删除草稿状态的盘点单'}, 400
            
            # 删除盘点明细
            db.session.query(InventoryCheckDetail).filter_by(check_id=check_id).delete()
            
            # 删除盘点单
            db.session.delete(check)
            db.session.commit()
            
            return {'message': '盘点单删除成功'}

class InventoryAlertResource(Resource):
    """库存预警API"""

    method_decorators = {'get': [require_material_perm('material:inventory_view')]}

    def get(self):
        """获取库存预警信息"""
        from enhanced_app import app
        with app.app_context():
            db = get_db()
            
            # 获取低库存预警（库存数量低于安全库存）
            low_stock_alerts = db.session.query(
                Material.code,
                Material.name,
                Material.safety_stock,
                func.sum(Inventory.quantity).label('current_stock'),
                Warehouse.name.label('warehouse_name')
            ).join(
                Inventory, Material.id == Inventory.material_id
            ).join(
                Warehouse, Inventory.warehouse_id == Warehouse.id
            ).group_by(
                Material.id, Material.code, Material.name, Material.safety_stock, Warehouse.name
            ).having(
                func.sum(Inventory.quantity) < Material.safety_stock
            ).all()
            
            # 获取呆滞料预警（长期未发生移动的物料）
            stagnant_materials = db.session.query(
                Material.code,
                Material.name,
                func.max(InventoryTransaction.created_at).label('last_transaction_date'),
                func.sum(Inventory.quantity).label('current_stock'),
                Warehouse.name.label('warehouse_name')
            ).join(
                Inventory, Material.id == Inventory.material_id
            ).join(
                Warehouse, Inventory.warehouse_id == Warehouse.id
            ).outerjoin(
                InventoryTransaction, Material.id == InventoryTransaction.material_id
            ).group_by(
                Material.id, Material.code, Material.name, Warehouse.name
            ).having(
                (func.max(InventoryTransaction.created_at) < (datetime.now(timezone.utc) - timedelta(days=90))) |
                (func.max(InventoryTransaction.created_at).is_(None))
            ).all()
            
            return {
                'low_stock_alerts': [{
                    'material_code': row.material_code,
                    'material_name': row.name,
                    'safety_stock': float(row.safety_stock),
                    'current_stock': float(row.current_stock or 0),
                    'warehouse_name': row.warehouse_name,
                    'alert_type': '低库存预警'
                } for row in low_stock_alerts],
                'stagnant_alerts': [{
                    'material_code': row.material_code,
                    'material_name': row.name,
                    'last_transaction_date': row.last_transaction_date.isoformat() if row.last_transaction_date else '无交易记录',
                    'current_stock': float(row.current_stock or 0),
                    'warehouse_name': row.warehouse_name,
                    'alert_type': '呆滞料预警'
                } for row in stagnant_materials]
            }

# 物料导入导出 API (F-004-08)
@materials_bp.route('/export', methods=['GET'])
@require_material_perm('material:export')
def export_materials():
    """导出物料数据为Excel"""
    from enhanced_app import app
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    
    with app.app_context():
        db = get_db()
        
        category_id = request.args.get('category_id', type=int)
        
        query = db.session.query(Material)
        if category_id:
            query = query.filter(Material.category_id == category_id)
        
        materials = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "物料数据"
        
        headers = ['物料编码', '物料名称', '规格型号', '单位', '分类', 
                   '安全库存', '供应商', '供应商编码', '单价', '状态', '创建时间']
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        category_map = {}
        categories = db.session.query(MaterialCategory).all()
        for cat in categories:
            category_map[cat.id] = cat.name
        
        for row_idx, m in enumerate(materials, 2):
            ws.cell(row=row_idx, column=1, value=m.code)
            ws.cell(row=row_idx, column=2, value=m.name)
            ws.cell(row=row_idx, column=3, value=m.spec or '')
            ws.cell(row=row_idx, column=4, value=m.unit or '')
            ws.cell(row=row_idx, column=5, value=category_map.get(m.category_id, ''))
            ws.cell(row=row_idx, column=6, value=m.safety_stock or 0)
            ws.cell(row=row_idx, column=7, value=m.supplier or '')
            ws.cell(row=row_idx, column=8, value=m.supplier_code or '')
            ws.cell(row=row_idx, column=9, value=float(m.unit_cost or 0))
            ws.cell(row=row_idx, column=10, value=m.status or 'active')
            ws.cell(row=row_idx, column=11, value=m.created_at.strftime('%Y-%m-%d %H:%M:%S') if m.created_at else '')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"materials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return output.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename={filename}'
        }


@materials_bp.route('/import', methods=['POST'])
@require_material_perm('material:import')
def import_materials():
    """导入物料数据（Excel格式）"""
    from enhanced_app import app
    from openpyxl import load_workbook
    
    with app.app_context():
        db = get_db()
        
        if 'file' not in request.files:
            return jsonify({'error': '请上传文件'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '只支持Excel文件格式(.xlsx, .xls)'}), 400
        
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            header_map = {
                '物料编码': 'material_code',
                '物料名称': 'name',
                '规格型号': 'specification',
                '单位': 'unit',
                '分类': 'category_name',
                '安全库存': 'safety_stock',
                '供应商': 'supplier',
                '供应商编码': 'supplier_code',
                '单价': 'unit_cost'
            }
            
            category_map = {}
            categories = db.session.query(MaterialCategory).all()
            for cat in categories:
                category_map[cat.name] = cat.id
            
            imported_count = 0
            error_rows = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                    
                try:
                    row_data = dict(zip(headers, row))
                    material_code = row_data.get('物料编码', '')
                    
                    if not material_code:
                        error_rows.append(f"第{row_num}行: 物料编码不能为空")
                        continue
                    
                    existing = db.session.query(Material).filter_by(code=material_code).first()
                    if existing:
                        error_rows.append(f"第{row_num}行: 物料编码 {material_code} 已存在")
                        continue
                    
                    category_name = row_data.get('分类', '')
                    category_id = category_map.get(category_name)
                    
                    material = Material(
                        code=material_code,
                        name=row_data.get('物料名称', ''),
                        spec=row_data.get('规格型号', ''),
                        unit=row_data.get('单位', '个') or '个',
                        category_id=category_id,
                        safety_stock=float(row_data.get('安全库存', 0) or 0),
                        supplier=row_data.get('供应商', ''),
                        supplier_code=row_data.get('供应商编码', ''),
                        unit_cost=float(row_data.get('单价', 0) or 0),
                        status='active'
                    )
                    
                    db.session.add(material)
                    imported_count += 1
                    
                except Exception as e:
                    error_rows.append(f"第{row_num}行: {str(e)}")
            
            db.session.commit()
            
            return jsonify({
                'message': f'成功导入 {imported_count} 条数据',
                'imported_count': imported_count,
                'errors': error_rows[:10] if len(error_rows) > 10 else error_rows
            }), 201
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'导入失败: {str(e)}'}), 500


# ==================== 物料统计报表 API (F-007-05) ====================

@materials_bp.route('/reports/inventory-value', methods=['GET'])
@require_material_perm('material:view_reports')
def get_inventory_value():
    """获取库存总价值统计"""
    from enhanced_app import app
    with app.app_context():
        db = get_db()
        Material, MaterialCategory, Inventory, Warehouse, _, _, _, _ = get_models()
        
        category_id = request.args.get('category_id', type=int)
        warehouse_id = request.args.get('warehouse_id', type=int)
        
        query = db.session.query(
            Material.id,
            Material.code,
            Material.name,
            MaterialCategory.name.label('category_name'),
            Warehouse.name.label('warehouse_name'),
            func.sum(Inventory.quantity).label('total_quantity'),
            Material.unit_cost
        ).join(
            MaterialCategory, Material.category_id == MaterialCategory.id
        ).join(
            Inventory, Material.id == Inventory.material_id
        ).join(
            Warehouse, Inventory.warehouse_id == Warehouse.id
        )
        
        if category_id:
            query = query.filter(Material.category_id == category_id)
        if warehouse_id:
            query = query.filter(Inventory.warehouse_id == warehouse_id)
        
        query = query.group_by(
            Material.id, Material.code, Material.name, MaterialCategory.name, Warehouse.name, Material.unit_cost
        )
        
        results = query.all()
        
        total_value = 0
        inventory_data = []
        
        for row in results:
            quantity = float(row.total_quantity or 0)
            unit_cost = float(row.unit_cost or 0)
            item_value = quantity * unit_cost
            total_value += item_value
            
            inventory_data.append({
                'material_id': row.id,
                'material_code': row.code,
                'material_name': row.name,
                'category_name': row.category_name,
                'warehouse_name': row.warehouse_name,
                'total_quantity': quantity,
                'unit_cost': unit_cost,
                'total_value': round(item_value, 2)
            })
        
        category_stats = {}
        for item in inventory_data:
            cat = item['category_name']
            if cat not in category_stats:
                category_stats[cat] = {'total_value': 0, 'count': 0}
            category_stats[cat]['total_value'] += item['total_value']
            category_stats[cat]['count'] += 1
        
        return jsonify({
            'total_value': round(total_value, 2),
            'inventory_data': inventory_data,
            'category_stats': [
                {'category': k, 'total_value': round(v['total_value'], 2), 'count': v['count']}
                for k, v in category_stats.items()
            ]
        })


@materials_bp.route('/reports/top-materials', methods=['GET'])
@require_material_perm('material:view_reports')
def get_top_materials():
    """获取出入库频率Top10物料"""
    from enhanced_app import app
    with app.app_context():
        db = get_db()
        Material, MaterialCategory, Inventory, Warehouse, InventoryTransaction, _, _, _ = get_models()
        
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        top_n = request.args.get('top_n', 10, type=int)
        warehouse_id = request.args.get('warehouse_id', type=int)
        
        query = db.session.query(
            Material.id,
            Material.code,
            Material.name,
            MaterialCategory.name.label('category_name'),
            func.count(InventoryTransaction.id).label('transaction_count'),
            func.sum(
                db.case(
                    (InventoryTransaction.transaction_type == 'in', InventoryTransaction.quantity),
                    else_=0
                )
            ).label('total_in'),
            func.sum(
                db.case(
                    (InventoryTransaction.transaction_type == 'out', InventoryTransaction.quantity),
                    else_=0
                )
            ).label('total_out')
        ).join(
            InventoryTransaction, Material.id == InventoryTransaction.material_id
        ).join(
            MaterialCategory, Material.category_id == MaterialCategory.id
        )
        
        if start_date:
            query = query.filter(InventoryTransaction.created_at >= start_date)
        if end_date:
            query = query.filter(InventoryTransaction.created_at <= end_date)
        if warehouse_id:
            query = query.filter(
                (InventoryTransaction.from_warehouse_id == warehouse_id) |
                (InventoryTransaction.to_warehouse_id == warehouse_id)
            )
        
        query = query.group_by(
            Material.id, Material.code, Material.name, MaterialCategory.name
        ).order_by(
            func.count(InventoryTransaction.id).desc()
        ).limit(top_n)
        
        results = query.all()
        
        top_materials = []
        for row in results:
            top_materials.append({
                'material_id': row.id,
                'material_code': row.code,
                'material_name': row.name,
                'category_name': row.category_name,
                'transaction_count': row.transaction_count,
                'total_in': float(row.total_in or 0),
                'total_out': float(row.total_out or 0)
            })
        
        return jsonify({
            'top_materials': top_materials,
            'start_date': start_date,
            'end_date': end_date
        })


@materials_bp.route('/reports/material-turnover', methods=['GET'])
@require_material_perm('material:view_reports')
def get_material_turnover():
    """获取物料周转率统计"""
    from enhanced_app import app
    with app.app_context():
        db = get_db()
        Material, MaterialCategory, Inventory, InventoryTransaction, _, _, _, _ = get_models()
        
        days = request.args.get('days', 30, type=int)
        category_id = request.args.get('category_id', type=int)
        
        start_date = datetime.utcnow() - timedelta(days=days)
        
        query = db.session.query(
            Material.id,
            Material.code,
            Material.name,
            MaterialCategory.name.label('category_name'),
            func.sum(Inventory.quantity).label('current_stock'),
            func.sum(
                db.case(
                    (InventoryTransaction.transaction_type == 'out', InventoryTransaction.quantity),
                    else_=0
                )
            ).label('total_out')
        ).join(
            Inventory, Material.id == Inventory.material_id
        ).outerjoin(
            InventoryTransaction, 
            db.and_(
                Material.id == InventoryTransaction.material_id,
                InventoryTransaction.created_at >= start_date
            )
        ).join(
            MaterialCategory, Material.category_id == MaterialCategory.id
        )
        
        if category_id:
            query = query.filter(Material.category_id == category_id)
        
        query = query.group_by(
            Material.id, Material.code, Material.name, MaterialCategory.name
        )
        
        results = query.all()
        
        turnover_data = []
        for row in results:
            current_stock = float(row.current_stock or 0)
            total_out = float(row.total_out or 0)
            turnover_rate = (total_out / current_stock * 100) if current_stock > 0 else 0
            
            turnover_data.append({
                'material_id': row.id,
                'material_code': row.code,
                'material_name': row.name,
                'category_name': row.category_name,
                'current_stock': current_stock,
                'total_out': total_out,
                'turnover_rate': round(turnover_rate, 2)
            })
        
        turnover_data.sort(key=lambda x: x['turnover_rate'], reverse=True)
        
        return jsonify({
            'turnover_data': turnover_data[:50],
            'period_days': days
        })


# 注册API路由
materials_api.add_resource(MaterialCategoryResource,
                          '/categories',
                          '/categories/<int:category_id>')
materials_api.add_resource(MaterialResource,
                          '/',
                          '/<int:material_id>')
materials_api.add_resource(WarehouseResource, 
                          '/warehouses', 
                          '/warehouses/<int:warehouse_id>')
materials_api.add_resource(LocationResource, 
                          '/locations', 
                          '/locations/<int:location_id>')
materials_api.add_resource(InventoryResource, 
                          '/inventory')
materials_api.add_resource(InventoryStatsResource,
                          '/inventory/stats')
materials_api.add_resource(InventoryTransactionResource, 
                          '/transactions', 
                          '/transactions/<int:transaction_id>')
materials_api.add_resource(SerialNumberResource, 
                          '/serial-numbers', 
                          '/serial-numbers/<int:serial_number_id>')
materials_api.add_resource(MaterialRelationshipResource, 
                          '/relationships', 
                          '/relationships/<int:relationship_id>')
materials_api.add_resource(SerialNumberTraceResource, 
                          '/serial-trace/<string:serial_number>')
materials_api.add_resource(MaterialReportResource, 
                          '/reports/inventory-summary')
materials_api.add_resource(MaterialFlowReportResource, 
                          '/reports/material-flow')
materials_api.add_resource(MaterialAlertReportResource, 
                          '/reports/material-alerts')
materials_api.add_resource(InventoryCheckResource, 
                          '/inventory-checks', 
                          '/inventory-checks/<int:check_id>')
materials_api.add_resource(InventoryAlertResource, 
                          '/alerts')
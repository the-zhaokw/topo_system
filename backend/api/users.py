from flask import Blueprint, request, jsonify, make_response
from werkzeug.security import generate_password_hash
from flask_jwt_extended import jwt_required, get_jwt_identity
import re
import logging
from datetime import datetime
from logging_config import get_log_manager
from logging_decorators import log_api_call, log_business_operation

logger = logging.getLogger(__name__)

users_bp = Blueprint('users', __name__, url_prefix='/users')

@users_bp.after_request
def add_cors_headers(response):
    """为所有响应添加CORS头"""
    origin = request.headers.get('Origin')
    if origin:
        response.headers['Access-Control-Allow-Origin'] = origin
    else:
        response.headers['Access-Control-Allow-Origin'] = 'http://localhost:3000'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With'
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response

# 延迟导入数据库和装饰器，模型延迟导入以避免循环导入
def get_db():
    """延迟获取数据库实例，避免循环导入"""
    from enhanced_app import get_db_instance
    return get_db_instance()

def get_models():
    """延迟获取数据库模型"""
    from enhanced_app import User, UserRole, Department, Position
    return User, UserRole, Department, Position

def get_create_audit_log():
    from enhanced_app import create_audit_log
    return create_audit_log

def get_db_and_models():
    """获取数据库实例和模型，避免循环导入"""
    from enhanced_app import db, User, UserRole, create_audit_log
    return db, User, UserRole, create_audit_log

def get_db_and_models_with_notification():
    """获取数据库实例和模型，包含Notification模型"""
    from enhanced_app import db, User, UserRole, create_audit_log, Notification
    return db, User, UserRole, create_audit_log, Notification

# 获取用户列表
@users_bp.route('/', methods=['GET'])
@log_api_call
@log_business_operation
@jwt_required()
def get_users():
    
    # 延迟导入数据库模型
    db = get_db()
    User, UserRole, Department, Position = get_models()
    
    # 获取当前用户ID
    current_user_id = get_jwt_identity()
    
    # 获取当前用户
    current_user = User.query.get(current_user_id)
    if not current_user:
        log_manager = get_log_manager()
        log_manager.log_error('get_users_user_not_found', f"用户不存在: {current_user_id}")
        return jsonify({'error': '用户不存在'}), 404
    
    # 权限检查：所有已登录用户都可以查看用户列表（用于选择指派人等场景）
    # 但普通用户只能看到基本用户信息（不包含敏感字段）
    position_info = current_user.get_position_info()
    is_admin = current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))

    # 使用增强日志系统记录请求
    try:
        log_manager = get_log_manager()
        log_manager.log_request(
            user_id=current_user_id,
            action='get_users_request',
            details={
                'page': request.args.get('page', 1),
                'per_page': request.args.get('per_page', 20),
                'department': request.args.get('department'),
                'search': request.args.get('search', '').strip()
            }
        )
    except Exception as e:
        logger.error(f"获取用户列表请求日志记录失败: {str(e)}")

    # 获取查询参数
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    department = request.args.get('department')
    search = request.args.get('search', '').strip()

    # 构建查询
    query = User.query

    # 应用筛选条件
    if department:
        query = query.filter_by(department=department)

    if search:
        query = query.filter(
            (User.username.contains(search)) |
            (User.email.contains(search)) |
            (User.first_name.contains(search)) |
            (User.last_name.contains(search))
        )
    
    # 分页查询
    pagination = query.order_by(User.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # 构建响应数据
    users_data = []
    for user in pagination.items:
        # 普通用户只能看到基本信息，管理员可以看到完整信息
        if is_admin:
            users_data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'position': user.position,
                'department': user.department,
                'is_active': user.is_active,
                'created_at': user.created_at.isoformat() if user.created_at else None,
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'employee_id': getattr(user, 'employee_id', ''),
                'company_phone': getattr(user, 'company_phone', ''),
                'phone': getattr(user, 'mobile_phone', ''),
                'mobile_phone': getattr(user, 'mobile_phone', ''),
                'birthday': user.birthday.strftime('%Y-%m-%d') if user.birthday else '',
                'gender': getattr(user, 'gender', ''),
                'work_language': getattr(user, 'work_language', '')
            })
        else:
            # 普通用户只能看到基本用户信息（用于选择指派人等场景）
            users_data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'position': user.position,
                'department': user.department,
                'is_active': user.is_active
            })
    
    # 使用增强日志系统记录业务操作
    try:
        log_manager = get_log_manager()
        log_manager.log_business(
            operation='users_listed',
            user_id=current_user_id,
            details={
                'total_users': pagination.total,
                'page': pagination.page,
                'per_page': pagination.per_page,
                'filtered_by_department': department,
                'search_term': search
            }
        )
    except Exception as e:
        logger.error(f"获取用户列表业务日志记录失败: {str(e)}")

    return jsonify({
        'users': users_data,
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
        'per_page': pagination.per_page
    }), 200

# 搜索用户
@users_bp.route('/search', methods=['GET'])
@jwt_required()
def search_users():

    # 延迟导入数据库模型
    db = get_db()
    User, UserRole, Department, Position = get_models()

    # 获取当前用户ID
    current_user_id = get_jwt_identity()

    # 获取当前用户
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    # 权限检查：只有管理员可以搜索用户
    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403
    
    query = request.args.get('q', '').strip()
    
    if not query or len(query) < 2:
        return jsonify({'users': []}), 200
    
    # 搜索用户
    users = User.query.filter(
        (User.username.contains(query)) |
        (User.email.contains(query)) |
        (User.first_name.contains(query)) |
        (User.last_name.contains(query))
    ).limit(10).all()
    
    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'position': user.position,
            'department': user.department
        })
    
    return jsonify({'users': users_data}), 200

# 获取审批人列表（所有登录用户可访问，用于请假审批人选择）
@users_bp.route('/approvers', methods=['GET'])
@jwt_required()
def get_approvers():
    
    # 延迟导入数据库模型
    db = get_db()
    User, UserRole, Department, Position = get_models()
    
    # 获取当前用户ID
    current_user_id = get_jwt_identity()
    
    # 获取当前用户
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404
    
    # 获取所有活跃用户作为审批人选项
    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    
    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'position': user.position,
            'department': user.department
        })
    
    return jsonify({'users': users_data, 'total': len(users_data)}), 200

# 获取用户详情
@users_bp.route('/<int:user_id>', methods=['GET'])
@jwt_required()
def get_user(user_id):
    
    # 延迟导入数据库模型
    db, User, UserRole, create_audit_log = get_db_and_models()
    
    current_user_id = get_jwt_identity()
    
    # 获取当前用户
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    # 权限检查：只能查看自己的信息或管理员可以查看所有
    position_info = current_user.get_position_info()
    is_admin = current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))
    if current_user_id != user_id and not is_admin:
        return jsonify({'error': '权限不足'}), 403

    # 获取目标用户
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    user_data = {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'position': user.position,
        'department': user.department,
        'is_active': user.is_active,
        'created_at': user.created_at.isoformat() if user.created_at else None,
        'last_login': user.last_login.isoformat() if user.last_login else None,
        'avatar': getattr(user, 'avatar', ''),
        'email_notification_enabled': getattr(user, 'email_notification_enabled', True),
        'email_on_bug_assigned': getattr(user, 'email_on_bug_assigned', True),
        'email_on_bug_closed': getattr(user, 'email_on_bug_closed', True),
        'employee_id': getattr(user, 'employee_id', ''),
        'company_phone': getattr(user, 'company_phone', ''),
        'phone': getattr(user, 'mobile_phone', ''),
        'mobile_phone': getattr(user, 'mobile_phone', ''),
        'birthday': user.birthday.strftime('%Y-%m-%d') if user.birthday else '',
        'gender': getattr(user, 'gender', ''),
        'work_language': getattr(user, 'work_language', '')
    }

    return jsonify(user_data), 200

# 获取用户主页信息（包含活动记录）
@users_bp.route('/<int:user_id>/home', methods=['GET'])
@jwt_required()
def get_user_home(user_id):
    from datetime import timedelta

    db, User, UserRole, create_audit_log = get_db_and_models()
    Activity, Bug, Project, Task = get_activity_models()

    current_user_id = get_jwt_identity()

    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    is_admin = current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))
    is_self = current_user_id == user_id

    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    user_data = {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'position': user.position,
        'department': user.department,
        'is_active': user.is_active,
        'created_at': user.created_at.isoformat() if user.created_at else None,
        'last_login': user.last_login.isoformat() if user.last_login else None,
        'avatar': getattr(user, 'avatar', ''),
        'employee_id': getattr(user, 'employee_id', ''),
        'company_phone': getattr(user, 'company_phone', ''),
        'phone': getattr(user, 'mobile_phone', ''),
        'mobile_phone': getattr(user, 'mobile_phone', ''),
        'birthday': user.birthday.strftime('%Y-%m-%d') if user.birthday else '',
        'gender': getattr(user, 'gender', ''),
        'work_language': getattr(user, 'work_language', '')
    }

    response_data = {
        'user': user_data,
        'is_admin_view': is_admin or is_self,
        'is_self': is_self
    }
    
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    activities = Activity.query.filter(
        Activity.performed_by == user_id,
        Activity.created_at >= thirty_days_ago
    ).order_by(Activity.created_at.desc()).limit(50).all()
    
    activity_list = []
    for activity in activities:
        activity_dict = {
            'id': activity.id,
            'action': activity.action,
            'description': activity.description,
            'target_type': activity.target_type,
            'target_id': activity.target_id,
            'created_at': activity.created_at.isoformat() if activity.created_at else None
        }
        
        if activity.target_type == 'bug':
            bug = Bug.query.get(activity.target_id)
            activity_dict['resource_name'] = bug.title if bug else '未知Bug'
        elif activity.target_type == 'project':
            project = Project.query.get(activity.target_id)
            activity_dict['resource_name'] = project.name if project else '未知项目'
        elif activity.target_type == 'task':
            task = Task.query.get(activity.target_id)
            activity_dict['resource_name'] = task.title if task else '未知任务'
        else:
            activity_dict['resource_name'] = f'{activity.target_type} #{activity.target_id}'
        
        activity_list.append(activity_dict)
    
    response_data['activities'] = activity_list
    response_data['activity_count'] = Activity.query.filter_by(performed_by=user_id).count()
    
    if is_admin or is_self:
        response_data['statistics'] = {
            'total_bugs': Bug.query.filter_by(assigned_to=user_id).count() if hasattr(Bug, 'assigned_to') else 0,
            'total_projects': Project.query.filter(
                (Project.owner_id == user_id) | (Project.manager_id == user_id)
            ).count() if hasattr(Project, 'owner_id') else 0,
            'total_tasks': Task.query.filter_by(assigned_to=user_id).count() if hasattr(Task, 'assigned_to') else 0
        }
    else:
        response_data['statistics'] = {}
    
    return jsonify(response_data), 200

def get_activity_models():
    """获取活动相关模型"""
    from enhanced_app import Activity, Bug, Project, Task
    return Activity, Bug, Project, Task

# 创建用户（管理员专用）
@users_bp.route('/', methods=['POST'])
@log_api_call
@log_business_operation
@jwt_required()
def create_user():
    # 延迟导入数据库模型
    db, User, UserRole, create_audit_log = get_db_and_models()
    
    data = request.get_json()
    
    # 获取当前用户ID
    current_user_id = get_jwt_identity()

    # 获取当前用户
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    # 权限检查：只有管理员可以创建用户
    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    # 使用增强日志系统记录请求
    try:
        log_manager = get_log_manager()
        log_manager.log_request(
            user_id=current_user_id,
            action='create_user_request',
            details={
                'username': data.get('username'),
                'email': data.get('email'),
                'department': data.get('department')
            }
        )
    except Exception as e:
        logger.error(f"创建用户请求日志记录失败: {str(e)}")
    
    # 验证必填字段
    required_fields = ['username', 'email', 'password']
    for field in required_fields:
        if not data.get(field):
            log_manager = get_log_manager()
            log_manager.log_error('create_user_validation_error', f"缺少必填字段: {field}")
            return jsonify({'error': f'缺少必填字段: {field}'}), 400
    
    # 验证用户名格式
    if not re.match(r'^[a-zA-Z0-9_]{3,20}$', data['username']):
        log_manager = get_log_manager()
        log_manager.log_error('create_user_validation_error', f"用户名格式不正确: {data['username']}")
        return jsonify({'error': '用户名只能包含字母、数字和下划线，长度3-20位'}), 400
    
    # 验证邮箱格式
    if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', data['email']):
        log_manager = get_log_manager()
        log_manager.log_error('create_user_validation_error', f"邮箱格式不正确: {data['email']}")
        return jsonify({'error': '邮箱格式不正确'}), 400
    
    # 职位处理 - 如果职位不存在则自动创建
    position = data.get('position', '')
    if position:
        position_obj = Position.query.filter_by(name=position).first()
        if not position_obj:
            try:
                # 判断职位是否为管理员或经理
                admin_position_names = ['admin', 'manager', 'administrator', '系统管理员', '经理', '管理员', '部门经理']
                manager_position_names = ['manager', 'project_manager', 'department_manager', '经理', '项目经理', '部门经理', '主管']
                position_lower = position.lower()
                is_admin = any(admin_name in position_lower or position_lower in admin_name
                              for admin_name in admin_position_names)
                is_manager = any(manager_name in position_lower or position_lower in manager_name
                                for manager_name in manager_position_names)

                new_pos = Position(name=position, is_admin=is_admin, is_manager=is_manager, permissions='[]')
                db.session.add(new_pos)
                db.session.commit()
                log_manager = get_log_manager()
                log_manager.log_business(
                    operation='position_auto_created',
                    user_id=current_user_id,
                    details={'position': position, 'is_admin': is_admin, 'is_manager': is_manager}
                )
            except Exception as e:
                db.session.rollback()
                log_manager = get_log_manager()
                log_manager.log_error('create_position_error', f"自动创建职位失败: {str(e)}")
    
    # 检查用户名是否已存在
    if User.query.filter_by(username=data['username']).first():
        log_manager = get_log_manager()
        log_manager.log_error('create_user_validation_error', f"用户名已存在: {data['username']}")
        return jsonify({'error': '用户名已存在'}), 400
    
    # 检查邮箱是否已存在
    if User.query.filter_by(email=data['email']).first():
        log_manager = get_log_manager()
        log_manager.log_error('create_user_validation_error', f"邮箱已存在: {data['email']}")
        return jsonify({'error': '邮箱已存在'}), 400
    
    # 创建用户
    # 生成随机salt值
    import secrets
    salt = secrets.token_hex(16)
    
    user = User(
        username=data['username'],
        email=data['email'],
        password_hash=generate_password_hash(data['password']),
        salt=salt,  # 添加salt值
        first_name=data.get('first_name', ''),
        last_name=data.get('last_name', ''),
        position=data.get('position', ''),
        department=data.get('department', ''),
        is_active=data.get('is_active', True)
    )
    
    try:
        db.session.add(user)
        db.session.commit()
        
        # 使用增强日志系统记录业务操作
        log_manager.log_business(
            operation='user_created',
            user_id=current_user_id,
            details={
                'username': user.username,
                'email': user.email,
                'department': user.department,
                'position': user.position
            }
        )
        
        # 创建log
        create_audit_log(
            user_id=current_user_id,
            action='create_user',
            resource_type='user',
            resource_id=user.id,
            details={'username': user.username, 'position': user.position}
        )
        
        return jsonify({
            'message': '用户创建成功',
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'position': user.position,
                'department': user.department,
                'is_active': user.is_active
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        log_manager = get_log_manager()
        log_manager.log_error('create_user_database_error', f"创建用户失败: {str(e)}")
        return jsonify({'error': '创建用户失败'}), 500

# 更新用户信息
@users_bp.route('/<int:user_id>', methods=['PUT'])
@log_api_call
@log_business_operation
@jwt_required()
def update_user(user_id):
    # 延迟导入数据库模型
    db, User, UserRole, create_audit_log = get_db_and_models()
    
    data = request.get_json()
    
    # 获取当前用户ID
    current_user_id = get_jwt_identity()
    
    # 记录请求日志
    try:
        log_manager = get_log_manager()
        log_manager.log_request(
            user_id=current_user_id,
            action='update_user',
            resource_type='user',
            resource_id=user_id,
            details={
                'updated_fields': list(data.keys()),
                'email': data.get('email'),
                'position': data.get('position'),
                'is_active': data.get('is_active')
            }
        )
    except Exception as e:
        logger.error(f"更新用户请求日志记录失败: {str(e)}")
    
    user = User.query.get(user_id)
    if not user:
        try:
            log_manager = get_log_manager()
            log_manager.log_error('update_user_not_found', f"用户不存在: {user_id}")
        except Exception as e:
            logger.error(f"记录用户不存在日志失败: {str(e)}")
        return jsonify({'error': '用户不存在'}), 404

    # 防止修改管理员账号
    current_user = User.query.get(current_user_id)
    user_position = user.get_position_info()
    if user_position and user_position.is_admin and current_user.id != user.id and not current_user.is_super_admin:
        try:
            log_manager = get_log_manager()
            log_manager.audit(
                user_id=current_user_id,
                action='update_user_admin_denied',
                resource_type='user',
                resource_id=user_id,
                details={'reason': '无法修改其他管理员账号'}
            )
        except Exception as e:
            logger.error(f"记录管理员权限拒绝日志失败: {str(e)}")
        return jsonify({'error': '无法修改其他管理员账号'}), 403
    
    # 职位更新处理 - 优先使用 position 字段
    if 'position' in data and data['position']:
        user.position = data['position']
    
    # 更新用户信息
    update_fields = [
        'first_name', 'last_name', 'position', 'department',
        'is_active', 'email_notification_enabled',
        'email_on_bug_assigned', 'email_on_bug_closed',
        'company_phone', 'birthday', 'gender', 'work_language',
        'employee_id', 'mobile_phone'
    ]
    
    for field in update_fields:
        if field in data:
            if field == 'birthday':
                if data[field] is None or data[field] == '':
                    user.birthday = None
                else:
                    try:
                        if isinstance(data[field], str):
                            user.birthday = datetime.strptime(data[field], '%Y-%m-%d')
                        else:
                            user.birthday = data[field]
                    except ValueError:
                        try:
                            log_manager = get_log_manager()
                            log_manager.log_error('update_user_validation_error', f"生日格式不正确: {data[field]}")
                        except Exception as e:
                            logger.error(f"记录生日格式错误日志失败: {str(e)}")
                        return jsonify({'error': '生日格式不正确，请使用 YYYY-MM-DD 格式'}), 400
            else:
                setattr(user, field, data[field])
    
    # 处理 phone 字段映射到 mobile_phone
    if 'phone' in data and data['phone'] is not None:
        user.mobile_phone = data['phone']

    # 用户名更新处理
    if 'username' in data and data['username']:
        new_username = data['username']
        if new_username != user.username:
            if not re.match(r'^[a-zA-Z0-9_]{3,20}$', new_username):
                try:
                    log_manager = get_log_manager()
                    log_manager.log_error('update_user_validation_error', f"用户名格式不正确: {new_username}")
                except Exception as e:
                    logger.error(f"记录用户名格式错误日志失败: {str(e)}")
                return jsonify({'error': '用户名只能包含字母、数字和下划线，长度3-20位'}), 400

            existing_user = User.query.filter_by(username=new_username).first()
            if existing_user and existing_user.id != user_id:
                try:
                    log_manager = get_log_manager()
                    log_manager.log_error('update_user_validation_error', f"用户名已被其他用户使用: {new_username}")
                except Exception as e:
                    logger.error(f"记录用户名重复错误日志失败: {str(e)}")
                return jsonify({'error': '用户名已被其他用户使用'}), 400

            user.username = new_username

    # 特殊字段处理
    if 'email' in data and data['email']:
        # 验证邮箱格式
        if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', data['email']):
            try:
                log_manager = get_log_manager()
                log_manager.log_error('update_user_validation_error', f"邮箱格式不正确: {data['email']}")
            except Exception as e:
                logger.error(f"记录邮箱格式错误日志失败: {str(e)}")
            return jsonify({'error': '邮箱格式不正确'}), 400
        
        # 检查邮箱是否已被其他用户使用
        existing_user = User.query.filter_by(email=data['email']).first()
        if existing_user and existing_user.id != user_id:
            try:
                log_manager = get_log_manager()
                log_manager.log_error('update_user_validation_error', f"邮箱已被其他用户使用: {data['email']}")
            except Exception as e:
                logger.error(f"记录邮箱重复错误日志失败: {str(e)}")
            return jsonify({'error': '邮箱已被其他用户使用'}), 400
        
        user.email = data['email']
    
    if 'password' in data and data['password']:
        if len(data['password']) < 6:
            try:
                log_manager = get_log_manager()
                log_manager.log_error('update_user_validation_error', f"密码长度不足: {len(data['password'])}位")
            except Exception as e:
                logger.error(f"记录密码长度错误日志失败: {str(e)}")
            return jsonify({'error': '密码长度至少为6个字符'}), 400
        user.password_hash = generate_password_hash(data['password'])
    
    try:
        db.session.commit()
        
        # 使用增强日志系统记录业务操作
        log_manager = get_log_manager()
        log_manager.log_business(
            operation='user_updated',
            user_id=current_user_id,
            details={
                'username': user.username,
                'updated_fields': list(data.keys()),
                'email': user.email,
                'position': user.position,
                'is_active': user.is_active
            }
        )
        
        # 创建log
        create_audit_log(
            user_id=current_user_id,
            action='update_user',
            resource_type='user',
            resource_id=user_id,
            details={'username': user.username, 'updated_fields': list(data.keys())}
        )
        
        return jsonify({
            'message': '用户信息更新成功',
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'position': user.position,
                'department': user.department,
                'is_active': user.is_active,
                'email_notification_enabled': getattr(user, 'email_notification_enabled', True),
                'email_on_bug_assigned': getattr(user, 'email_on_bug_assigned', True),
                'email_on_bug_closed': getattr(user, 'email_on_bug_closed', True)
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        try:
            log_manager = get_log_manager()
            log_manager.log_error('update_user_database_error', f"更新用户信息失败: {str(e)}")
        except Exception as log_e:
            logger.error(f"记录数据库错误日志失败: {str(log_e)}")
        return jsonify({'error': '更新用户信息失败'}), 500

# 删除用户
@users_bp.route('/<int:user_id>', methods=['DELETE'])
@log_api_call
@log_business_operation
@jwt_required()
def delete_user(user_id):
    # 延迟导入数据库模型
    db, User, UserRole, create_audit_log, Notification = get_db_and_models_with_notification()
    
    # 记录请求日志
    try:
        current_user_id = get_jwt_identity()
        log_manager = get_log_manager()
        log_manager.log_request(
            user_id=current_user_id,
            action='delete_user',
            resource_type='user',
            resource_id=user_id,
            details={'action': 'user_deletion'}
        )
    except Exception as e:
        logger.error(f"删除用户请求日志记录失败: {str(e)}")
    
    user = User.query.get(user_id)
    if not user:
        try:
            log_manager = get_log_manager()
            log_manager.log_error('delete_user_not_found', f"用户不存在: {user_id}")
        except Exception as e:
            logger.error(f"记录用户不存在日志失败: {str(e)}")
        return jsonify({'error': '用户不存在'}), 404

    # 防止删除管理员账号
    current_user = User.query.get(current_user_id)

    user_position = user.get_position_info()
    if user_position and user_position.is_admin:
        if not current_user.is_super_admin:
            try:
                log_manager = get_log_manager()
                log_manager.log_audit(
                    user_id=current_user_id,
                    action='delete_user_admin_denied',
                    resource_type='user',
                    resource_id=user_id,
                    details={'reason': '权限不足，无法删除管理员账号'}
                )
            except Exception as e:
                logger.error(f"记录删除管理员权限拒绝日志失败: {str(e)}")
            return jsonify({'error': '权限不足，无法删除管理员账号'}), 403
        elif current_user.id == user_id:
            try:
                log_manager = get_log_manager()
                log_manager.log_audit(
                    user_id=current_user_id,
                    action='delete_user_self_admin_denied',
                    resource_type='user',
                    resource_id=user_id,
                    details={'reason': '无法删除自己的管理员账号'}
                )
            except Exception as e:
                logger.error(f"记录删除自己管理员账号拒绝日志失败: {str(e)}")
            return jsonify({'error': '无法删除自己的管理员账号'}), 403
    
    try:
        Notification.query.filter_by(user_id=user_id).delete()
        db.session.delete(user)
        db.session.commit()
        
        # 使用增强日志系统记录业务操作
        log_manager = get_log_manager()
        log_manager.log_business(
            operation='user_deleted',
            user_id=current_user_id,
            details={
                'username': user.username,
                'email': user.email,
                'department': user.department,
                'position': user.position
            }
        )
        
        # 创建log
        create_audit_log(
            user_id=current_user_id,
            action='delete_user',
            resource_type='user',
            resource_id=user_id,
            details={'username': user.username}
        )
        
        return jsonify({'message': '用户删除成功'}), 200
    except Exception as e:
        db.session.rollback()
        try:
            log_manager = get_log_manager()
            log_manager.log_error('delete_user_database_error', f"删除用户失败: {str(e)}")
        except Exception as log_e:
            logger.error(f"记录删除用户数据库错误日志失败: {str(log_e)}")
        return jsonify({'error': '删除用户失败'}), 500

# 重置用户密码
@users_bp.route('/<int:user_id>/reset-password', methods=['POST'])
@log_api_call
@jwt_required()
def reset_user_password(user_id):
    db, User, UserRole, create_audit_log = get_db_and_models()

    current_user_id = get_jwt_identity()

    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    try:
        user.password_hash = generate_password_hash('123456')
        db.session.commit()

        log_manager = get_log_manager()
        log_manager.log_business(
            operation='user_password_reset',
            user_id=current_user_id,
            details={
                'reset_user_id': user_id,
                'username': user.username
            }
        )

        create_audit_log(
            user_id=current_user_id,
            action='reset_user_password',
            resource_type='user',
            resource_id=user_id,
            details={'username': user.username}
        )

        return jsonify({'success': True, 'message': '密码重置成功，新密码为：123456'}), 200
    except Exception as e:
        db.session.rollback()
        try:
            log_manager = get_log_manager()
            log_manager.log_error('reset_password_error', f"重置用户密码失败: {str(e)}")
        except Exception:
            pass
        return jsonify({'error': '重置密码失败'}), 500

# 批量删除用户
@users_bp.route('/batch-delete', methods=['POST'])
@jwt_required()
def batch_delete_users():
    # 延迟导入数据库模型
    db, User, UserRole, create_audit_log = get_db_and_models()
    
    data = request.get_json()
    
    if not data or 'user_ids' not in data or not isinstance(data['user_ids'], list):
        return jsonify({'error': '请提供有效的用户ID列表'}), 400
    
    user_ids = data['user_ids']

    # 检查是否包含管理员ID
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    is_admin = current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))
    if not is_admin:
        return jsonify({'error': '权限不足，无法批量删除用户'}), 403

    admin_users = []
    for user_id in user_ids:
        user = User.query.get(user_id)
        if user:
            user_pos = user.get_position_info()
            if user_pos and user_pos.is_admin:
                admin_users.append(user.username)

    if admin_users:
        return jsonify({'error': f'无法删除管理员账号: {", ".join(admin_users)}'}), 403
    
    try:
        # 获取要删除的用户
        users = User.query.filter(User.id.in_(user_ids)).all()
        
        if not users:
            return jsonify({'error': '未找到要删除的用户'}), 404
        
        # 创建log
        current_user_id = get_jwt_identity()
        for user in users:
            create_audit_log(
                user_id=current_user_id,
                action='delete_user',
                resource_type='user',
                resource_id=user.id,
                details={'username': user.username, 'batch_delete': True}
            )
        
        # 批量删除
        User.query.filter(User.id.in_(user_ids)).delete(synchronize_session=False)
        db.session.commit()
        
        return jsonify({
            'message': f'成功删除 {len(users)} 个用户'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': '批量删除失败'}), 500


# 导出用户数据
@users_bp.route('/export/<format_type>', methods=['GET'])
@jwt_required()
def export_users(format_type):
    from flask import send_file, make_response, request
    from io import BytesIO
    import csv
    import datetime
    
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = db.session.get(User, int(current_user_id))
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    users = User.query.all()

    if format_type == 'csv':
        from io import StringIO
        output = StringIO()
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow(['ID', '用户名', '邮箱', '姓', '名', '部门', '职位', '工号', '公司电话', '手机号码', '移动电话', '生日', '性别', '工作语言', '状态', '创建时间'])
        for user in users:
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.last_name or '',
                user.first_name or '',
                user.department or '',
                user.position or '',
                getattr(user, 'employee_id', ''),
                getattr(user, 'company_phone', ''),
                getattr(user, 'mobile_phone', ''),
                getattr(user, 'mobile_phone', ''),
                user.birthday.strftime('%Y/%m/%d') if user.birthday else '',
                getattr(user, 'gender', ''),
                getattr(user, 'work_language', ''),
                '激活' if user.is_active else '禁用',
                user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ''
            ])
        output.seek(0)
        # 将StringIO转换为BytesIO并编码为utf-8
        output_bytes = BytesIO(output.getvalue().encode('utf-8'))
        output_bytes.seek(0)
        response = make_response(send_file(
            output_bytes,
            mimetype='text/csv;charset=utf-8',
            as_attachment=True,
            download_name=f'users_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        ))
        filename = f'users_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        # 添加 CORS 头
        origin = request.headers.get('Origin')
        if origin:
            response.headers['Access-Control-Allow-Origin'] = origin
        else:
            response.headers['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        return response
    elif format_type == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '用户数据'

            headers = ['ID', '用户名', '邮箱', '姓', '名', '部门', '职位', '工号', '公司电话', '手机号码', '移动电话', '生日', '性别', '工作语言', '状态', '创建时间']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row, user in enumerate(users, 2):
                ws.cell(row=row, column=1, value=user.id)
                ws.cell(row=row, column=2, value=user.username)
                ws.cell(row=row, column=3, value=user.email)
                ws.cell(row=row, column=4, value=user.last_name or '')
                ws.cell(row=row, column=5, value=user.first_name or '')
                ws.cell(row=row, column=6, value=user.department or '')
                ws.cell(row=row, column=7, value=user.position or '')
                ws.cell(row=row, column=8, value=getattr(user, 'employee_id', ''))
                ws.cell(row=row, column=9, value=getattr(user, 'company_phone', ''))
                ws.cell(row=row, column=10, value=getattr(user, 'mobile_phone', ''))
                ws.cell(row=row, column=11, value=getattr(user, 'mobile_phone', ''))
                ws.cell(row=row, column=12, value=user.birthday.strftime('%Y/%m/%d') if user.birthday else '')
                ws.cell(row=row, column=13, value=getattr(user, 'gender', ''))
                ws.cell(row=row, column=14, value=getattr(user, 'work_language', ''))
                ws.cell(row=row, column=15, value='激活' if user.is_active else '禁用')
                ws.cell(row=row, column=16, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '')
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            output_bytes = output.getvalue()
            output.close()
            output = BytesIO(output_bytes)
            filename = f'users_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response = make_response(send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            ))
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            origin = request.headers.get('Origin')
            if origin:
                response.headers['Access-Control-Allow-Origin'] = origin
            else:
                response.headers['Access-Control-Allow-Origin'] = 'http://localhost:3000'
            response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With'
            response.headers['Access-Control-Allow-Credentials'] = 'true'
            return response
        except ImportError:
            return jsonify({'error': 'Excel导出需要安装openpyxl库'}), 500
        except Exception as e:
            logger.error(f"导出xlsx错误: {str(e)}")
            return jsonify({'error': f'导出失败: {str(e)}'}), 500
    else:
        return jsonify({'error': '不支持的导出格式'}), 400

# 导入用户数据
@users_bp.route('/import', methods=['POST'])
@jwt_required()
def import_users():
    import csv
    from io import StringIO, BytesIO
    import datetime
    
    db = get_db()
    User, UserRole, Department, Position = get_models()
    create_audit_log = get_create_audit_log()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    if 'file' not in request.files:
        return jsonify({'error': '请上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    try:
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            if content.startswith('\ufeff'):
                content = content[1:]
            input_file = StringIO(content)
            reader = csv.DictReader(input_file)

            required_columns = ['用户名', '邮箱']
            if not all(col in reader.fieldnames for col in required_columns):
                return jsonify({'error': f'CSV 文件缺少必要的列：{", ".join(required_columns)}'}), 400

            rows = list(reader)
        elif filename.endswith('.xlsx'):
            import openpyxl
            input_file = BytesIO(file.read())
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            required_columns = ['用户名', '邮箱']
            if not all(col in headers for col in required_columns):
                return jsonify({'error': f'Excel 文件缺少必要的列：{", ".join(required_columns)}'}), 400
            
            # 读取数据行
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if any(cell is not None for cell in row):  # 跳过空行
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            row_dict[header] = row[col_idx] if row[col_idx] is not None else ''
                        else:
                            row_dict[header] = ''
                    rows.append(row_dict)
        else:
            return jsonify({'error': '不支持的文件格式，请上传 CSV 或 Excel 文件'}), 400
        
        # 处理导入数据
        success_count = 0
        error_count = 0
        errors = []
        
        for row_idx, row in enumerate(rows, start=2):
            try:
                username = str(row.get('用户名', '')).strip()
                email = str(row.get('邮箱', '')).strip()

                if not username:
                    errors.append(f'第{row_idx}行：用户名为空')
                    error_count += 1
                    continue

                if not email:
                    errors.append(f'第{row_idx}行：邮箱为空')
                    error_count += 1
                    continue

                if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
                    errors.append(f'第{row_idx}行：邮箱格式不正确：{email}')
                    error_count += 1
                    continue

                existing_user = User.query.filter_by(username=username).first()

                if existing_user:
                    if row.get('邮箱'):
                        existing_user.email = email
                    if row.get('姓'):
                        existing_user.last_name = str(row.get('姓', '')).strip()
                    if row.get('名'):
                        existing_user.first_name = str(row.get('名', '')).strip()
                    if row.get('部门'):
                        existing_user.department = str(row.get('部门', '')).strip()
                    if row.get('职位'):
                        existing_user.position = str(row.get('职位', '')).strip()
                    if '工号' in row:
                        setattr(existing_user, 'employee_id', str(row.get('工号', '')).strip())
                    if '公司电话' in row:
                        setattr(existing_user, 'company_phone', str(row.get('公司电话', '')).strip())
                    if '手机号码' in row:
                        setattr(existing_user, 'mobile_phone', str(row.get('手机号码', '')).strip())
                    if '移动电话' in row:
                        setattr(existing_user, 'mobile_phone', str(row.get('移动电话', '')).strip())
                    if '生日' in row and row.get('生日'):
                        try:
                            existing_user.birthday = datetime.datetime.strptime(str(row.get('生日')), '%Y/%m/%d')
                        except ValueError:
                            try:
                                existing_user.birthday = datetime.datetime.strptime(str(row.get('生日')), '%Y-%m-%d')
                            except ValueError:
                                errors.append(f'第{row_idx}行：生日格式不正确：{row.get("生日")}')
                    if '性别' in row:
                        setattr(existing_user, 'gender', str(row.get('性别', '')).strip())
                    if '工作语言' in row:
                        setattr(existing_user, 'work_language', str(row.get('工作语言', '')).strip())
                    if row.get('状态') == '激活':
                        existing_user.is_active = True
                    elif row.get('状态') == '禁用':
                        existing_user.is_active = False

                    db.session.add(existing_user)
                    success_count += 1
                else:
                    import secrets
                    from werkzeug.security import generate_password_hash

                    default_password = 'User@' + secrets.token_hex(4)

                    new_user = User(
                        username=username,
                        email=email,
                        password_hash=generate_password_hash(default_password),
                        salt=secrets.token_hex(16),
                        first_name=str(row.get('名', '')).strip(),
                        last_name=str(row.get('姓', '')).strip(),
                        department=str(row.get('部门', '')).strip() if row.get('部门') else '',
                        position=str(row.get('职位', '')).strip() if row.get('职位') else '',
                        is_active=(row.get('状态') != '禁用')
                    )
                    
                    # 设置其他字段
                    if '工号' in row:
                        setattr(new_user, 'employee_id', str(row.get('工号', '')).strip())
                    if '公司电话' in row:
                        setattr(new_user, 'company_phone', str(row.get('公司电话', '')).strip())
                    if '手机号码' in row:
                        setattr(new_user, 'mobile_phone', str(row.get('手机号码', '')).strip())
                    if '移动电话' in row:
                        setattr(new_user, 'mobile_phone', str(row.get('移动电话', '')).strip())
                    if '生日' in row and row.get('生日'):
                        try:
                            new_user.birthday = datetime.datetime.strptime(str(row.get('生日')), '%Y/%m/%d')
                        except ValueError:
                            try:
                                new_user.birthday = datetime.datetime.strptime(str(row.get('生日')), '%Y-%m-%d')
                            except ValueError:
                                pass
                    if '性别' in row:
                        setattr(new_user, 'gender', str(row.get('性别', '')).strip())
                    if '工作语言' in row:
                        setattr(new_user, 'work_language', str(row.get('工作语言', '')).strip())
                    
                    db.session.add(new_user)
                    success_count += 1
                    
            except Exception as e:
                errors.append(f'第{row_idx}行：处理失败 - {str(e)}')
                error_count += 1
        
        # 提交数据库事务
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='import_users',
            resource_type='user',
            resource_id=0,
            details={
                'success_count': success_count,
                'error_count': error_count,
                'filename': file.filename
            }
        )
        
        # 返回结果
        result = {
            'message': f'导入完成，成功{success_count}条，失败{error_count}条',
            'processed': success_count + error_count,
            'success_count': success_count,
            'error_count': error_count
        }
        
        if errors:
            result['errors'] = errors[:10]  # 只返回前 10 个错误
        
        return jsonify(result), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"导入用户失败：{str(e)}")
        return jsonify({'error': f'导入失败：{str(e)}'}), 500

# 获取所有部门列表
@users_bp.route('/departments', methods=['GET'])
@jwt_required()
def get_departments():
    db = get_db()
    User, UserRole, Department, Position = get_models()
    
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    departments = Department.query.order_by(Department.name).all()
    department_list = [d.name for d in departments]

    return jsonify({'departments': department_list}), 200

# 获取部门人员列表
@users_bp.route('/department/<path:department_name>/members', methods=['GET'])
@jwt_required()
def get_department_members(department_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    query = User.query.filter_by(department=department_name)
    
    pagination = query.order_by(User.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    members_data = []
    for user in pagination.items:
        members_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'position': user.position,
            'department': user.department,
            'is_active': user.is_active,
            'created_at': user.created_at.isoformat() if user.created_at else None,
            'last_login': user.last_login.isoformat() if user.last_login else None,
            'employee_id': getattr(user, 'employee_id', ''),
            'company_phone': getattr(user, 'company_phone', ''),
            'mobile_phone': getattr(user, 'mobile_phone', ''),
            'birthday': user.birthday.strftime('%Y-%m-%d') if user.birthday else '',
            'gender': getattr(user, 'gender', ''),
            'work_language': getattr(user, 'work_language', '')
        })
    
    return jsonify({
        'department': department_name,
        'members': members_data,
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
        'per_page': pagination.per_page
    }), 200

# 创建部门
@users_bp.route('/departments', methods=['POST'])
@jwt_required()
def create_department():
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    data = request.get_json()
    department_name = data.get('department', '').strip()
    
    if not department_name:
        return jsonify({'error': '部门名称不能为空'}), 400
    
    existing_department = Department.query.filter_by(name=department_name).first()
    
    if existing_department:
        return jsonify({'error': '部门已存在'}), 400
    
    new_department = Department(name=department_name)
    db.session.add(new_department)
    db.session.commit()
    
    return jsonify({
        'message': '部门创建成功',
        'department': department_name
    }), 201

# 更新部门名称
@users_bp.route('/departments/<path:old_department_name>', methods=['PUT'])
@jwt_required()
def update_department(old_department_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    data = request.get_json()
    new_department_name = data.get('department', '').strip()
    
    if not new_department_name:
        return jsonify({'error': '部门名称不能为空'}), 400
    
    if old_department_name == new_department_name:
        return jsonify({'message': '部门名称未变化'}), 200
    
    existing_department = Department.query.filter_by(name=new_department_name).first()
    
    if existing_department:
        return jsonify({'error': '新部门名称已存在'}), 400
    
    department = Department.query.filter_by(name=old_department_name).first()
    if department:
        department.name = new_department_name
    
    users_to_update = User.query.filter_by(department=old_department_name).all()
    for user in users_to_update:
        user.department = new_department_name
    
    db.session.commit()
    
    return jsonify({
        'message': '部门更新成功',
        'old_department': old_department_name,
        'new_department': new_department_name,
        'updated_count': len(users_to_update)
    }), 200

# 删除部门
@users_bp.route('/departments/<path:department_name>', methods=['DELETE'])
@jwt_required()
def delete_department(department_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()
    
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    users_in_department = User.query.filter_by(department=department_name).all()

    if users_in_department:
        return jsonify({
            'error': '该部门下还有成员，无法删除',
            'member_count': len(users_in_department)
        }), 400

    department = Department.query.filter_by(name=department_name).first()
    if department:
        db.session.delete(department)
        db.session.commit()

    return jsonify({
        'message': '部门删除成功',
        'department': department_name
    }), 200

# 批量添加用户到部门
@users_bp.route('/departments/<path:department_name>/members/batch-add', methods=['POST'])
@jwt_required()
def batch_add_department_members(department_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    data = request.get_json()
    user_ids = data.get('user_ids', [])
    
    if not user_ids or not isinstance(user_ids, list):
        return jsonify({'error': '请提供有效的用户ID列表'}), 400
    
    success_count = 0
    failed_users = []
    
    for user_id in user_ids:
        try:
            user = User.query.get(user_id)
            if user:
                user.department = department_name
                success_count += 1
            else:
                failed_users.append({'id': user_id, 'reason': '用户不存在'})
        except Exception as e:
            failed_users.append({'id': user_id, 'reason': str(e)})
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'批量添加失败: {str(e)}'}), 500
    
    return jsonify({
        'message': f'成功添加 {success_count} 个用户到部门',
        'success_count': success_count,
        'failed_users': failed_users
    }), 200

# 批量移除部门用户
@users_bp.route('/departments/<path:department_name>/members/batch-remove', methods=['POST'])
@jwt_required()
def batch_remove_department_members(department_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    data = request.get_json()
    user_ids = data.get('user_ids', [])

    if not user_ids or not isinstance(user_ids, list):
        return jsonify({'error': '请提供有效的用户ID列表'}), 400

    success_count = 0
    failed_users = []

    for user_id in user_ids:
        try:
            user = User.query.get(user_id)
            if user and user.department == department_name:
                user.department = ''
                success_count += 1
            else:
                failed_users.append({'id': user_id, 'reason': '用户不存在或不在该部门'})
        except Exception as e:
            failed_users.append({'id': user_id, 'reason': str(e)})
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'批量移除失败: {str(e)}'}), 500
    
    return jsonify({
        'message': f'成功从部门移除 {success_count} 个用户',
        'success_count': success_count,
        'failed_users': failed_users
    }), 200

# 获取所有职位列表
@users_bp.route('/positions', methods=['GET'])
@jwt_required()
def get_positions():
    db = get_db()
    User, UserRole, Department, Position = get_models()
    
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404
    
    positions = Position.query.order_by(Position.name).all()
    position_list = [p.name for p in positions]
    
    return jsonify({'positions': position_list}), 200

# 获取职位人员列表
@users_bp.route('/position/<path:position_name>/members', methods=['GET'])
@jwt_required()
def get_position_members(position_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    query = User.query.filter_by(position=position_name)

    pagination = query.order_by(User.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    members_data = []
    for user in pagination.items:
        members_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'position': user.position,
            'department': user.department,
            'is_active': user.is_active,
            'created_at': user.created_at.isoformat() if user.created_at else None,
            'last_login': user.last_login.isoformat() if user.last_login else None,
            'employee_id': getattr(user, 'employee_id', ''),
            'company_phone': getattr(user, 'company_phone', ''),
            'mobile_phone': getattr(user, 'mobile_phone', ''),
            'birthday': user.birthday.strftime('%Y-%m-%d') if user.birthday else '',
            'gender': getattr(user, 'gender', ''),
            'work_language': getattr(user, 'work_language', '')
        })

    return jsonify({
        'position': position_name,
        'members': members_data,
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
        'per_page': pagination.per_page
    }), 200

# 创建职位
@users_bp.route('/positions', methods=['POST'])
@jwt_required()
def create_position():
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    data = request.get_json()
    position_name = data.get('position', '').strip()

    if not position_name:
        return jsonify({'error': '职位名称不能为空'}), 400

    existing_position = Position.query.filter_by(name=position_name).first()

    if existing_position:
        return jsonify({'error': '职位已存在'}), 400

    new_position = Position(name=position_name)
    db.session.add(new_position)
    db.session.commit()

    return jsonify({
        'message': '职位创建成功',
        'position': position_name
    }), 201

# 更新职位名称
@users_bp.route('/positions/<path:old_position_name>', methods=['PUT'])
@jwt_required()
def update_position(old_position_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    data = request.get_json()
    new_position_name = data.get('position', '').strip()

    if not new_position_name:
        return jsonify({'error': '职位名称不能为空'}), 400

    if old_position_name == new_position_name:
        return jsonify({'message': '职位名称未变化'}), 200

    existing_position = Position.query.filter_by(name=new_position_name).first()

    if existing_position:
        return jsonify({'error': '新职位名称已存在'}), 400

    position = Position.query.filter_by(name=old_position_name).first()
    if position:
        position.name = new_position_name

    users_to_update = User.query.filter_by(position=old_position_name).all()
    for user in users_to_update:
        user.position = new_position_name

    db.session.commit()

    return jsonify({
        'message': '职位更新成功',
        'old_position': old_position_name,
        'new_position': new_position_name,
        'updated_count': len(users_to_update)
    }), 200

# 删除职位
@users_bp.route('/positions/<path:position_name>', methods=['DELETE'])
@jwt_required()
def delete_position(position_name):
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    users_with_position = User.query.filter_by(position=position_name).all()

    if users_with_position:
        return jsonify({
            'error': '该职位下还有成员，无法删除',
            'member_count': len(users_with_position)
        }), 400
    
    position = Position.query.filter_by(name=position_name).first()
    if position:
        db.session.delete(position)
        db.session.commit()
    
    return jsonify({
        'message': '职位删除成功',
        'position': position_name
    }), 200

# 获取所有可用权限列表
@users_bp.route('/permissions', methods=['GET'])
@jwt_required()
def get_all_permissions():
    """获取所有可分配的权限列表"""
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    from models.permissions import PermissionCodes, DEFAULT_ROLES

    permission_categories = {
        'system': {
            'name': '系统权限',
            'permissions': [
                {'code': PermissionCodes.SYSTEM_ADMIN, 'name': '系统管理员', 'description': '拥有系统全部权限'},
                {'code': PermissionCodes.STATISTICS_VIEW, 'name': '查看统计', 'description': '查看系统统计数据'},
                {'code': PermissionCodes.DATA_EXPORT, 'name': '导出数据', 'description': '导出系统数据'},
                {'code': PermissionCodes.DATA_IMPORT, 'name': '导入数据', 'description': '导入系统数据'},
            ]
        },
        'user': {
            'name': '用户管理',
            'permissions': [
                {'code': PermissionCodes.USER_VIEW, 'name': '查看用户', 'description': '查看用户信息'},
                {'code': PermissionCodes.USER_CREATE, 'name': '创建用户', 'description': '创建新用户'},
                {'code': PermissionCodes.USER_EDIT, 'name': '编辑用户', 'description': '编辑用户信息'},
                {'code': PermissionCodes.USER_DELETE, 'name': '删除用户', 'description': '删除用户'},
                {'code': PermissionCodes.USER_IMPORT, 'name': '导入用户', 'description': '批量导入用户'},
                {'code': PermissionCodes.USER_EXPORT, 'name': '导出用户', 'description': '导出用户数据'},
                {'code': PermissionCodes.USER_ASSIGN_ROLE, 'name': '分配角色', 'description': '分配用户角色'},
            ]
        },
        'project': {
            'name': '项目管理',
            'permissions': [
                {'code': PermissionCodes.PROJECT_VIEW, 'name': '查看项目', 'description': '查看项目信息'},
                {'code': PermissionCodes.PROJECT_CREATE, 'name': '创建项目', 'description': '创建新项目'},
                {'code': PermissionCodes.PROJECT_EDIT, 'name': '编辑项目', 'description': '编辑项目信息'},
                {'code': PermissionCodes.PROJECT_DELETE, 'name': '删除项目', 'description': '删除项目'},
                {'code': PermissionCodes.PROJECT_ASSIGN_MEMBER, 'name': '分配成员', 'description': '分配项目成员'},
                {'code': PermissionCodes.PROJECT_MANAGE_SETTINGS, 'name': '管理设置', 'description': '管理项目设置'},
            ]
        },
        'bug': {
            'name': 'Bug管理',
            'permissions': [
                {'code': PermissionCodes.BUG_VIEW, 'name': '查看Bug', 'description': '查看Bug列表和详情'},
                {'code': PermissionCodes.BUG_CREATE, 'name': '创建Bug', 'description': '创建新Bug'},
                {'code': PermissionCodes.BUG_EDIT, 'name': '编辑Bug', 'description': '编辑Bug信息'},
                {'code': PermissionCodes.BUG_DELETE, 'name': '删除Bug', 'description': '删除Bug'},
                {'code': PermissionCodes.BUG_ASSIGN, 'name': '分配Bug', 'description': '分配Bug给其他人'},
                {'code': PermissionCodes.BUG_RESOLVE, 'name': '解决Bug', 'description': '标记Bug为已解决'},
                {'code': PermissionCodes.BUG_CLOSE, 'name': '关闭Bug', 'description': '关闭Bug'},
            ]
        },
        'task': {
            'name': '任务管理',
            'permissions': [
                {'code': PermissionCodes.TASK_VIEW, 'name': '查看任务', 'description': '查看任务列表和详情'},
                {'code': PermissionCodes.TASK_CREATE, 'name': '创建任务', 'description': '创建新任务'},
                {'code': PermissionCodes.TASK_EDIT, 'name': '编辑任务', 'description': '编辑任务信息'},
                {'code': PermissionCodes.TASK_DELETE, 'name': '删除任务', 'description': '删除任务'},
                {'code': PermissionCodes.TASK_ASSIGN, 'name': '分配任务', 'description': '分配任务给其他人'},
                {'code': PermissionCodes.TASK_UPDATE_STATUS, 'name': '更新状态', 'description': '更新任务状态'},
            ]
        },
        'attendance': {
            'name': '考勤管理',
            'permissions': [
                {'code': PermissionCodes.ATTENDANCE_VIEW, 'name': '查看考勤', 'description': '查看考勤记录'},
                {'code': PermissionCodes.ATTENDANCE_MANAGE, 'name': '管理考勤', 'description': '管理考勤设置'},
                {'code': PermissionCodes.CLOCK_IN, 'name': '上班打卡', 'description': '上班签到'},
                {'code': PermissionCodes.CLOCK_OUT, 'name': '下班打卡', 'description': '下班签退'},
                {'code': PermissionCodes.LEAVE_APPLY, 'name': '申请请假', 'description': '提交请假申请'},
                {'code': PermissionCodes.LEAVE_APPROVE, 'name': '审批请假', 'description': '审批请假申请'},
                {'code': PermissionCodes.OVERTIME_APPLY, 'name': '申请加班', 'description': '提交加班申请'},
                {'code': PermissionCodes.OVERTIME_APPROVE, 'name': '审批加班', 'description': '审批加班申请'},
                {'code': PermissionCodes.EXCEPTION_HANDLE, 'name': '异常处理', 'description': '处理考勤异常'},
                {'code': PermissionCodes.EXCEPTION_APPROVE, 'name': '审批异常', 'description': '审批考勤异常'},
                {'code': PermissionCodes.ATTENDANCE_REPORT, 'name': '考勤报表', 'description': '查看考勤报表'},
            ]
        }
    }

    return jsonify({
        'permissions': permission_categories,
        'roles': DEFAULT_ROLES
    }), 200

# 获取用户的自定义权限
@users_bp.route('/<int:user_id>/permissions', methods=['GET'])
@jwt_required()
def get_user_permissions(user_id):
    """获取指定用户的自定义权限"""
    from models.permissions import DEFAULT_ROLES

    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    target_user = User.query.get(user_id)
    if not target_user:
        return jsonify({'error': '用户不存在'}), 404

    target_position = target_user.get_position_info()
    custom_perms = target_user.get_custom_permissions()

    return jsonify({
        'user_id': user_id,
        'username': target_user.username,
        'position': target_user.position,
        'is_super_admin': target_user.is_super_admin,
        'custom_permissions': custom_perms,
        'all_permissions': custom_perms.get('allowed', []) + custom_perms.get('denied', [])
    }), 200

# 更新用户的自定义权限
@users_bp.route('/<int:user_id>/permissions', methods=['PUT'])
@log_api_call
@log_business_operation
@jwt_required()
def update_user_permissions(user_id):
    """更新指定用户的自定义权限"""
    db, User, UserRole, create_audit_log = get_db_and_models()

    data = request.get_json()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    target_user = User.query.get(user_id)
    if not target_user:
        return jsonify({'error': '用户不存在'}), 404

    target_position = target_user.get_position_info()
    if target_position and target_position.is_admin:
        return jsonify({'error': '无法修改管理员的权限'}), 403

    if target_user.is_super_admin:
        return jsonify({'error': '无法修改系统管理员的权限'}), 403

    if 'allowed' in data:
        if not isinstance(data['allowed'], list):
            return jsonify({'error': 'allowed权限必须为列表'}), 400

    if 'denied' in data:
        if not isinstance(data['denied'], list):
            return jsonify({'error': 'denied权限必须为列表'}), 400

    custom_perms = {
        'allowed': data.get('allowed', []),
        'denied': data.get('denied', [])
    }

    target_user.set_custom_permissions(custom_perms)

    try:
        db.session.commit()

        log_manager = get_log_manager()
        log_manager.log_business(
            operation='user_permissions_updated',
            user_id=current_user_id,
            details={
                'target_user_id': user_id,
                'username': target_user.username,
                'new_permissions': custom_perms
            }
        )

        create_audit_log(
            user_id=current_user_id,
            action='update_user_permissions',
            resource_type='user',
            resource_id=user_id,
            details={'username': target_user.username, 'permissions': custom_perms}
        )

        return jsonify({
            'message': '权限更新成功',
            'user_id': user_id,
            'username': target_user.username,
            'custom_permissions': custom_perms
        }), 200
    except Exception as e:
        db.session.rollback()
        log_manager = get_log_manager()
        log_manager.log_error('update_user_permissions_error', f"更新用户权限失败: {str(e)}")
        return jsonify({'error': '更新权限失败'}), 500

# 批量更新用户角色
@users_bp.route('/batch-update-role', methods=['POST'])
@log_api_call
@jwt_required()
def batch_update_user_role():
    """批量更新用户角色"""
    db = get_db()
    User, UserRole, Department, Position = get_models()

    data = request.get_json()

    if not data or 'user_ids' not in data or 'role' not in data:
        return jsonify({'error': '请提供用户ID列表和新角色'}), 400

    user_ids = data['user_ids']
    new_position = data.get('position', '')

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    if not new_position:
        return jsonify({'error': '职位不能为空'}), 400

    existing_position = Position.query.filter_by(name=new_position).first()
    if not existing_position:
        return jsonify({'error': '无效的职位'}), 400

    success_count = 0
    failed_users = []

    for user_id in user_ids:
        try:
            user = User.query.get(user_id)
            if not user:
                failed_users.append({'id': user_id, 'reason': '用户不存在'})
                continue
            if user.is_super_admin:
                failed_users.append({'id': user_id, 'reason': '无法修改系统管理员职位'})
                continue
            user.position = new_position
            success_count += 1
        except Exception as e:
            failed_users.append({'id': user_id, 'reason': str(e)})

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'批量更新失败: {str(e)}'}), 500

    return jsonify({
        'message': f'成功更新 {success_count} 个用户的职位',
        'success_count': success_count,
        'failed_users': failed_users
    }), 200

# 获取职位列表
@users_bp.route('/positions', methods=['GET'])
@jwt_required()
def get_position_list():
    """获取所有职位列表"""
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    position_info = current_user.get_position_info()
    if not (current_user.is_super_admin or (position_info and (position_info.is_admin or position_info.is_manager))):
        return jsonify({'error': '权限不足'}), 403

    positions = Position.query.all()
    positions_data = []
    for pos in positions:
        user_count = User.query.filter_by(position=pos.name).count()
        positions_data.append({
            'name': pos.name,
            'description': pos.description,
            'is_admin': pos.is_admin,
            'is_manager': pos.is_manager,
            'user_count': user_count
        })

    return jsonify({'positions': positions_data}), 200

def _build_user_data(user):
    """构建用户数据字典"""
    return {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'position': user.position,
        'department': user.department,
        'is_active': user.is_active,
        'created_at': user.created_at.isoformat() if user.created_at else None,
        'last_login': user.last_login.isoformat() if user.last_login else None,
        'employee_id': getattr(user, 'employee_id', ''),
        'company_phone': getattr(user, 'company_phone', ''),
        'mobile_phone': getattr(user, 'mobile_phone', ''),
        'birthday': user.birthday.strftime('%Y-%m-%d') if user.birthday else '',
        'gender': getattr(user, 'gender', ''),
        'work_language': getattr(user, 'work_language', '')
    }


def _get_department_members(db, User, department_name, page=1, per_page=20, search=''):
    """获取部门成员列表"""
    query = User.query.filter_by(department=department_name)

    if search:
        query = query.filter(
            (User.username.contains(search)) |
            (User.email.contains(search)) |
            (User.first_name.contains(search)) |
            (User.last_name.contains(search))
        )

    pagination = query.order_by(User.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    members_data = [_build_user_data(user) for user in pagination.items]

    return {
        'members': members_data,
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
        'per_page': pagination.per_page,
        'statistics': {
            'total': pagination.total,
            'active': User.query.filter_by(department=department_name, is_active=True).count(),
            'inactive': User.query.filter_by(department=department_name, is_active=False).count()
        }
    }


@users_bp.route('/my-department', methods=['GET'])
@jwt_required()
def get_my_department():
    """
    获取我的部门信息

    权限规则：
    1. 系统管理员（is_super_admin=True）：查看所有部门和所有人员
    2. 部门经理（position.is_manager=True）：查看自己部门的成员
    3. 普通员工：只能查看自己的部门信息，不能查看成员列表
    """
    db = get_db()
    User, UserRole, Department, Position = get_models()

    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    if not current_user:
        return jsonify({'error': '用户不存在'}), 404

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    department_filter = request.args.get('department', '').strip()

    current_user_data = {
        'id': current_user.id,
        'username': current_user.username,
        'position': current_user.position,
        'department': current_user.department
    }

    if current_user.is_system_admin():
        departments = Department.query.order_by(Department.name).all()
        departments_data = [
            {
                'id': dept.id,
                'name': dept.name,
                'description': dept.description,
                'member_count': User.query.filter_by(department=dept.name).count()
            }
            for dept in departments
        ]

        query = User.query
        if department_filter:
            query = query.filter_by(department=department_filter)
        if search:
            query = query.filter(
                (User.username.contains(search)) |
                (User.email.contains(search)) |
                (User.first_name.contains(search)) |
                (User.last_name.contains(search))
            )

        pagination = query.order_by(User.id.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )

        members_data = [_build_user_data(user) for user in pagination.items]

        total_users = User.query.count()
        active_users = User.query.filter_by(is_active=True).count()
        inactive_users = User.query.filter_by(is_active=False).count()

        return jsonify({
            'has_department': True,
            'is_manager': True,
            'is_system_admin': True,
            'can_view_all': True,
            'departments': departments_data,
            'members': members_data,
            'total_members': pagination.total,
            'page': pagination.page,
            'pages': pagination.pages,
            'per_page': pagination.per_page,
            'current_user': current_user_data,
            'statistics': {
                'total': total_users,
                'active': active_users,
                'inactive': inactive_users
            }
        }), 200

    if not current_user.department:
        return jsonify({
            'has_department': False,
            'is_manager': False,
            'is_system_admin': False,
            'can_view_all': False,
            'current_user': current_user_data,
            'message': '您尚未加入任何部门'
        }), 200

    department = Department.query.filter_by(name=current_user.department).first()

    response_data = {
        'has_department': True,
        'is_manager': current_user.is_department_manager(),
        'is_system_admin': False,
        'can_view_all': False,
        'department': {
            'id': department.id if department else None,
            'name': current_user.department,
            'description': department.description if department else None
        },
        'current_user': current_user_data
    }

    if current_user.is_department_manager():
        members_result = _get_department_members(
            db, User, current_user.department,
            page=page, per_page=per_page, search=search
        )
        response_data.update(members_result)
    else:
        response_data['message'] = '您不是部门经理，无法查看部门员工列表'

    return jsonify(response_data), 200
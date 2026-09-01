"""
活动记录API - 系统审计日志
"""
from flask import Blueprint, request, jsonify, send_file, make_response
from datetime import datetime, timedelta
from utils.time_utils import now_china
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import func
from io import BytesIO
import json
import re
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建活动记录蓝图
activities_bp = Blueprint('activities', __name__, url_prefix='/activities')

def get_models():
    from enhanced_app import User, Activity, Bug, Project, WorkLog, app, db
    return User, Activity, Bug, Project, WorkLog, app, db

# 获取活动记录统计
@activities_bp.route('/statistics', methods=['GET'])
@jwt_required()
def get_statistics():
    """获取活动记录统计数据（全量，非当前页）"""
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        total = db.session.query(func.count(Activity.id)).scalar() or 0
        create_count = db.session.query(func.count(Activity.id)).filter(
            Activity.action.like('%create%')
        ).scalar() or 0
        update_count = db.session.query(func.count(Activity.id)).filter(
            Activity.action.like('%update%')
        ).scalar() or 0
        delete_count = db.session.query(func.count(Activity.id)).filter(
            Activity.action.like('%delete%')
        ).scalar() or 0

        today_start = now_china().replace(hour=0, minute=0, second=0, microsecond=0)
        today_count = db.session.query(func.count(Activity.id)).filter(
            Activity.created_at >= today_start
        ).scalar() or 0

        return jsonify({
            'total': total,
            'create_count': create_count,
            'update_count': update_count,
            'delete_count': delete_count,
            'today_count': today_count
        }), 200

# 获取活动记录列表
@activities_bp.route('/', methods=['GET'])
@jwt_required()
def get_activities():
    """获取活动记录列表"""
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        resource_type = request.args.get('resource_type')
        action = request.args.get('action')
        user_name = request.args.get('user_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        query = Activity.query

        if resource_type:
            query = query.filter(Activity.target_type == resource_type)

        if action:
            query = query.filter(Activity.action.like(f'%{action}%'))

        if user_name:
            query = query.join(User).filter(User.username.contains(user_name))

        if start_date:
            query = query.filter(Activity.created_at >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(Activity.created_at <= datetime.fromisoformat(end_date))

        pagination = query.order_by(Activity.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )

        activities = []
        for activity in pagination.items:
            activity_dict = {
                'id': activity.id,
                'action': activity.action,
                'description': activity.description,
                'target_type': activity.target_type,
                'target_id': activity.target_id,
                'resource_type': activity.target_type,
                'resource_id': activity.target_id,
                'user_id': activity.performed_by,
                'performed_by': activity.performed_by,
                'created_at': activity.created_at.isoformat() if activity.created_at else None,
                'field_changes': json.loads(activity.field_changes) if activity.field_changes else [],
                'resource_name': resolve_resource_name(activity),
            }

            performer = User.query.get(activity.performed_by)
            if performer:
                activity_dict['user_name'] = performer.username
                activity_dict['user_role'] = performer.role

            activities.append(activity_dict)

        return jsonify({
            'activities': activities,
            'total': pagination.total,
            'page': page,
            'pages': pagination.pages,
            'per_page': per_page
        }), 200


# 操作类型中文映射
ACTION_TEXT_MAP = {
    'create': '创建', 'create_bug': '创建Bug', 'create_project': '创建项目',
    'create_project_log': '创建日志', 'create_project_member': '添加项目成员',
    'create_user': '创建用户', 'create_work_log': '创建工作日志',
    'create_leave_application': '创建请假申请', 'create_overtime_application': '创建加班申请',
    'create_risk': '创建风险', 'create_requirement_document': '创建需求文档',
    'create_requirement_item': '创建需求条目', 'create_knowledge_article': '创建知识文章',
    'create_knowledge_category': '创建知识分类', 'create_personal_task': '创建个人任务',
    'create_personal_template': '创建个人模板', 'create_test_suite': '创建测试套件',
    'create_test_case': '创建测试用例', 'create_shift_schedule': '创建班次',
    'create_user_shift': '分配班次', 'update': '更新', 'update_bug': '更新Bug',
    'update_project': '更新项目', 'update_project_log': '更新日志',
    'update_user': '更新用户', 'update_work_log': '更新工作日志',
    'update_risk': '更新风险', 'delete': '删除', 'delete_project': '删除项目',
    'delete_user': '删除用户', 'delete_work_log': '删除工作日志',
    'delete_risk': '删除风险', 'bug_status_update': '状态更新',
    'bug_status_transition': '状态转换', 'status_change': '状态变更',
    'assign_bug': '分配', 'assign': '分配', 'add_project_member': '添加成员',
    'remove_project_member': '移除成员', 'approve': '审批通过',
    'approve_leave_application': '审批请假', 'approve_overtime_application': '审批加班',
    'reject': '审批拒绝', 'clock_in': '上班打卡', 'clock_out': '下班打卡',
    'upload_attachment': '上传附件', 'delete_attachment': '删除附件',
    'export': '导出', 'data_import': '数据导入', 'data_export': '数据导出',
    'import': '导入', 'user_register': '用户注册', 'user_login': '用户登录',
    'batch_create': '批量创建', 'login': '登录', 'register': '注册',
}

# 资源类型中文映射
RESOURCE_TYPE_TEXT_MAP = {
    'project': '项目', 'bug': '缺陷', 'user': '用户', 'project_member': '项目成员',
    'work_log': '工作日志', 'project_log': '项目日志',
    'requirement_document': '需求文档', 'requirement_item': '需求条目',
    'personal_task': '个人任务', 'personal_template': '任务模板',
    'knowledge_article': '知识文章', 'knowledge_category': '知识分类',
    'leave_application': '请假申请', 'overtime_application': '加班申请',
    'attendance': '考勤', 'shift_schedule': '班次', 'user_shift': '排班',
    'risk': '风险', 'test_suite': '测试套件', 'test_case': '测试用例',
    'material': '物料', 'contract': '合同', 'data': '数据',
}


def _build_activity_export_rows(query, User, Activity):
    """构造活动记录导出数据行"""
    rows = []
    activities = query.all()
    # 预加载所有执行人，避免 N+1 查询
    performer_ids = list({a.performed_by for a in activities if a.performed_by})
    performers = {u.id: u for u in User.query.filter(User.id.in_(performer_ids)).all()} if performer_ids else {}

    for activity in activities:
        performer = performers.get(activity.performed_by)
        # 解析变更详情
        changes_text = ''
        if activity.field_changes:
            try:
                changes = json.loads(activity.field_changes)
                if isinstance(changes, list):
                    parts = []
                    for c in changes:
                        field = c.get('field', c.get('field_label', ''))
                        old_val = c.get('old_value', c.get('from', ''))
                        new_val = c.get('new_value', c.get('to', ''))
                        parts.append(f'{field}: {old_val} -> {new_val}')
                    changes_text = '; '.join(parts)
                elif isinstance(changes, dict):
                    parts = []
                    for field, val in changes.items():
                        old_val = val.get('from', val.get('old_value', ''))
                        new_val = val.get('to', val.get('new_value', ''))
                        parts.append(f'{field}: {old_val} -> {new_val}')
                    changes_text = '; '.join(parts)
            except (json.JSONDecodeError, TypeError):
                changes_text = str(activity.field_changes)

        rows.append({
            'ID': activity.id,
            '操作类型': ACTION_TEXT_MAP.get(activity.action, activity.action or ''),
            '资源类型': RESOURCE_TYPE_TEXT_MAP.get(activity.target_type, activity.target_type or ''),
            '资源名称': resolve_resource_name(activity),
            '资源ID': activity.target_id,
            '操作描述': activity.description or '',
            '变更详情': changes_text,
            '操作用户': performer.username if performer else '',
            '用户角色': performer.role if performer else '',
            '操作时间': activity.created_at.strftime('%Y-%m-%d %H:%M:%S') if activity.created_at else '',
        })
    return rows


def _generate_xlsx(rows, sheet_name='活动记录'):
    """生成 xlsx 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        ws.append(['暂无数据'])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # 表头
    headers = list(rows[0].keys())
    ws.append(headers)

    # 表头样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='38BDF8', end_color='38BDF8', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 数据行
    data_font = Font(name='微软雅黑', size=10)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.font = data_font
            cell.alignment = left_align
            cell.border = thin_border

    # 自动列宽
    for col_idx, key in enumerate(headers, 1):
        max_len = len(str(key))
        for row in rows:
            val_len = len(str(row.get(key, '')))
            if val_len > max_len:
                max_len = val_len
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # 冻结表头
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generate_csv(rows):
    """生成 csv 文件（UTF-8 BOM，兼容 Excel 中文）"""
    import io as _io
    output = BytesIO()
    if not rows:
        output.write('\ufeff暂无数据\n'.encode('utf-8'))
        output.seek(0)
        return output

    headers = list(rows[0].keys())
    text_buf = _io.StringIO()
    writer = csv.DictWriter(
        text_buf,
        fieldnames=headers,
        extrasaction='ignore',
        lineterminator='\n'
    )
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    output.write('\ufeff'.encode('utf-8'))
    output.write(text_buf.getvalue().encode('utf-8'))
    output.seek(0)
    return output


# 导出活动记录
@activities_bp.route('/export', methods=['GET'])
@jwt_required()
def export_activities():
    """导出活动记录为 Excel/CSV 文件"""
    from urllib.parse import quote
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        # 解析筛选条件（与列表接口一致）
        resource_type = request.args.get('resource_type')
        action = request.args.get('action')
        user_name = request.args.get('user_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        format_type = request.args.get('format', 'xlsx').lower()

        query = Activity.query

        if resource_type:
            query = query.filter(Activity.target_type == resource_type)

        if action:
            query = query.filter(Activity.action.like(f'%{action}%'))

        if user_name:
            query = query.join(User).filter(User.username.contains(user_name))

        if start_date:
            query = query.filter(Activity.created_at >= datetime.fromisoformat(start_date))
        if end_date:
            # 结束日期包含当天全天
            end_dt = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59)
            query = query.filter(Activity.created_at <= end_dt)

        query = query.order_by(Activity.created_at.desc())
        rows = _build_activity_export_rows(query, User, Activity)

        # 生成文件
        date_str = now_china().strftime('%Y%m%d')
        if format_type == 'xlsx':
            output = _generate_xlsx(rows)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f'活动记录_{date_str}.xlsx'
        elif format_type == 'csv':
            output = _generate_csv(rows)
            mimetype = 'text/csv'
            filename = f'活动记录_{date_str}.csv'
        else:
            return jsonify({'error': f'不支持的格式: {format_type}，仅支持 xlsx 和 csv'}), 400

        response = make_response(send_file(
            output,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        ))
        origin = request.headers.get('Origin')
        response.headers['Access-Control-Allow-Origin'] = origin or '*'
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        # 使用 RFC 5987 编码中文文件名，避免 WSGI latin-1 编码错误
        ascii_name = f'activity_log_{date_str}.{format_type}'
        encoded_name = quote(filename)
        response.headers['Content-Disposition'] = (
            f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
        )

        return response


def resolve_resource_name(activity):
    """根据 target_type 和 target_id 解析资源名称"""
    try:
        if activity.target_id == 0:
            return activity.description or '系统操作'

        target_type = activity.target_type
        target_id = activity.target_id

        if target_type == 'bug':
            from enhanced_app import Bug
            bug = Bug.query.get(target_id)
            return bug.title if bug else '缺陷已删除'
        elif target_type == 'project':
            from enhanced_app import Project
            project = Project.query.get(target_id)
            return project.name if project else '项目已删除'
        elif target_type == 'user':
            from enhanced_app import User
            user = User.query.get(target_id)
            return user.username if user else '用户已删除'
        elif target_type == 'work_log':
            from enhanced_app import WorkLog
            work_log = WorkLog.query.get(target_id)
            if work_log:
                return work_log.title
            match = re.search(r'工作日志[：:](.+)', activity.description)
            return match.group(1) if match else '工作日志已删除'
        elif target_type == 'project_log':
            from enhanced_app import ProjectLog
            log = ProjectLog.query.get(target_id)
            return log.title if log else '项目日志已删除'
        elif target_type == 'project_member':
            from enhanced_app import ProjectMember
            member = ProjectMember.query.get(target_id)
            if member:
                user = member.user
                return f'{user.username} (项目成员)' if user else '项目成员'
            return '项目成员'
        elif target_type == 'leave_application':
            from enhanced_app import LeaveApplication
            app_obj = LeaveApplication.query.get(target_id)
            if app_obj:
                return f'请假申请 #{app_obj.id}'
            return '请假申请'
        elif target_type == 'overtime_application':
            from enhanced_app import OvertimeApplication
            app_obj = OvertimeApplication.query.get(target_id)
            if app_obj:
                return f'加班申请 #{app_obj.id}'
            return '加班申请'
        elif target_type == 'attendance':
            from enhanced_app import AttendanceRecord
            record = AttendanceRecord.query.get(target_id)
            if record:
                return f'考勤记录 #{record.id}'
            return '考勤记录'
        elif target_type == 'shift_schedule':
            from enhanced_app import ShiftSchedule
            shift = ShiftSchedule.query.get(target_id)
            return shift.name if shift else '班次'
        elif target_type == 'risk':
            from enhanced_app import Risk
            risk = Risk.query.get(target_id)
            return risk.title if risk else '风险已删除'
        elif target_type == 'test_suite':
            from enhanced_app import TestSuite
            suite = TestSuite.query.get(target_id)
            return suite.name if suite else '测试套件已删除'
        elif target_type == 'test_case':
            from enhanced_app import TestCase
            case = TestCase.query.get(target_id)
            return case.title if case else '测试用例已删除'
        else:
            # 从描述中提取资源名称
            return activity.description or target_type
    except Exception:
        return activity.description or '未知'


# 获取最近的活動记录
@activities_bp.route('/recent', methods=['GET'])
@jwt_required()
def get_recent_activities():
    """获取最近的活動记录"""
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        recent_date = now_china() - timedelta(days=7)
        activities = Activity.query.filter(
            Activity.created_at >= recent_date
        ).order_by(Activity.created_at.desc()).limit(50).all()

        result = []
        for activity in activities:
            activity_dict = {
                'id': activity.id,
                'action': activity.action,
                'description': activity.description,
                'target_type': activity.target_type,
                'target_id': activity.target_id,
                'created_at': activity.created_at.isoformat() if activity.created_at else None,
                'field_changes': json.loads(activity.field_changes) if activity.field_changes else [],
                'resource_name': resolve_resource_name(activity),
            }

            performer = User.query.get(activity.performed_by)
            if performer:
                activity_dict['user_name'] = performer.username

            result.append(activity_dict)

        return jsonify(result), 200

# 获取特定资源的活动记录
@activities_bp.route('/<string:resource_type>/<int:resource_id>', methods=['GET'])
@jwt_required()
def get_activities_by_resource(resource_type, resource_id):
    """获取特定资源的活动记录"""
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        activities = Activity.query.filter(
            Activity.target_type == resource_type,
            Activity.target_id == resource_id
        ).order_by(Activity.created_at.desc()).all()

        result = []
        for activity in activities:
            activity_dict = {
                'id': activity.id,
                'action': activity.action,
                'description': activity.description,
                'target_type': activity.target_type,
                'target_id': activity.target_id,
                'user_id': activity.performed_by,
                'created_at': activity.created_at.isoformat() if activity.created_at else None,
                'field_changes': json.loads(activity.field_changes) if activity.field_changes else [],
                'resource_name': resolve_resource_name(activity),
            }

            performer = User.query.get(activity.performed_by)
            if performer:
                activity_dict['user_name'] = performer.username

            result.append(activity_dict)

        return jsonify(result), 200

# 获取单个活动记录详情
@activities_bp.route('/<int:activity_id>', methods=['GET'])
@jwt_required()
def get_activity(activity_id):
    """获取单个活动记录详情"""
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        activity = Activity.query.get(activity_id)
        if not activity:
            return jsonify({'error': '活动记录不存在'}), 404

        activity_dict = {
            'id': activity.id,
            'action': activity.action,
            'description': activity.description,
            'target_type': activity.target_type,
            'target_id': activity.target_id,
            'user_id': activity.performed_by,
            'created_at': activity.created_at.isoformat() if activity.created_at else None,
            'field_changes': json.loads(activity.field_changes) if activity.field_changes else [],
            'resource_name': resolve_resource_name(activity),
        }

        performer = User.query.get(activity.performed_by)
        if performer:
            activity_dict['user_name'] = performer.username
            activity_dict['user_role'] = performer.role

        return jsonify(activity_dict), 200

# 创建活动记录
@activities_bp.route('/', methods=['POST'])
@jwt_required()
def create_activity():
    """创建活动记录"""
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        current_user_id = get_jwt_identity()
        data = request.get_json()

        activity = Activity(
            action=data.get('action'),
            description=data.get('description'),
            performed_by=int(current_user_id),
            target_type=data.get('target_type'),
            target_id=data.get('target_id')
        )

        db.session.add(activity)
        db.session.commit()

        return jsonify({
            'message': '活动记录创建成功',
            'activity': {
                'id': activity.id,
                'action': activity.action,
                'description': activity.description
            }
        }), 201

# 删除活动记录
@activities_bp.route('/<int:activity_id>', methods=['DELETE'])
@jwt_required()
def delete_activity(activity_id):
    """删除活动记录（仅管理员）"""
    User, Activity, Bug, Project, WorkLog, app, db = get_models()

    with app.app_context():
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))

        if current_user.role != 'admin':
            return jsonify({'error': '无权限删除活动记录'}), 403

        activity = Activity.query.get(activity_id)
        if not activity:
            return jsonify({'error': '活动记录不存在'}), 404

        db.session.delete(activity)
        db.session.commit()

        return jsonify({'message': '活动记录删除成功'}), 200

"""
测试管理API - 测试集与测试用例管理
"""
from flask import Blueprint, request, jsonify
from datetime import datetime
import json
from flask_jwt_extended import jwt_required, get_jwt_identity
import logging

logger = logging.getLogger(__name__)

test_management_bp = Blueprint('test_management', __name__, url_prefix='/test-management')

def get_db_and_models():
    from enhanced_app import db
    from enhanced_app import TestSuite, TestCase, TestStep, TestExecution, TestResult, TestCaseRequirementLink
    from enhanced_app import User, Project, RequirementItem, Bug
    from enhanced_app import create_audit_log
    return db, TestSuite, TestCase, TestStep, TestExecution, TestResult, TestCaseRequirementLink, User, Project, RequirementItem, Bug, create_audit_log


# ==================== 测试集管理 ====================

@test_management_bp.route('/suites/<int:project_id>', methods=['GET'])
@jwt_required()
def get_test_suites(project_id):
    """获取项目下的所有测试集"""
    try:
        db, TestSuite, TestCase, _, _, _, _, User, _, _, _, _ = get_db_and_models()
        include_cases = request.args.get('include_cases', 'false').lower() == 'true'
        
        suites = TestSuite.query.filter_by(project_id=project_id).all()
        result = [s.to_dict(include_children=True, include_cases=include_cases) for s in suites]
        
        return jsonify(result)
    except Exception as e:
        logger.error(f'获取测试集列表失败: {str(e)}')
        return jsonify({'error': '获取测试集列表失败'}), 500


@test_management_bp.route('/suites/<int:suite_id>', methods=['GET'])
@jwt_required()
def get_test_suite_by_id(suite_id):
    """根据ID获取测试集详情"""
    try:
        db, TestSuite, _, _, _, _, _, _, _, _, _, _ = get_db_and_models()
        include_cases = request.args.get('include_cases', 'true').lower() == 'true'
        
        suite = TestSuite.query.get(suite_id)
        if not suite:
            return jsonify({'error': '测试集不存在'}), 404
        
        return jsonify(suite.to_dict(include_children=True, include_cases=include_cases))
    except Exception as e:
        logger.error(f'获取测试集详情失败: {str(e)}')
        return jsonify({'error': '获取测试集详情失败'}), 500


@test_management_bp.route('/suites', methods=['POST'])
@jwt_required()
def create_test_suite():
    """创建测试集"""
    try:
        db, TestSuite, _, _, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        data = request.get_json()
        if not data.get('name'):
            return jsonify({'error': '测试集名称不能为空'}), 400
        
        suite = TestSuite(
            project_id=data.get('project_id'),
            parent_id=data.get('parent_id'),
            name=data.get('name'),
            description=data.get('description', ''),
            type=data.get('type', 'functional'),
            status=data.get('status', 'designing'),
            priority=data.get('priority', 2),
            owner_id=data.get('owner_id'),
            expected_duration=data.get('expected_duration', 0),
            created_by=current_user_id
        )
        
        db.session.add(suite)
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='create',
            resource_type='test_suite',
            resource_id=suite.id,
            details=f'创建测试集: {suite.name}'
        )
        
        return jsonify(suite.to_dict(include_children=True)), 201
    except Exception as e:
        logger.error(f'创建测试集失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '创建测试集失败'}), 500


@test_management_bp.route('/suites/<int:suite_id>', methods=['PUT'])
@jwt_required()
def update_test_suite(suite_id):
    """更新测试集"""
    try:
        db, TestSuite, _, _, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        suite = TestSuite.query.get(suite_id)
        if not suite:
            return jsonify({'error': '测试集不存在'}), 404
        
        data = request.get_json()
        
        # 更新字段
        if 'name' in data:
            suite.name = data['name']
        if 'description' in data:
            suite.description = data['description']
        if 'type' in data:
            suite.type = data['type']
        if 'status' in data:
            suite.status = data['status']
        if 'priority' in data:
            suite.priority = data['priority']
        if 'owner_id' in data:
            suite.owner_id = data['owner_id']
        if 'expected_duration' in data:
            suite.expected_duration = data['expected_duration']
        if 'parent_id' in data:
            suite.parent_id = data['parent_id']
        
        suite.version += 1
        suite.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='update',
            resource_type='test_suite',
            resource_id=suite.id,
            details=f'更新测试集: {suite.name}'
        )
        
        return jsonify(suite.to_dict(include_children=True))
    except Exception as e:
        logger.error(f'更新测试集失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '更新测试集失败'}), 500


@test_management_bp.route('/suites/<int:suite_id>', methods=['DELETE'])
@jwt_required()
def delete_test_suite(suite_id):
    """删除测试集"""
    try:
        db, TestSuite, _, _, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        suite = TestSuite.query.get(suite_id)
        if not suite:
            return jsonify({'error': '测试集不存在'}), 404
        
        suite_name = suite.name
        db.session.delete(suite)
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='delete',
            resource_type='test_suite',
            resource_id=suite_id,
            details=f'删除测试集: {suite_name}'
        )
        
        return jsonify({'message': '删除成功'})
    except Exception as e:
        logger.error(f'删除测试集失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '删除测试集失败'}), 500


@test_management_bp.route('/suites/batch', methods=['POST'])
@jwt_required()
def batch_operate_suites():
    """批量操作测试集"""
    try:
        db, TestSuite, _, _, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        data = request.get_json()
        operation = data.get('operation')
        suite_ids = data.get('suite_ids', [])
        
        if operation == 'copy':
            # 复制测试集
            for suite_id in suite_ids:
                original = TestSuite.query.get(suite_id)
                if original:
                    new_suite = TestSuite(
                        project_id=data.get('target_project_id', original.project_id),
                        parent_id=data.get('target_suite_id', original.parent_id),
                        name=data.get('new_name', f'{original.name} (副本)'),
                        description=original.description,
                        type=original.type,
                        status='designing',
                        priority=original.priority,
                        owner_id=original.owner_id,
                        expected_duration=original.expected_duration,
                        created_by=current_user_id
                    )
                    db.session.add(new_suite)
            
            db.session.commit()
            return jsonify({'message': '复制成功'})
        
        return jsonify({'error': '不支持的操作'}), 400
    except Exception as e:
        logger.error(f'批量操作测试集失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '批量操作失败'}), 500


@test_management_bp.route('/suites/<int:suite_id>/version-history', methods=['GET'])
@jwt_required()
def get_suite_version_history(suite_id):
    """获取测试集版本历史（简化版）"""
    try:
        db, TestSuite, TestCase, _, _, _, _, _, _, _, _, _ = get_db_and_models()
        
        suite = TestSuite.query.get(suite_id)
        if not suite:
            return jsonify({'error': '测试集不存在'}), 404
        
        cases = TestCase.query.filter_by(suite_id=suite_id).all()
        
        case_summary = {
            'total': len(cases),
            'approved': len([c for c in cases if c.status == 'approved']),
            'reviewed': len([c for c in cases if c.status == 'reviewed']),
            'pending_review': len([c for c in cases if c.status == 'pending_review'])
        }
        
        return jsonify({
            'suite': suite.to_dict(),
            'case_summary': case_summary,
            'cases': [c.to_dict() for c in cases]
        })
    except Exception as e:
        logger.error(f'获取测试集版本历史失败: {str(e)}')
        return jsonify({'error': '获取版本历史失败'}), 500


# ==================== 测试用例管理 ====================

@test_management_bp.route('/cases/by-suite/<int:suite_id>', methods=['GET'])
@jwt_required()
def get_cases_by_suite(suite_id):
    """获取测试集下的所有用例"""
    try:
        db, _, TestCase, _, _, _, _, _, _, _, _, _ = get_db_and_models()
        
        cases = TestCase.query.filter_by(suite_id=suite_id).all()
        return jsonify([c.to_dict(include_steps=True, include_links=True) for c in cases])
    except Exception as e:
        logger.error(f'获取用例列表失败: {str(e)}')
        return jsonify({'error': '获取用例列表失败'}), 500


@test_management_bp.route('/cases/<int:case_id>', methods=['GET'])
@jwt_required()
def get_case_by_id(case_id):
    """根据ID获取测试用例"""
    try:
        db, _, TestCase, _, _, _, _, _, _, _, _, _ = get_db_and_models()
        
        case = TestCase.query.get(case_id)
        if not case:
            return jsonify({'error': '用例不存在'}), 404
        
        return jsonify(case.to_dict(include_steps=True, include_links=True))
    except Exception as e:
        logger.error(f'获取用例详情失败: {str(e)}')
        return jsonify({'error': '获取用例详情失败'}), 500


@test_management_bp.route('/cases', methods=['POST'])
@jwt_required()
def create_test_case():
    """创建测试用例"""
    try:
        db, _, TestCase, TestStep, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        data = request.get_json()
        if not data.get('title') or not data.get('suite_id'):
            return jsonify({'error': '用例标题和所属测试集不能为空'}), 400
        
        # 生成用例标识符
        suite_id = data.get('suite_id')
        count = TestCase.query.filter_by(suite_id=suite_id).count()
        identifier = f'TC-{suite_id}-{count + 1:03d}'
        
        case = TestCase(
            suite_id=suite_id,
            identifier=identifier,
            title=data.get('title'),
            description=data.get('description', ''),
            priority=data.get('priority', 2),
            type=data.get('type', 'functional'),
            status=data.get('status', 'designing'),
            precondition=data.get('precondition', ''),
            test_data=data.get('test_data', ''),
            environment=data.get('environment', ''),
            is_automated=data.get('is_automated', False),
            automation_script=data.get('automation_script'),
            tags=data.get('tags'),
            estimated_duration=data.get('estimated_duration', 0),
            designer_id=data.get('designer_id', current_user_id),
            created_by=current_user_id
        )
        
        db.session.add(case)
        db.session.flush()
        
        # 处理步骤
        if 'steps' in data:
            for idx, step_data in enumerate(data['steps']):
                step = TestStep(
                    case_id=case.id,
                    step_number=step_data.get('step_number', idx + 1),
                    action=step_data.get('action', ''),
                    expected_result=step_data.get('expected_result', '')
                )
                db.session.add(step)
        
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='create',
            resource_type='test_case',
            resource_id=case.id,
            details=f'创建测试用例: {case.title}'
        )
        
        return jsonify(case.to_dict(include_steps=True)), 201
    except Exception as e:
        logger.error(f'创建测试用例失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '创建测试用例失败'}), 500


@test_management_bp.route('/cases/<int:case_id>', methods=['PUT'])
@jwt_required()
def update_test_case(case_id):
    """更新测试用例"""
    try:
        db, _, TestCase, TestStep, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        case = TestCase.query.get(case_id)
        if not case:
            return jsonify({'error': '用例不存在'}), 404
        
        data = request.get_json()
        
        # 更新基本信息
        if 'title' in data:
            case.title = data['title']
        if 'description' in data:
            case.description = data['description']
        if 'priority' in data:
            case.priority = data['priority']
        if 'type' in data:
            case.type = data['type']
        if 'status' in data:
            case.status = data['status']
        if 'precondition' in data:
            case.precondition = data['precondition']
        if 'test_data' in data:
            case.test_data = data['test_data']
        if 'environment' in data:
            case.environment = data['environment']
        if 'is_automated' in data:
            case.is_automated = data['is_automated']
        if 'automation_script' in data:
            case.automation_script = data['automation_script']
        if 'tags' in data:
            case.tags = data['tags']
        if 'estimated_duration' in data:
            case.estimated_duration = data['estimated_duration']
        if 'designer_id' in data:
            case.designer_id = data['designer_id']
        if 'reviewer_id' in data:
            case.reviewer_id = data['reviewer_id']
        if 'approved_by' in data:
            case.approved_by = data['approved_by']
        
        # 更新步骤
        if 'steps' in data:
            # 删除旧步骤
            TestStep.query.filter_by(case_id=case_id).delete()
            # 添加新步骤
            for step_data in data['steps']:
                step = TestStep(
                    case_id=case.id,
                    step_number=step_data.get('step_number'),
                    action=step_data.get('action', ''),
                    expected_result=step_data.get('expected_result', '')
                )
                db.session.add(step)
        
        case.version += 1
        case.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='update',
            resource_type='test_case',
            resource_id=case.id,
            details=f'更新测试用例: {case.title}'
        )
        
        return jsonify(case.to_dict(include_steps=True))
    except Exception as e:
        logger.error(f'更新测试用例失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '更新测试用例失败'}), 500


@test_management_bp.route('/cases/<int:case_id>', methods=['DELETE'])
@jwt_required()
def delete_test_case(case_id):
    """删除测试用例"""
    try:
        db, _, TestCase, _, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        case = TestCase.query.get(case_id)
        if not case:
            return jsonify({'error': '用例不存在'}), 404
        
        case_title = case.title
        db.session.delete(case)
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='delete',
            resource_type='test_case',
            resource_id=case_id,
            details=f'删除测试用例: {case_title}'
        )
        
        return jsonify({'message': '删除成功'})
    except Exception as e:
        logger.error(f'删除测试用例失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '删除测试用例失败'}), 500


@test_management_bp.route('/cases/batch', methods=['POST'])
@jwt_required()
def batch_operate_cases():
    """批量操作用例"""
    try:
        db, _, TestCase, _, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        data = request.get_json()
        operation = data.get('operation')
        case_ids = data.get('case_ids', [])
        
        if operation == 'move':
            target_suite_id = data.get('target_suite_id')
            for case_id in case_ids:
                case = TestCase.query.get(case_id)
                if case:
                    case.suite_id = target_suite_id
            db.session.commit()
            return jsonify({'message': '移动成功'})
        
        return jsonify({'error': '不支持的操作'}), 400
    except Exception as e:
        logger.error(f'批量操作用例失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '批量操作失败'}), 500


@test_management_bp.route('/cases/<int:case_id>/history', methods=['GET'])
@jwt_required()
def get_case_history(case_id):
    """获取用例历史（简化版）"""
    try:
        db, _, TestCase, _, _, _, _, _, _, _, _, _ = get_db_and_models()
        
        case = TestCase.query.get(case_id)
        if not case:
            return jsonify({'error': '用例不存在'}), 404
        
        return jsonify([{
            'version': case.version,
            'title': case.title,
            'status': case.status,
            'updated_at': case.updated_at.isoformat() if case.updated_at else None
        }])
    except Exception as e:
        logger.error(f'获取用例历史失败: {str(e)}')
        return jsonify({'error': '获取用例历史失败'}), 500


@test_management_bp.route('/cases/<int:case_id>/review', methods=['POST'])
@jwt_required()
def submit_case_review(case_id):
    """提交用例评审"""
    try:
        db, _, TestCase, _, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        case = TestCase.query.get(case_id)
        if not case:
            return jsonify({'error': '用例不存在'}), 404
        
        data = request.get_json()
        review_status = data.get('status')
        
        if review_status == 'reviewed':
            case.status = 'reviewed'
            case.reviewer_id = current_user_id
        elif review_status == 'approved':
            case.status = 'approved'
            case.approved_by = current_user_id
        
        case.version += 1
        case.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify(case.to_dict())
    except Exception as e:
        logger.error(f'提交评审失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '提交评审失败'}), 500


@test_management_bp.route('/cases/<int:case_id>/copy', methods=['POST'])
@jwt_required()
def copy_case(case_id):
    """复制用例"""
    try:
        db, _, TestCase, TestStep, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        original = TestCase.query.get(case_id)
        if not original:
            return jsonify({'error': '用例不存在'}), 404
        
        data = request.get_json()
        target_suite_id = data.get('suite_id', original.suite_id)
        
        count = TestCase.query.filter_by(suite_id=target_suite_id).count()
        identifier = f'TC-{target_suite_id}-{count + 1:03d}'
        
        new_case = TestCase(
            suite_id=target_suite_id,
            identifier=identifier,
            title=data.get('title', f'{original.title} (副本)'),
            description=original.description,
            priority=original.priority,
            type=original.type,
            status='designing',
            precondition=original.precondition,
            test_data=original.test_data,
            environment=original.environment,
            is_automated=original.is_automated,
            automation_script=original.automation_script,
            tags=original.tags,
            estimated_duration=original.estimated_duration,
            designer_id=current_user_id,
            created_by=current_user_id
        )
        
        db.session.add(new_case)
        db.session.flush()
        
        # 复制步骤
        for step in original.steps:
            new_step = TestStep(
                case_id=new_case.id,
                step_number=step.step_number,
                action=step.action,
                expected_result=step.expected_result
            )
            db.session.add(new_step)
        
        db.session.commit()
        
        return jsonify(new_case.to_dict(include_steps=True)), 201
    except Exception as e:
        logger.error(f'复制用例失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '复制用例失败'}), 500


# ==================== 测试执行管理 ====================

@test_management_bp.route('/executions/by-project/<int:project_id>', methods=['GET'])
@jwt_required()
def get_executions_by_project(project_id):
    """获取项目的执行记录"""
    try:
        db, _, _, _, TestExecution, _, _, _, _, _, _, _ = get_db_and_models()
        
        executions = TestExecution.query.filter_by(project_id=project_id).order_by(TestExecution.created_at.desc()).all()
        return jsonify([e.to_dict(include_results=True) for e in executions])
    except Exception as e:
        logger.error(f'获取执行记录失败: {str(e)}')
        return jsonify({'error': '获取执行记录失败'}), 500


@test_management_bp.route('/executions/<int:execution_id>', methods=['GET'])
@jwt_required()
def get_execution_by_id(execution_id):
    """获取执行记录详情"""
    try:
        db, _, _, _, TestExecution, _, _, _, _, _, _, _ = get_db_and_models()
        
        execution = TestExecution.query.get(execution_id)
        if not execution:
            return jsonify({'error': '执行记录不存在'}), 404
        
        return jsonify(execution.to_dict(include_results=True))
    except Exception as e:
        logger.error(f'获取执行详情失败: {str(e)}')
        return jsonify({'error': '获取执行详情失败'}), 500


@test_management_bp.route('/executions', methods=['POST'])
@jwt_required()
def create_execution():
    """创建执行记录"""
    try:
        db, _, _, _, TestExecution, TestResult, TestCase, _, User, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        data = request.get_json()
        if not data.get('name') or not data.get('project_id'):
            return jsonify({'error': '执行名称和项目ID不能为空'}), 400
        
        execution = TestExecution(
            project_id=data.get('project_id'),
            suite_id=data.get('suite_id'),
            name=data.get('name'),
            executor_id=current_user_id,
            environment=data.get('environment'),
            test_version=data.get('test_version'),
            build_number=data.get('build_number'),
            notes=data.get('notes')
        )
        
        db.session.add(execution)
        db.session.flush()
        
        # 自动为测试集的用例创建结果记录
        if data.get('suite_id'):
            cases = TestCase.query.filter_by(suite_id=data.get('suite_id')).all()
            for case in cases:
                result = TestResult(
                    execution_id=execution.id,
                    case_id=case.id,
                    result='not_executed',
                    executor_id=current_user_id
                )
                db.session.add(result)
        
        db.session.commit()
        
        # 记录审计日志
        create_audit_log(
            user_id=current_user_id,
            action='create',
            resource_type='test_execution',
            resource_id=execution.id,
            details=f'创建测试执行: {execution.name}'
        )
        
        return jsonify(execution.to_dict(include_results=True)), 201
    except Exception as e:
        logger.error(f'创建执行记录失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '创建执行记录失败'}), 500


@test_management_bp.route('/executions/<int:execution_id>', methods=['PUT'])
@jwt_required()
def update_execution(execution_id):
    """更新执行记录"""
    try:
        db, _, _, _, TestExecution, _, _, _, User, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        execution = TestExecution.query.get(execution_id)
        if not execution:
            return jsonify({'error': '执行记录不存在'}), 404
        
        data = request.get_json()
        
        if 'name' in data:
            execution.name = data['name']
        if 'status' in data:
            execution.status = data['status']
            if data['status'] == 'completed':
                execution.completed_at = datetime.utcnow()
        if 'environment' in data:
            execution.environment = data['environment']
        if 'test_version' in data:
            execution.test_version = data['test_version']
        if 'build_number' in data:
            execution.build_number = data['build_number']
        if 'notes' in data:
            execution.notes = data['notes']
        
        execution.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify(execution.to_dict(include_results=True))
    except Exception as e:
        logger.error(f'更新执行记录失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '更新执行记录失败'}), 500


@test_management_bp.route('/executions/<int:execution_id>', methods=['DELETE'])
@jwt_required()
def delete_execution(execution_id):
    """删除执行记录"""
    try:
        db, _, _, _, TestExecution, _, _, _, User, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        execution = TestExecution.query.get(execution_id)
        if not execution:
            return jsonify({'error': '执行记录不存在'}), 404
        
        execution_name = execution.name
        db.session.delete(execution)
        db.session.commit()
        
        return jsonify({'message': '删除成功'})
    except Exception as e:
        logger.error(f'删除执行记录失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '删除执行记录失败'}), 500


@test_management_bp.route('/executions/<int:execution_id>/results', methods=['POST'])
@jwt_required()
def submit_test_result(execution_id):
    """提交测试结果"""
    try:
        db, _, _, _, TestExecution, TestResult, _, _, User, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        execution = TestExecution.query.get(execution_id)
        if not execution:
            return jsonify({'error': '执行记录不存在'}), 404
        
        data = request.get_json()
        
        result = TestResult.query.filter_by(
            execution_id=execution_id,
            case_id=data.get('case_id')
        ).first()
        
        if not result:
            result = TestResult(
                execution_id=execution_id,
                case_id=data.get('case_id'),
                executor_id=current_user_id
            )
            db.session.add(result)
        
        result.result = data.get('result')
        result.actual_result = data.get('actual_result')
        result.defect_id = data.get('defect_id')
        result.notes = data.get('notes')
        result.executed_at = datetime.utcnow()
        result.duration = data.get('duration', 0)
        
        db.session.commit()
        
        return jsonify(result.to_dict())
    except Exception as e:
        logger.error(f'提交测试结果失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '提交测试结果失败'}), 500


@test_management_bp.route('/results/by-case/<int:case_id>', methods=['GET'])
@jwt_required()
def get_results_by_case(case_id):
    """获取用例的执行结果"""
    try:
        db, _, _, _, _, TestResult, _, _, _, _, _, _ = get_db_and_models()
        
        results = TestResult.query.filter_by(case_id=case_id).order_by(TestResult.created_at.desc()).all()
        return jsonify([r.to_dict() for r in results])
    except Exception as e:
        logger.error(f'获取用例执行结果失败: {str(e)}')
        return jsonify({'error': '获取用例执行结果失败'}), 500


# ==================== 需求关联管理 ====================

@test_management_bp.route('/links', methods=['POST'])
@jwt_required()
def create_requirement_link():
    """创建用例与需求关联"""
    try:
        db, _, _, _, _, _, TestCaseRequirementLink, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        data = request.get_json()
        
        link = TestCaseRequirementLink(
            test_case_id=data.get('test_case_id'),
            requirement_id=data.get('requirement_id'),
            link_type=data.get('link_type', 'tests'),
            created_by=current_user_id
        )
        
        db.session.add(link)
        db.session.commit()
        
        return jsonify(link.to_dict()), 201
    except Exception as e:
        logger.error(f'创建关联失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '创建关联失败'}), 500


@test_management_bp.route('/links/<int:link_id>', methods=['DELETE'])
@jwt_required()
def delete_requirement_link(link_id):
    """删除用例与需求关联"""
    try:
        db, _, _, _, _, _, TestCaseRequirementLink, _, _, _, _, _ = get_db_and_models()
        
        link = TestCaseRequirementLink.query.get(link_id)
        if not link:
            return jsonify({'error': '关联不存在'}), 404
        
        db.session.delete(link)
        db.session.commit()
        
        return jsonify({'message': '删除成功'})
    except Exception as e:
        logger.error(f'删除关联失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': '删除关联失败'}), 500


@test_management_bp.route('/links/by-case/<int:case_id>', methods=['GET'])
@jwt_required()
def get_links_by_case(case_id):
    """获取用例的关联需求"""
    try:
        db, _, _, _, _, _, TestCaseRequirementLink, _, _, _, _, _ = get_db_and_models()
        
        links = TestCaseRequirementLink.query.filter_by(test_case_id=case_id).all()
        return jsonify([l.to_dict() for l in links])
    except Exception as e:
        logger.error(f'获取用例关联失败: {str(e)}')
        return jsonify({'error': '获取用例关联失败'}), 500


@test_management_bp.route('/links/by-requirement/<int:requirement_id>', methods=['GET'])
@jwt_required()
def get_links_by_requirement(requirement_id):
    """获取需求的关联用例"""
    try:
        db, _, _, _, _, _, TestCaseRequirementLink, _, _, _, _, _ = get_db_and_models()
        
        links = TestCaseRequirementLink.query.filter_by(requirement_id=requirement_id).all()
        return jsonify([l.to_dict() for l in links])
    except Exception as e:
        logger.error(f'获取需求关联失败: {str(e)}')
        return jsonify({'error': '获取需求关联失败'}), 500


# ==================== 统计与报表 ====================

@test_management_bp.route('/statistics/project/<int:project_id>', methods=['GET'])
@jwt_required()
def get_project_statistics(project_id):
    """获取项目测试统计"""
    try:
        db, TestSuite, TestCase, _, TestExecution, _, _, _, Project, _, _, _ = get_db_and_models()
        
        suites = TestSuite.query.filter_by(project_id=project_id).all()
        cases = TestCase.query.join(TestSuite).filter(TestSuite.project_id == project_id).all()
        executions = TestExecution.query.filter_by(project_id=project_id).order_by(TestExecution.created_at.desc()).limit(10).all()
        
        total_suites = len(suites)
        total_cases = len(cases)
        
        status_counts = {}
        for case in cases:
            status = case.status or 'unknown'
            status_counts[status] = status_counts.get(status, 0) + 1
        
        return jsonify({
            'total_suites': total_suites,
            'total_cases': total_cases,
            'status_counts': status_counts,
            'latest_executions': [e.to_dict() for e in executions]
        })
    except Exception as e:
        logger.error(f'获取项目统计失败: {str(e)}')
        return jsonify({'error': '获取项目统计失败'}), 500


@test_management_bp.route('/reports/execution/<int:execution_id>', methods=['GET'])
@jwt_required()
def get_execution_report(execution_id):
    """获取执行报告"""
    try:
        db, _, _, _, TestExecution, TestResult, _, _, _, _, _, _ = get_db_and_models()
        
        execution = TestExecution.query.get(execution_id)
        if not execution:
            return jsonify({'error': '执行记录不存在'}), 404
        
        results = TestResult.query.filter_by(execution_id=execution_id).all()
        
        total = len(results)
        passed = sum(1 for r in results if r.result == 'passed')
        failed = sum(1 for r in results if r.result == 'failed')
        blocked = sum(1 for r in results if r.result == 'blocked')
        skipped = sum(1 for r in results if r.result == 'skipped')
        
        pass_rate = (passed / total * 100) if total > 0 else 0
        
        return jsonify({
            'execution': execution.to_dict(),
            'summary': {
                'total': total,
                'passed': passed,
                'failed': failed,
                'blocked': blocked,
                'skipped': skipped,
                'pass_rate': round(pass_rate, 2)
            },
            'failed_cases': [r.to_dict() for r in results if r.result == 'failed']
        })
    except Exception as e:
        logger.error(f'获取执行报告失败: {str(e)}')
        return jsonify({'error': '获取执行报告失败'}), 500


@test_management_bp.route('/reports/project/<int:project_id>', methods=['GET'])
@jwt_required()
def get_project_test_report(project_id):
    """获取项目测试报告"""
    try:
        db, TestSuite, TestCase, _, TestExecution, TestResult, _, _, Project, _, _, _ = get_db_and_models()
        
        suites = TestSuite.query.filter_by(project_id=project_id).all()
        cases = TestCase.query.join(TestSuite).filter(TestSuite.project_id == project_id).all()
        executions = TestExecution.query.filter_by(project_id=project_id).all()
        
        total_suites = len(suites)
        total_cases = len(cases)
        
        # 用例状态统计
        case_status = {}
        for case in cases:
            status = case.status or 'unknown'
            case_status[status] = case_status.get(status, 0) + 1
        
        # 执行统计
        exec_stats = {}
        for execution in executions:
            for result in execution.results:
                r = result.result
                exec_stats[r] = exec_stats.get(r, 0) + 1
        
        return jsonify({
            'total_suites': total_suites,
            'total_cases': total_cases,
            'case_status': case_status,
            'execution_stats': exec_stats,
            'total_executions': len(executions)
        })
    except Exception as e:
        logger.error(f'获取项目测试报告失败: {str(e)}')
        return jsonify({'error': '获取项目测试报告失败'}), 500


@test_management_bp.route('/dashboard/<int:project_id>', methods=['GET'])
@jwt_required()
def get_test_dashboard(project_id):
    """获取测试仪表盘"""
    try:
        db, TestSuite, TestCase, _, TestExecution, TestResult, _, _, Project, _, _, _ = get_db_and_models()
        
        suites = TestSuite.query.filter_by(project_id=project_id).all()
        cases = TestCase.query.join(TestSuite).filter(TestSuite.project_id == project_id).all()
        executions = TestExecution.query.filter_by(project_id=project_id).order_by(TestExecution.created_at.desc()).limit(5).all()
        
        total_suites = len(suites)
        total_cases = len(cases)
        
        # 计算自动化覆盖率
        automated_cases = len([c for c in cases if c.is_automated])
        automation_rate = (automated_cases / total_cases * 100) if total_cases > 0 else 0
        
        # 计算已评审用例数
        reviewed_cases = len([c for c in cases if c.status in ['reviewed', 'approved']])
        
        return jsonify({
            'total_suites': total_suites,
            'total_cases': total_cases,
            'automated_cases': automated_cases,
            'automation_rate': round(automation_rate, 2),
            'reviewed_cases': reviewed_cases,
            'latest_executions': [e.to_dict() for e in executions]
        })
    except Exception as e:
        logger.error(f'获取测试仪表盘失败: {str(e)}')
        return jsonify({'error': '获取测试仪表盘失败'}), 500


# ==================== 导入导出功能 ====================

@test_management_bp.route('/cases/export/<int:suite_id>', methods=['GET'])
@jwt_required()
def export_test_cases(suite_id):
    """导出测试用例为Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        db, _, TestCase, _, _, _, _, _, _, _, _, _ = get_db_and_models()
        
        cases = TestCase.query.filter_by(suite_id=suite_id).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        headers = ['用例标识', '用例标题', '优先级', '类型', '状态', '前提条件', '测试数据', 
                   '环境要求', '是否自动化', '步骤数', '创建时间', '更新时间']
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        priority_map = {0: 'P0', 1: 'P1', 2: 'P2', 3: 'P3'}
        status_map = {
            'designing': '设计中',
            'pending_review': '待评审',
            'reviewed': '已评审',
            'approved': '已批准',
            'deprecated': '已废弃'
        }
        
        for row, case in enumerate(cases, 2):
            ws.cell(row=row, column=1, value=case.identifier)
            ws.cell(row=row, column=2, value=case.title)
            ws.cell(row=row, column=3, value=priority_map.get(case.priority, 'P2'))
            ws.cell(row=row, column=4, value=case.type)
            ws.cell(row=row, column=5, value=status_map.get(case.status, case.status))
            ws.cell(row=row, column=6, value=case.precondition or '')
            ws.cell(row=row, column=7, value=case.test_data or '')
            ws.cell(row=row, column=8, value=case.environment or '')
            ws.cell(row=row, column=9, value='是' if case.is_automated else '否')
            ws.cell(row=row, column=10, value=len(case.steps) if case.steps else 0)
            ws.cell(row=row, column=11, value=case.created_at.strftime('%Y-%m-%d %H:%M') if case.created_at else '')
            ws.cell(row=row, column=12, value=case.updated_at.strftime('%Y-%m-%d %H:%M') if case.updated_at else '')
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'测试用例_{suite_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f'导出测试用例失败: {str(e)}')
        return jsonify({'error': '导出测试用例失败'}), 500


@test_management_bp.route('/cases/import', methods=['POST'])
@jwt_required()
def import_test_cases():
    """导入测试用例"""
    try:
        from openpyxl import load_workbook
        
        db, _, TestCase, TestStep, _, _, _, User, _, _, _, create_audit_log = get_db_and_models()
        current_user_id = get_jwt_identity()
        
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
        
        file = request.files['file']
        suite_id = request.form.get('suite_id', type=int)
        
        if not suite_id:
            return jsonify({'error': '缺少测试集ID'}), 400
        
        wb = load_workbook(file)
        ws = wb.active
        
        priority_reverse_map = {'P0': 0, 'P1': 1, 'P2': 2, 'P3': 3}
        status_reverse_map = {
            '设计中': 'designing',
            '待评审': 'pending_review',
            '已评审': 'reviewed',
            '已批准': 'approved',
            '已废弃': 'deprecated'
        }
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row[1]:
                    continue
                
                count = TestCase.query.filter_by(suite_id=suite_id).count()
                identifier = f'TC-{suite_id}-{count + 1:03d}'
                
                priority_str = str(row[2]) if row[2] else 'P2'
                priority = priority_reverse_map.get(priority_str, 2)
                
                status_str = str(row[4]) if row[4] else '设计中'
                status = status_reverse_map.get(status_str, 'designing')
                
                case = TestCase(
                    suite_id=suite_id,
                    identifier=identifier,
                    title=str(row[1]),
                    priority=priority,
                    type=str(row[3]) if row[3] else 'functional',
                    status=status,
                    precondition=str(row[5]) if row[5] else '',
                    test_data=str(row[6]) if row[6] else '',
                    environment=str(row[7]) if row[7] else '',
                    is_automated=str(row[8]) == '是' if row[8] else False,
                    designer_id=current_user_id,
                    created_by=current_user_id
                )
                
                db.session.add(case)
                db.session.flush()
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f'第{row_idx}行导入失败: {str(e)}')
        
        db.session.commit()
        
        create_audit_log(
            user_id=current_user_id,
            action='import',
            resource_type='test_case',
            resource_id=suite_id,
            details=f'导入测试用例: 成功{imported_count}个'
        )
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {imported_count} 个用例',
            'imported_count': imported_count,
            'errors': errors
        })
    except Exception as e:
        logger.error(f'导入测试用例失败: {str(e)}')
        db.session.rollback()
        return jsonify({'error': f'导入失败: {str(e)}'}), 500


@test_management_bp.route('/reports/execution/<int:execution_id>/export', methods=['GET'])
@jwt_required()
def export_execution_report(execution_id):
    """导出执行报告为Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        db, _, _, _, TestExecution, TestResult, _, _, _, _, _, _ = get_db_and_models()
        
        execution = TestExecution.query.get(execution_id)
        if not execution:
            return jsonify({'error': '执行记录不存在'}), 404
        
        results = TestResult.query.filter_by(execution_id=execution_id).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "执行报告"
        
        ws.merge_cells('A1:H1')
        title_cell = ws.cell(row=1, column=1, value=f'测试执行报告 - {execution.name}')
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=3, column=1, value='执行名称:')
        ws.cell(row=3, column=2, value=execution.name)
        ws.cell(row=4, column=1, value='执行人:')
        ws.cell(row=4, column=2, value=execution.executor.username if execution.executor else '')
        ws.cell(row=5, column=1, value='开始时间:')
        ws.cell(row=5, column=2, value=execution.started_at.strftime('%Y-%m-%d %H:%M') if execution.started_at else '')
        ws.cell(row=6, column=1, value='完成时间:')
        ws.cell(row=6, column=2, value=execution.completed_at.strftime('%Y-%m-%d %H:%M') if execution.completed_at else '')
        
        total = len(results)
        passed = sum(1 for r in results if r.result == 'passed')
        failed = sum(1 for r in results if r.result == 'failed')
        blocked = sum(1 for r in results if r.result == 'blocked')
        pass_rate = (passed / total * 100) if total > 0 else 0
        
        ws.cell(row=8, column=1, value='总用例数:')
        ws.cell(row=8, column=2, value=total)
        ws.cell(row=8, column=3, value='通过数:')
        ws.cell(row=8, column=4, value=passed)
        ws.cell(row=8, column=5, value='失败数:')
        ws.cell(row=8, column=6, value=failed)
        ws.cell(row=8, column=7, value='阻塞数:')
        ws.cell(row=8, column=8, value=blocked)
        ws.cell(row=9, column=1, value='通过率:')
        ws.cell(row=9, column=2, value=f'{pass_rate:.2f}%')
        
        headers = ['用例标识', '用例标题', '执行结果', '执行人', '执行时间', '实际结果', '关联缺陷', '备注']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        result_map = {
            'passed': '通过',
            'failed': '失败',
            'blocked': '阻塞',
            'skipped': '跳过',
            'not_executed': '未执行'
        }
        
        for row, result in enumerate(results, 12):
            ws.cell(row=row, column=1, value=result.test_case.identifier if result.test_case else '')
            ws.cell(row=row, column=2, value=result.test_case.title if result.test_case else '')
            ws.cell(row=row, column=3, value=result_map.get(result.result, result.result))
            ws.cell(row=row, column=4, value=result.executor.username if result.executor else '')
            ws.cell(row=row, column=5, value=result.executed_at.strftime('%Y-%m-%d %H:%M') if result.executed_at else '')
            ws.cell(row=row, column=6, value=result.actual_result or '')
            ws.cell(row=row, column=7, value=result.defect.title if result.defect else '')
            ws.cell(row=row, column=8, value=result.notes or '')
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'执行报告_{execution.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f'导出执行报告失败: {str(e)}')
        return jsonify({'error': '导出执行报告失败'}), 500
